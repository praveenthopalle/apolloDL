# -*- coding: utf-8 -*-

from __future__ import unicode_literals
import torch

CUDA_LAUNCH_BLOCKING = "1"
import os

os.environ["CUDA_DEVICE_ORDER"] = "PCI_BUS_ID"
os.environ["CUDA_VISIBLE_DEVICES"] = "0,1,2,3"

import multiprocessing
import random
from threading import Thread

import botocore
from django.contrib import auth
from django.contrib.auth import authenticate
from django.core.files.uploadedfile import UploadedFile
from django.shortcuts import render
from django.template import RequestContext
from django.utils.datetime_safe import datetime
from django.views.decorators.csrf import csrf_exempt
import json
from nltk.corpus import stopwords
from nltk.stem import WordNetLemmatizer as wn
from django.http import HttpResponse, JsonResponse
import re
from openpyxl import load_workbook
from openpyxl.writer.excel import save_virtual_workbook
from apollo4.ComputeOptimalParameters import getOptimalParameterForMNB_alpha, getOptimalParameterForLR_alpha, \
    getOptimalParameterForSVM_alpha, getOptimalParameterForOVRMNB_alpha, getOptimalParameterForOVRLR_alpha, \
    getOptimalParameterForOVRSVM_alpha, getOptimalParametersForDeepLearning, getBestModelAndHyperParameters
from sklearn.feature_extraction.text import TfidfVectorizer, CountVectorizer, TfidfTransformer
from sklearn.naive_bayes import MultinomialNB
from sklearn.linear_model import SGDClassifier
from sklearn.calibration import CalibratedClassifierCV
from sklearn.multiclass import OneVsRestClassifier
from sklearn.cluster import MiniBatchKMeans
from sklearn.decomposition import LatentDirichletAllocation
from sklearn.model_selection import KFold
from sklearn import preprocessing
from sklearn import metrics
from sklearn.neighbors import KNeighborsClassifier
import pandas as pd
import numpy as np
import copy
from dateutil.relativedelta import relativedelta
import os, sys, signal
from os.path import splitext
import uuid
import apollo4.globals
import boto3
from elasticsearch import Elasticsearch
from apollo4.connection import MyConnection
import pickle
from sklearn.metrics.pairwise import cosine_similarity
from sklearn.decomposition import TruncatedSVD
from zipfile import ZipFile
import io
from django.contrib.auth.decorators import login_required
from django.conf import settings
from apollo4.secondviews import Process_All_Files
from apollo4.DeepLearningModel import DeepLearningModel
from collections import Counter

HOST_URLS = ["https://search-apollo4-5xxq4s5dugv4fenh4e2bbt3xmi.us-east-1.es.amazonaws.com"]
#  In production or deploying to aws uncomment this line and comment local server settings.
# if os.name == 'nt':
# #     # In local server uncomment the below code and comment above es_conn.
# #     HOST_URLS = ["https://search-apollolocal-yrsi6dysaadh7xyeotkeeoybqu.us-east-1.es.amazonaws.com"]
#     es_conn = Elasticsearch(HOST_URLS, connection_class=MyConnection,
#                             proxies={'https': 'http://root:Samsung1%21@105.128.219.200:8080'},timeout=30)
#     os.environ["HTTPS_PROXY"] = "https://root:Samsung1%21@105.128.219.200:8080"
#     os.environ["HTTP_PROXY"] = "http://root:Samsung1%21@105.128.219.200:8080"
# else:
es_conn = Elasticsearch(HOST_URLS, timeout=60)

print('osname', os.name)
TYPE_NAME_USER = '_doc'

AWS_STORAGE_BUCKET_NAME = 'apollo4'

response1 = None
response2 = None
response = None
response4 = None
responseTrain = None

stopwords = ["i", "me", "my", "myself", "we", "our", "ours", "ourselves", "you", "you're", "you've", "you'll", "you'd",
             "your", "yours", "yourself", "yourselves", "he", "him", "his", "himself", "she", "she's", "her", "hers",
             "herself", "it", "it's", "its", "itself", "they", "them", "their", "theirs", "themselves", "what", "which",
             "who", "whom", "this", "that", "that'll", "these", "those", "am", "is", "are", "was", "were", "be", "been",
             "being",
             "have", "has", "had", "having", "do", "does", "did", "doing", "a", "an", "the", "and", "but", "if", "or",
             "because", "as", "until", "while", "of", "at", "by", "for", "with", "about", "against", "between", "into",
             "through",
             "during", "before", "after", "above", "below", "to", "from", "up", "down", "in", "out", "on", "off",
             "over", "under", "again", "further", "then", "once", "here", "there", "when", "where", "why", "how", "all",
             "any", "both",
             "each", "few", "more", "most", "other", "some", "such", "no", "nor", "not", "only", "own", "same", "so",
             "than", "too", "very", "s", "t", "can", "will", "just", "don", "don't", "should", "should've", "now", "d",
             "ll", "m",
             "o", "re", "ve", "y", "ain", "aren", "aren't", "couldn", "couldn't", "didn", "didn't", "doesn", "doesn't",
             "hadn", "hadn't", "hasn", "hasn't", "haven", "haven't", "isn", "isn't", "ma", "mightn", "mightn't",
             "mustn", "mustn't",
             "needn", "needn't", "shan", "shan't", "shouldn", "shouldn't", "wasn", "wasn't", "weren", "weren't", "won",
             "won't", "wouldn", "wouldn't"]


@csrf_exempt
def redirectChange(request):
    try:
        if request.method == "POST":
            get_value = request.body
            get_value = get_value.decode("utf-8")
            get_value = json.loads(get_value)
            redirectURL = get_value['redirectURL']
            settings.LOGIN_REDIRECT_URL = redirectURL
            if redirectURL == '/':
                settings.LOGIN_REDIRECT_URL = '/home'
        return HttpResponse('done')
    except Exception as e:
        return HttpResponse(
            "Error running the program.")


def index(request):
    return render(request, 'registration/login.html')


@csrf_exempt
def testing_data_upload_view(request):
    # global training_data
    index_name = 'testingfiledata'
    try:
        if request.method == 'POST':
            # userName = request.user.username
            # # es_conn.indices.create(index=index_name)
            # testing_data = request.FILES.getlist('testFile')
            # # testing_data = open(request.FILES.get('file').temporary_file_path(), 'r').read()
            # finalTestingData = Process_All_Files(testing_data)
            # query = {"query": {"bool": {"must": {"match": {"username.keyword": userName}}}}}
            # es_conn.delete_by_query(index=index_name, body=query)
            # if len(testing_data) > 0:
            #     datafile = {
            #         'username': userName,
            #         'testing_data': finalTestingData
            #     }
            #
            #     es_conn.create(index=index_name, doc_type='_doc', body=datafile, id=uuid.uuid4())
            return HttpResponse('sucess')
    except Exception as e:
        return HttpResponse(
            "Error running the program.")


@csrf_exempt
def training_data_upload_view(request):
    # global training_data
    index_name = 'trainingfiledata'
    try:
        if request.method == 'POST':
            # userName = request.user.username;
            # # training_data = open(request.FILES.get('file').temporary_file_path(), 'r').read()
            # # es_conn.indices.create(index=index_name)
            # training_data = request.FILES.getlist('trainFile')
            # # training_data = request.FILES.getlist('file').read().decode("ISO-8859-1")
            # finalTrainingData = Process_All_Files(training_data)
            # # res = es_conn.update(index=index_name,body={"doc": {"match_all": {}}})
            # query = {"query": {"bool": {"must": {"match": {"username.keyword": userName}}}}}
            # es_conn.delete_by_query(index=index_name, body=query)
            # if len(training_data) > 0:
            #     datafile = {
            #         'username': userName,
            #         'training_data': finalTrainingData
            #     }
            #
            #     es_conn.create(index=index_name, doc_type='_doc', body=datafile, id=uuid.uuid4())

            return HttpResponse('sucess')
    except Exception as e:
        return HttpResponse(
            "Error running the program.")


# need to work on this
def xls_response(request):
    workbook = ''
    response1 = ''
    try:
        if request.method == "GET":
            userName = request.user.username;
            index_name_globals = 'apolloglobals'
            query = {"query": {"bool": {"must": {"match": {"username.keyword": userName}}}}}
            res = es_conn.search(index=index_name_globals, body=query)
            testingDataType = res['hits']['hits'][0]['_source']['testingDataType']
            trainingDataType = res['hits']['hits'][0]['_source']['trainingDataType']
            testingFileName = res['hits']['hits'][0]['_source']['testingFileName']
            if testingDataType == 'Patent':
                workbook = load_workbook('./static/template_patent.xlsx')

            elif testingDataType == 'Journal':

                workbook = load_workbook('./static/template_journal.xlsx')

            if trainingDataType == 'Patent':
                workbook_training_data = load_workbook('./static/template_patent_training_data.xlsx')

            elif trainingDataType == 'Journal':
                workbook_training_data = load_workbook('./static/template_journal_training_data.xlsx')

            resultsSheet = workbook["Results"]
            rawDataSheet = workbook["Raw_Data"]

            resultsSheet_training_data = workbook_training_data["Results"]
            rawDataSheet_training_data = workbook_training_data["Raw_Data"]

            # Write results for predicted probabilities and class names to the sheet containing raw data
            # In excel, the indices start from 1, instead of 0
            index_name_apollo = 'apolloglobals'
            query = {"query": {"bool": {"must": {"match": {"username.keyword": userName}}}}}
            res = es_conn.search(index=index_name_apollo, body=query)
            model_prob_all = res['hits']['hits'][0]['_source']['model_prob_all']
            model_predicted = res['hits']['hits'][0]['_source']['model_predicted']
            model_predicted = json.loads(model_predicted)
            model_predicted_training = res['hits']['hits'][0]['_source']['model_predicted_training']
            model_predicted_training = json.loads(model_predicted_training)
            file_test_proc = res['hits']['hits'][0]['_source']['file_test_proc']
            file_sample_proc = res['hits']['hits'][0]['_source']['file_sample_proc']
            file_test_open = res['hits']['hits'][0]['_source']['file_test_open']
            file_sample_open_training = res['hits']['hits'][0]['_source']['file_sample_open_training']
            if testingDataType == 'Patent':
                # Now, the first line is header, so remove the first line
                for row_index in np.arange(len(file_test_proc)):
                    doc = file_test_proc[row_index].split('\t')
                    for column_index in np.arange(8):
                        resultsSheet.cell(row=row_index + 2, column=column_index + 1).value = doc[column_index].strip(
                            '\"')
                    resultsSheet.cell(row=row_index + 2, column=9).value = model_predicted[row_index].strip('\"')

            elif testingDataType == 'Journal':
                for row_index in np.arange(len(file_test_open)):
                    doc = file_test_open[row_index].split('\t')
                    for column_index in np.arange(6):
                        resultsSheet.cell(row=row_index + 2, column=column_index + 1).value = doc[column_index].strip(
                            '\"')
                    resultsSheet.cell(row=row_index + 2, column=7).value = model_predicted[row_index].strip('\"')

            if trainingDataType == 'Patent':
                # Write predictions on training data in a separate file:
                for row_index in np.arange(len(file_sample_proc)):
                    doc = file_sample_proc[row_index].split('\t')
                    for column_index in np.arange(8):
                        resultsSheet_training_data.cell(row=row_index + 2, column=column_index + 1).value = doc[
                            column_index].strip('\"')
                    resultsSheet_training_data.cell(row=row_index + 2, column=9).value = model_predicted_training[
                        row_index].strip('\"')

            elif trainingDataType == 'Journal':
                # Write predictions on training data in a separate file:
                for row_index in np.arange(len(file_sample_open_training)):
                    doc = file_sample_open_training[row_index].split('\t')
                    for column_index in np.arange(7):
                        resultsSheet_training_data.cell(row=row_index + 2, column=column_index + 1).value = doc[
                            column_index].strip('\"')
                    resultsSheet_training_data.cell(row=row_index + 2, column=8).value = model_predicted_training[
                        row_index].strip('\"')

            # In the Raw_Data sheet, write the class names starting from column B

            model_prob_all = json.loads(model_prob_all)
            column_header_index = 2
            key_value = 'runDocumentClassifier/'
            key_value += userName + '/'
            s3 = boto3.client('s3')
            model = s3.get_object(Bucket=AWS_STORAGE_BUCKET_NAME,
                                  Key=key_value + 'trainedModel.pkl')

            model = model['Body'].read()
            model = pickle.loads(model)
            for cls in model.classes_:
                rawDataSheet.cell(row=1, column=column_header_index).value = cls.strip('\r')
                column_header_index += 1

            # Wirte all the probabilities for each class assgined by the model in the Raw_Data sheet
            for row_index in np.arange(len(model_prob_all)):
                for column_index in np.arange(len(model_prob_all[row_index])):
                    # The first column in template excel file is formula for 'OTHERS',
                    # hence start writing the probability values from second column in the excel sheet
                    rawDataSheet.cell(row=row_index + 2, column=column_index + 2).value = model_prob_all[
                        row_index][column_index]

            # workbook.save(re.sub('.txt', '_Threshold_Analysis.xlsx',gui_parameters['testing_file_name']))
            #
            thresholdAnalysisResultFileName = re.sub('.txt', '_Threshold_Analysis.xlsx', testingFileName)
            # response1 = HttpResponse(content_type='application/ms-excel')
            response1 = HttpResponse(content=save_virtual_workbook(workbook),
                                     content_type='application/vnd.ms-excel')
            response1['Content-Disposition'] = 'attachment; filename=' + thresholdAnalysisResultFileName
            # workbook.save(response1);
        return response1
    except Exception as e:
        return HttpResponse(
            "Error running the program.")


def training_data_xls_response(request):
    # global response2
    response2 = ''

    try:
        if request.method == "GET":
            userName = request.user.username;
            index_name_globals = 'apolloglobals'
            query = {"query": {"bool": {"must": {"match": {"username.keyword": userName}}}}}
            res = es_conn.search(index=index_name_globals, body=query)
            testingDataType = res['hits']['hits'][0]['_source']['testingDataType']
            trainingDataType = res['hits']['hits'][0]['_source']['trainingDataType']
            testingFileName = res['hits']['hits'][0]['_source']['testingFileName']
            if testingDataType == 'Patent':
                workbook = load_workbook('./static/template_patent.xlsx')

            elif testingDataType == 'Journal':

                workbook = load_workbook('./static/template_journal.xlsx')

            if trainingDataType == 'Patent':
                workbook_training_data = load_workbook('./static/template_patent_training_data.xlsx')

            elif trainingDataType == 'Journal':
                workbook_training_data = load_workbook('./static/template_journal_training_data.xlsx')

            resultsSheet = workbook["Results"]
            rawDataSheet = workbook["Raw_Data"]

            resultsSheet_training_data = workbook_training_data["Results"]
            rawDataSheet_training_data = workbook_training_data["Raw_Data"]

            # Write results for predicted probabilities and class names to the sheet containing raw data
            # In excel, the indices start from 1, instead of 0
            index_name_apollo = 'apolloglobals'
            query = {"query": {"bool": {"must": {"match": {"username.keyword": userName}}}}}
            res = es_conn.search(index=index_name_apollo, body=query)
            model_prob_all_training = res['hits']['hits'][0]['_source']['model_prob_all_training']
            model_predicted = res['hits']['hits'][0]['_source']['model_predicted']
            model_predicted = json.loads(model_predicted)
            model_predicted_training = res['hits']['hits'][0]['_source']['model_predicted_training']
            model_predicted_training = json.loads(model_predicted_training)
            file_test_proc = res['hits']['hits'][0]['_source']['file_test_proc']
            file_sample_proc = res['hits']['hits'][0]['_source']['file_sample_proc']
            file_test_open = res['hits']['hits'][0]['_source']['file_test_open']
            file_sample_open_training = res['hits']['hits'][0]['_source']['file_sample_open_training']
            if testingDataType == 'Patent':
                for row_index in np.arange(len(file_test_proc)):
                    doc = file_test_proc[row_index].split('\t')
                    for column_index in np.arange(8):
                        resultsSheet.cell(row=row_index + 2, column=column_index + 1).value = doc[column_index].strip(
                            '\"')
                    resultsSheet.cell(row=row_index + 2, column=9).value = model_predicted[row_index].strip('\"')

            elif testingDataType == 'Journal':
                for row_index in np.arange(len(file_test_open)):
                    doc = file_test_open[row_index].split('\t')
                    for column_index in np.arange(6):
                        resultsSheet.cell(row=row_index + 2, column=column_index + 1).value = doc[column_index].strip(
                            '\"')
                    resultsSheet.cell(row=row_index + 2, column=7).value = model_predicted[row_index].strip('\"')

            if trainingDataType == 'Patent':
                # Write predictions on training data in a separate file:
                for row_index in np.arange(len(file_sample_proc)):
                    doc = file_sample_proc[row_index].split('\t')
                    for column_index in np.arange(8):
                        resultsSheet_training_data.cell(row=row_index + 2, column=column_index + 1).value = doc[
                            column_index].strip('\"')
                    resultsSheet_training_data.cell(row=row_index + 2, column=9).value = model_predicted_training[
                        row_index].strip('\"')

            elif trainingDataType == 'Journal':
                # Write predictions on training data in a separate file:
                for row_index in np.arange(len(file_sample_open_training)):
                    doc = file_sample_open_training[row_index].split('\t')
                    for column_index in np.arange(7):
                        resultsSheet_training_data.cell(row=row_index + 2, column=column_index + 1).value = doc[
                            column_index].strip('\"')
                    resultsSheet_training_data.cell(row=row_index + 2, column=8).value = model_predicted_training[
                        row_index].strip('\"')

            # In the Raw_Data sheet, write the class names starting from column B

            model_prob_all_training = json.loads(model_prob_all_training)
            column_header_index = 2
            key_value = 'runDocumentClassifier/'
            key_value += userName + '/'
            s3 = boto3.client('s3')
            model = s3.get_object(Bucket=AWS_STORAGE_BUCKET_NAME,
                                  Key=key_value + 'trainedModel.pkl')

            model = model['Body'].read()
            model = pickle.loads(model)
            for cls in model.classes_:
                rawDataSheet.cell(row=1, column=column_header_index).value = cls.strip('\r')
                column_header_index += 1

            # Write all the probabilities for each class assgined by the model in the Raw_Data sheet
            for row_index in np.arange(len(model_prob_all_training)):
                for column_index in np.arange(len(model_prob_all_training[row_index])):
                    # The first column in template excel file is formula for 'OTHERS',
                    # hence start writing the probability values from second column in the excel sheet
                    rawDataSheet_training_data.cell(row=row_index + 2, column=column_index + 2).value = \
                    model_prob_all_training[row_index][column_index]

            # workbook_training_data.save(
            #     re.sub('.txt', '_Threshold_Analysis_Training_Data.xlsx', gui_parameters['testing_file_name']))

            thresholdAnalysisTrainingDataResultFileName = re.sub('.txt', '_Threshold_Analysis_Training_Data.xlsx',
                                                                 testingFileName)
            # # response2 = HttpResponse(content_type='application/ms-excel')
            response2 = HttpResponse(content=save_virtual_workbook(workbook_training_data),
                                     content_type='application/vnd.ms-excel')
            response2['Content-Disposition'] = 'attachment; filename=' + thresholdAnalysisTrainingDataResultFileName
            # workbook.save(response2);
        return response2
    except Exception as e:
        return HttpResponse(
            "Error  running the program. Please contact the IP Group Analytics Team () to resolve the issue. Please provide the error details below in your email. \nPlease provide all the steps to reproduce this issue. \n" + "-" * 40 + "\n" + str(
                e) + "\n" + "-" * 40)


# need to change till here

####################file download functions starts########################

####################file download functions ends########################

@csrf_exempt
def data(request):
    # global patentOrJournalTrainingData
    index_name = 'patentorjournal'
    try:
        if request.method == "POST":
            userName = request.user.username;
            get_value = request.body
            get_value = get_value.decode("utf-8")
            patentOrJournalTrainingData = ''

            if "identification number\ttitle\tabstract\tclaims\tapplication number\tapplication date\tcurrent assignee\tupc" in get_value.lower():
                patentOrJournalTrainingData = 'Patent'
            elif "meta data\ttitle\tabstract\tauthor\taffiliation\tpublished year" in get_value.lower():
                patentOrJournalTrainingData = 'Journal'
            elif 'nasca' in get_value.lower():
                patentOrJournalTrainingData = 'NASCA File Error.'
            else:
                patentOrJournalTrainingData = 'Training File Error.'

        return HttpResponse(patentOrJournalTrainingData)
    except Exception as e:
        return HttpResponse(
            "Error running the program.")


@login_required(login_url='/login/')
def home(request):
    return render(request, 'apollo4/home.html')


@csrf_exempt
@login_required(login_url='/login/')
def svl(request):
    return render(request, 'apollo4/superVisedLearning.html')


@csrf_exempt
@login_required(login_url='/login/')
def usvl(request):
    return render(request, 'apollo4/unSuperVised.html')


@csrf_exempt
@login_required(login_url='/login/')
def em(request):
    return render(request, 'apollo4/existingModel.html')


@csrf_exempt
@login_required(login_url='/login/')
def emus(request):
    return render(request, 'apollo4/existingModelUnsupervised.html')


@csrf_exempt
def da(request):
    return render(request, 'apollo4/dataannotate.html')


@csrf_exempt
@login_required(login_url='/login/')
def il(request):
    return render(request, 'apollo4/incrementalLearning.html')


@csrf_exempt
@login_required(login_url='/login/')
def ilu(request):
    return render(request, 'apollo4/incrementalLearningUnsupervised.html')


@csrf_exempt
@login_required(login_url='/login/')
def ps(request):
    return render(request, 'apollo4/patentScoring.html')


@csrf_exempt
@login_required(login_url='/login/')
def da(request):
    return render(request, 'apollo4/dataannotate.html')


def um(request):
    return render(request, 'apollo4/user_manual.html')


def logout(request):
    auth.logout(request)
    return render(request, 'registration/login.html')


@csrf_exempt
def fetch_update(request):
    try:
        if request.method == 'POST':
            userName = request.user.username;
            index_name_gloabals = 'apolloglobals'
            query = {"query": {"bool": {"must": {"match": {"username.keyword": userName}}}}}
            res = es_conn.search(index=index_name_gloabals, body=query)
            data = res['hits']['hits'][0]['_source']

        return JsonResponse({'data': data})
    except Exception as e:
        return HttpResponse(
            "Error running the program.")


@csrf_exempt
def userRunModelTrack(request):
    try:
        if request.method == 'POST':
            try:
                userName = request.user.username;
                userTrack = request.body.decode('utf-8');
                userTrack = json.loads(userTrack);
                testingFileName = userTrack['testing_file_name']
                trainingFileName = userTrack['training_file_name']
                testingDataType = userTrack['testing_data_type']
                trainingDataType = userTrack['training_data_type']
                str_model_name = userTrack['model']
                automaticMode = userTrack['automatic_mode']
                additional_stopwords = userTrack['additional_stopwords']
                target_performance_measure = userTrack['target_performance_measure']
            except Exception as e:
                print('parsing went wrong', e)

            time = datetime.now()
            time = time.strftime("%I:%M%p on %B %d, %Y")

            update_fields = {
                'username': userName,
                'testingFileName': testingFileName,
                'trainingFileName': trainingFileName,
                'testingDataType': testingDataType,
                'trainingDataType': trainingDataType,
                'str_model_name': str_model_name,
                'automaticMode': automaticMode,
                'additional_stopwords': additional_stopwords,
                'target_performance_measure': target_performance_measure,
                'time': time + ' UTC  Time'
            }
            index_name = 'userrunmodeltrack'
            if es_conn.indices.exists(index_name):
                es_conn.create(index=index_name, doc_type='_doc', body=update_fields, id=uuid.uuid4())
            else:
                save_response = es_conn.indices.create(index=index_name, ignore=400)
                es_conn.create(index=index_name, doc_type='_doc', body=update_fields, id=uuid.uuid4())

        return HttpResponse('sucess')
    except Exception as e:
        return HttpResponse(
            "Error running the program.")


@csrf_exempt
def runSupervisedSaving(request):
    try:
        if request.method == 'POST':
            userName = request.user.username;
            update_fields = {
                'username': userName,
                'progressbar_maximum': 0,  # progress bar max_value
                'progressbar_value': 0,  # progress bar value
                'saved_project_status': 0,
                'progressbarlabel_text': '',  # progress field
                'progress_text': '',  # progress text
                'trainingDataNumInstances': None,  # Training examples
                'trainingDataNumClasses': None,  # training classes
                'trainingDataStatistics': None,  # trainig data table
                'trainingDataPerformance': None,  # model evaluation before +/-
                'trainingDataPerformancesStandardDeviations': None,
                # model evaluation after +/-
                'str_parameter_name': None,  # Hyper parameter left before =
                'optimal_model_parameter': None,  # Hyper parameter right after =
                'testingDataNumInstances': None,  # Testing examples
                'testingDataStatistics': None,  # testing data table
                'final_progress_value': 0,
                'user_defined_stopwords': '',
                'testingFileName': '',
                'trainingFileName': '',
                'excel_status_code': 0,
                'current_tab': 0,
                'newclassString': '',
                'errorString': '',
                'file_test_proc': '',
                'model_predicted': '',
                'file_test_open': '',
                'file_sample_proc': '',
                'model_predicted_training': '',
                'file_sample_open_training': '',
                'model': '',
                'model_prob_all_training': '',
                'model_prob_all': '',
                'testingFileName': '',
                'trainingFileName': '',
                'testingDataType': '',
                'trainingDataType': '',
                'model_isotonic_calibration': None,
                'model_sigmoid_calibration': None,
                'tfidfVectorizer': None
            }
            index_name_gloabals = 'apolloglobals'
            query = {"query": {"bool": {"must": {"match": {"username.keyword": userName}}}}}
            es_conn.delete_by_query(index=index_name_gloabals, body=query)
            es_conn.create(index=index_name_gloabals, doc_type='_doc', body=update_fields, id=uuid.uuid4())
        return HttpResponse('sucess')
    except Exception as e:
        return HttpResponse(
            "Error running the program.")


####################runDocumentClassifierSupervised starts########################
@csrf_exempt
def runDocumentClassifierSupervised(request):
    relevanceThreshold = 0.0
    file_others = []
    file_classified = []
    model_classified = []
    global responsesupervised

    try:
        # Assuming that the client-side has already selected the options before running the program.
        # Assuming that the request from the client side will have all the fields necessary for running the program.
        if request.method == "GET":
            return responsesupervised
        if request.method == "POST":
            try:
                userName = request.user.username;
                finalTrainingData = request.FILES.getlist('trainFile')
                finalTestingData = request.FILES.getlist('testFile')
                gui_parameters = request.POST.getlist('inputData')[0]
                gui_parameters = json.loads(gui_parameters);
                testingFileName = gui_parameters['testing_file_name']
                trainingFileName = gui_parameters['training_file_name']
                testingDataType = gui_parameters['testing_data_type']
                trainingDataType = gui_parameters['training_data_type']
                str_model_name = gui_parameters['model']
            except Exception as e:
                print('parsing went wrong', e)

            # Thread(target=runSupervisedSaving,args=(userName,)).start()
            # set the progress bar values
            training_data = Process_All_Files(finalTrainingData)
            testing_data = Process_All_Files(finalTestingData)
            index_name_gloabals = 'apolloglobals'
            query = {"query": {"bool": {"must": {"match": {"username.keyword": userName}}}}}
            res = es_conn.search(index=index_name_gloabals, body=query)
            id = res['hits']['hits'][0]['_id']
            if gui_parameters['automatic_mode'] == 'True':
                progressbar_maximum = 235
            else:
                progressbar_maximum = 100

                # apollo4.globals.progressbar_maximum = 100

            es_conn.update(index=index_name_gloabals, id=id,
                           body={"doc": {"progressbar_maximum": progressbar_maximum, "current_tab": 1,
                                         "testingFileName": testingFileName, "trainingFileName": trainingFileName,
                                         "testingDataType": testingDataType, "trainingDataType": trainingDataType}})
            # apollo4.globals.progressbar_value = 0

            # 1. supervised learning

            # Set the text in progressbarlabel
            programRunStartTime = datetime.now()
            programRunStartTimeLabel = 'Progress: Program run started at ' + programRunStartTime.strftime(
                "%I:%M%p on %B %d, %Y") + ' (UTC time). '

            progressbarlabel_text = ''
            progressbarlabel_text = programRunStartTimeLabel
            progress_text = ''
            progress_text = progress_text + '-' * 75 + '\n' + "Program run started at " + programRunStartTime.strftime(
                "%I:%M%p on %B %d, %Y") + " (UTC time).\n" + \
                            '-' * 75 + '\n' + "Starting document classification process..."
            es_conn.update(index=index_name_gloabals, id=id,
                           body={"doc": {"progressbar_value": 0, 'progressbarlabel_text': progressbarlabel_text,
                                         'progress_text': progress_text}})

            # The code for loading and pre-processing the data is different for patent and journal data
            if gui_parameters['training_data_type'] == 'Patent':
                file_sample_open = training_data
                file_sample_open = file_sample_open.split('\n')  # split by new line
                file_sample_open = list(filter(None, file_sample_open))  # delete empty lines

                # Now, the first line is header, so remove the first line
                file_sample_open = file_sample_open[1:]

                progress_text = progress_text + "\nFound " + str(
                    len(file_sample_open)) + " documents! \nPreprocessing documents...\n"

                progressbar_value = 5

                # Set value of progressbar to 5 once the training dataset is loaded

                # Build the stop words
                stops = stopwords

                aux_stops = './static/AuxStops.txt'

                aux_stops = open(aux_stops, 'r').read()
                aux_stops = re.sub("[^a-zA-Z ]", "   ", aux_stops)  # remove non-alphanumeric
                aux_stops = " ".join(aux_stops.split())  # split by any whitespace and rejoin w/ space
                aux_stops = aux_stops.split(' ')
                aux_stops = list(filter(None, aux_stops))

                # append auxiliary stops
                stops = stops + aux_stops

                # append user-provided stop words
                user_defined_stopwords = (" ".join(gui_parameters['additional_stopwords'].lower().split(','))).split()
                es_conn.update(index=index_name_gloabals, id=id,
                               body={"doc": {"progress_text": progress_text, "progressbar_value": progressbar_value,
                                             "user_defined_stopwords": user_defined_stopwords}})
                stops = stops + user_defined_stopwords

                # Bulid stopterm dictionary
                stopterms = {}
                for stop in stops:
                    if stop in stopterms:
                        stopterms[stop] += 1
                    else:
                        stopterms[stop] = 1

                # Preprocess the sample file
                # MS: Need to check whether we will receive the file or the text from the GUI.
                # Baesd on that, change the function below.
                (file_sample_proc, file_sample_stem, temp) = preprocess_collection(file_sample_open, stopterms, True,
                                                                                   progress_text)
                file_sample = list(filter(None, file_sample_stem))
                title_samples = [doc.split('\t')[1].strip('\r').strip('\n') for doc in file_sample]
                abstract_samples = [doc.split('\t')[2].strip('\r').strip('\n') for doc in file_sample]
                claim_samples = [doc.split('\t')[3].strip('\r').strip('\n') for doc in file_sample]
                label_samples = [doc.split('\t')[8].lower().strip('\r').strip('\n') for doc in file_sample]
                labels = sorted(list(set(label_samples)))
                labels.append(u'others')
                train_data = [' '.join(doc) for doc in zip(title_samples, abstract_samples, claim_samples)]
                train_target = label_samples
                es_conn.update(index=index_name_gloabals, id=id,
                               body={"doc": {"progress_text": progress_text, "progressbar_value": progressbar_value,
                                             "user_defined_stopwords": user_defined_stopwords,
                                             "file_sample_proc": file_sample_proc}})
                # End patent training data

            elif gui_parameters['training_data_type'] == 'Journal':
                # MS: Need to check whether we will receive the file or the text from the GUI.
                # Baesd on that, change the function below.
                file_sample_open = training_data
                file_sample_open = file_sample_open.split('\n')  # split by new line
                file_sample_open = list(filter(None, file_sample_open))  # delete empty lines

                # Now, the first line is header, so remove the first line
                file_sample_open = file_sample_open[1:]

                progress_text = progress_text + "\nFound " + str(
                    len(file_sample_open)) + " documents! \nPreprocessing documents...\n"

                # Set value of progressbar to 5 once the training dataset is loaded
                progressbar_value = 5

                # Remove the duplicated documents based on "title"
                file_sample_open_training = dedup_collection_journal(file_sample_open, 1, 2)

                # Preprocessing for scoupus data
                file_sample_open = preprocess_collection_journal(file_sample_open_training)

                # Take the stopwords from the GUI and add them to the stopwords list
                user_defined_stopwords = gui_parameters['additional_stopwords'].lower()

                file_sample_data = [' '.join([stop_and_stem_journal(doc.split('\t')[1], user_defined_stopwords)
                                                 , stop_and_stem_journal(doc.split('\t')[2], user_defined_stopwords)
                                              ]) for doc in file_sample_open]

                # Training Phase
                label_samples = [doc.split('\t')[-1].lower().strip('\r').strip('\n') for doc in file_sample_open]
                labels = sorted(list(set(label_samples)))
                labels.append(u'others')
                train_data = file_sample_data
                train_target = label_samples
                # end journal training data preprocessing
                es_conn.update(index=index_name_gloabals, id=id,
                               body={"doc": {"progress_text": progress_text, "progressbar_value": progressbar_value,
                                             "user_defined_stopwords": user_defined_stopwords,
                                             "file_sample_open_training": file_sample_open_training}})

            progress_text = progress_text + "Removed duplicates and preprocessed " + str(
                len(train_data)) + " documents."

            # Set value of progressbar to 10 once the training dataset is preprocessed
            progressbar_value += 5

            trainingDataNumInstances = len(train_target)
            trainingDataNumClasses = len(set(train_target))
            es_conn.update(index=index_name_gloabals, id=id,
                           body={"doc": {"progress_text": progress_text, "progressbar_value": progressbar_value,
                                         "trainingDataNumInstances": trainingDataNumInstances,
                                         "trainingDataNumClasses": trainingDataNumClasses}})
            # Display information about training examples class distribution

            numInstancesTrainingData = len(train_data)
            trainingDataStats = []

            leastDocumentsForAClass = 5  # initialize to 5
            for label in set(train_target):
                distribution = str(np.round(train_target.count(label) * 100.0 / len(train_target) * 1.0, 2)) + '%'
                trainingDataStats.append([label, train_target.count(label), distribution])
                if train_target.count(label) < leastDocumentsForAClass:
                    leastDocumentsForAClass = train_target.count(label)

            # Update the treeview with the distribution of instances in the training data
            trainingDataStatistics = trainingDataStats

            # Make sure that there are at least 5 documents for each class:
            # this is required to perform 5-fold cross validation
            if leastDocumentsForAClass < 5:
                progress_text = progress_text + "*" * 50 + "\nThe program requires at least 5 training examples for each class. Please provide at least 5 training examples for each class and re-run the program." + '\n' + "*" * 50
                errorString = 'The program requires at least 5 training examples for each class. Please provide at least 5 training examples for each class and re-run the program.'
                final_progress_value = 200
                es_conn.update(index=index_name_gloabals, id=id,
                               body={"doc": {"progress_text": progress_text, "errorString": errorString,
                                             "final_progress_value": final_progress_value}})
                # Display a messagebox in the GUI
                return HttpResponse(
                    "The program requires at least 5 training examples for each class. Please provide at least 5 training examples for each class and re-run the program.")

            if len(set(train_target)) < 3 and 'One vs Rest' in str_model_name:
                # Display a messagebox in the GUI
                errorString = 'One vs Rest models are supported for only more than two classes in the data. There are less than three classes. Please select a model that is NOT One vs Rest.'
                final_progress_value = 200
                es_conn.update(index=index_name_gloabals, id=id,
                               body={"doc": {"errorString": errorString,
                                             "final_progress_value": final_progress_value}})
                return HttpResponse(
                    "One vs Rest models are supported for only more than two classes in the data. There are less than three classes. Please select a model that is NOT One vs Rest.")

            progress_text = progress_text + "\nStarting model training..."
            es_conn.update(index=index_name_gloabals, id=id,
                           body={"doc": {"trainingDataStatistics": json.dumps(trainingDataStatistics),
                                         "progress_text": progress_text}})

            cv = CountVectorizer()
            tfidf = TfidfTransformer()

            # Changed the n-grams to (1,5) in the line below, and max_df from 0.5 to 0.8, based on side-experiments
            tfidf_vect = TfidfVectorizer(analyzer='word',
                                         ngram_range=(1, 5),
                                         min_df=2,
                                         max_df=0.8,
                                         max_features=200000,
                                         stop_words='english',
                                         use_idf=True)

            # tf-idf with params
            train_tfidf = tfidf_vect.fit_transform(train_data)

            # Set value of progressbar to 15 once the training dataset is vectorized
            progressbar_value += 5

            tfidfVectorizer = tfidf_vect

            # Model and model parameters

            svm_alpha = 1.0  # default value
            mnb_alpha = 0.001  # default value
            lrl2_alpha = 1.0  # default value
            svm_kernel = 'linear'  # default value
            class_weight = None  # default value
            batchSize = 16
            maxSequenceLength = 256

            automatic_mode = False
            optimal_model_parameter = -1

            progress_text = progress_text + "\nOptimizing model parameters..."

            str_parameter_name = 'Alpha = '
            key_value = 'runDocumentClassifier/'
            key_value += userName + '/'
            s3 = boto3.client('s3')
            tfidf = pickle.dumps(tfidfVectorizer)
            s3.put_object(Body=tfidf, Bucket=AWS_STORAGE_BUCKET_NAME,
                          Key=key_value + 'tfidf_vect.pkl')
            es_conn.update(index=index_name_gloabals, id=id,
                           body={"doc": {"optimal_model_parameter": optimal_model_parameter,
                                         "progress_text": progress_text, "str_parameter_name": str_parameter_name,
                                         "progressbar_value": progressbar_value}})

            if gui_parameters['automatic_mode'] == True:
                str_model_name = 'automatic'

            if str_model_name == 'automatic':
                # determine the best model based on evaluating several models
                automatic_mode = True

                progress_text = progress_text + "\nEvaluating all the models with optimal parameter settings..."
                chosen_model, optimal_model_parameter, max_performance_measure, standard_deviation, all_measures_performances, all_measures_standardDeviations = getBestModelAndHyperParameters(
                    train_tfidf, train_data, train_target, gui_parameters['target_performance_measure'])
                trainingDataPerformances = all_measures_performances
                trainingDataPerformancesStandardDeviation = all_measures_standardDeviations

                # set model to chosen_model
                str_model_name = chosen_model
                str_parameter_name = 'Alpha = '

                progress_text = progress_text + "\n***************************************************************************" + \
                                "\nThe BEST model based on 5-fold cross validation on training data is: " + str_model_name + \
                                "\nPlease refer to the detailed results for each model in the PROGRESS frame above." + \
                                "\n***************************************************************************"
                es_conn.update(index=index_name_gloabals, id=id,
                               body={"doc": {"trainingDataPerformances": trainingDataPerformances,
                                             'trainingDataPerformancesStandardDeviation': trainingDataPerformancesStandardDeviation,
                                             "progress_text": progress_text, "str_parameter_name": str_parameter_name}})

            if str_model_name == 'Multinomial Naive Bayes':
                # Get optimal alpha for the model
                mnb_alpha = -1
                if automatic_mode == False:
                    mnb_alpha, max_performance_measure, standard_deviation, all_measures_performances, all_measures_standardDeviations = getOptimalParameterForMNB_alpha(
                        train_tfidf.todense(), train_data, train_target, gui_parameters['target_performance_measure'])
                    trainingDataPerformances = all_measures_performances
                    trainingDataPerformancesStandardDeviation = all_measures_standardDeviations
                    optimal_model_parameter = mnb_alpha
                else:
                    # The best model is already computed and best parameter is already determined
                    mnb_alpha = optimal_model_parameter

                model = MultinomialNB(alpha=mnb_alpha).partial_fit(train_tfidf.todense(), train_target,
                                                                   classes=np.unique(train_target))
                trainedModel = model
                trainedModelName = 'Multinomial_Naive_Bayes_Alpha=' + str(
                    mnb_alpha) + '_' + gui_parameters['target_performance_measure'].replace(' (default)', '')
                es_conn.update(index=index_name_gloabals, id=id,
                               body={"doc": {"trainingDataPerformance": list(trainingDataPerformances),
                                             "trainingDataPerformancesStandardDeviations": list(
                                                 trainingDataPerformancesStandardDeviation),
                                             "optimal_model_parameter": optimal_model_parameter,
                                             "trainedModelName": trainedModelName}})

            elif str_model_name == 'Logistics Regression':
                # Get optimal alpha for the model
                lrl2_alpha = -1
                if automatic_mode == False:
                    lrl2_alpha, max_performance_measure, standard_deviation, all_measures_performances, all_measures_standardDeviations = getOptimalParameterForLR_alpha(
                        train_tfidf.toarray(), train_data, train_target, gui_parameters['target_performance_measure'])
                    trainingDataPerformances = all_measures_performances
                    trainingDataPerformancesStandardDeviation = all_measures_standardDeviations
                    optimal_model_parameter = lrl2_alpha
                else:
                    lrl2_alpha = optimal_model_parameter

                random_state = np.random.RandomState(seed=87654)

                # output of the model is dependent on the interaction between alpha and the number of epochs (n_iter)
                # When alpha is very small, n_iter must be large to compensate for the slower learning rate
                # When using SGD, the partial_fit method has to be applied on different batches of the training data,
                # and we need to epoch multiple times
                model = SGDClassifier(loss='log', penalty='l2', alpha=lrl2_alpha, class_weight=class_weight,
                                      random_state=random_state)

                train_tfidf_dense = train_tfidf.toarray()

                def batches(l, n):
                    for i in np.arange(0, len(l), n):
                        yield l[i:i + n]

                n_iter = 25
                np.random.seed(5647)
                shuffledRange = np.arange(len(train_data))
                for n in np.arange(n_iter):
                    np.random.shuffle(shuffledRange)
                    shuffled_train_tfidf = [train_tfidf_dense[i] for i in shuffledRange]
                    shuffled_train_target = [train_target[i] for i in shuffledRange]

                    # Training the model in 10 batches
                    for batch in batches(np.arange(len(shuffled_train_target)), 5):
                        model.partial_fit(shuffled_train_tfidf[batch[0]:batch[-1] + 1],
                                          shuffled_train_target[batch[0]:batch[-1] + 1],
                                          classes=np.unique(train_target))

                trainedModel = model
                trainedModelName = 'Logistic_Regression_Alpha=' + str(
                    lrl2_alpha) + '_' + gui_parameters['target_performance_measure'].replace(' (default)', '')
                es_conn.update(index=index_name_gloabals, id=id,
                               body={"doc": {"trainingDataPerformance": list(trainingDataPerformances),
                                             "trainingDataPerformancesStandardDeviations": list(
                                                 trainingDataPerformancesStandardDeviation),
                                             "optimal_model_parameter": optimal_model_parameter,
                                             "trainedModelName": trainedModelName}})

            elif str_model_name == 'Support vector Machine':
                # Get optimal alpha for the model, performance of 5-fold CV, and standard deviation of performance
                if automatic_mode == False:
                    svm_alpha, max_performance_measure, standard_deviation, all_measures_performances, all_measures_standardDeviations = getOptimalParameterForSVM_alpha(
                        train_tfidf.toarray(), train_data, train_target, gui_parameters['target_performance_measure'])
                    trainingDataPerformances = all_measures_performances
                    trainingDataPerformancesStandardDeviation = all_measures_standardDeviations
                    optimal_model_parameter = svm_alpha

                else:
                    svm_alpha = optimal_model_parameter

                random_state = np.random.RandomState(seed=87654)

                # output of the model is dependent on the interaction between alpha and the number of epochs (n_iter)
                # When alpha is very small, n_iter must be large to compensate for the slower learning rate
                # When using SGD, the partial_fit method has to be applied on different batches of the training data,
                # and we need to epoch multiple times
                model = SGDClassifier(loss='hinge', penalty='l2', alpha=svm_alpha, class_weight=class_weight,
                                      random_state=random_state)

                train_tfidf_dense = train_tfidf.toarray()

                def batches(l, n):
                    for i in np.arange(0, len(l), n):
                        yield l[i:i + n]

                n_iter = 25
                np.random.seed(5647)
                shuffledRange = np.arange(len(train_data))
                for n in np.arange(n_iter):
                    np.random.shuffle(shuffledRange)
                    shuffled_train_tfidf = [train_tfidf_dense[i] for i in shuffledRange]
                    shuffled_train_target = [train_target[i] for i in shuffledRange]

                    # Training the model in 10 batches
                    for batch in batches(np.arange(len(shuffled_train_target)), 5):
                        model.partial_fit(shuffled_train_tfidf[batch[0]:batch[-1] + 1],
                                          shuffled_train_target[batch[0]:batch[-1] + 1],
                                          classes=np.unique(train_target))

                trainedModel = model
                trainedModelName = 'Support_Vector_Machine_Alpha=' + str(
                    svm_alpha) + '_' + gui_parameters['target_performance_measure'].replace(' (default)', '')
                es_conn.update(index=index_name_gloabals, id=id,
                               body={"doc": {"trainingDataPerformance": list(trainingDataPerformances),
                                             "trainingDataPerformancesStandardDeviations": list(
                                                 trainingDataPerformancesStandardDeviation),
                                             "optimal_model_parameter": optimal_model_parameter,
                                             "trainedModelName": trainedModelName}})

            if str_model_name == 'One vs Rest(Multinomial Naive Bayes)':
                # Get optimal alpha for the model
                mnb_alpha = -1
                if automatic_mode == False:
                    mnb_alpha, max_performance_measure, standard_deviation, all_measures_performances, all_measures_standardDeviations = getOptimalParameterForOVRMNB_alpha(
                        train_tfidf.todense(), train_data, train_target, gui_parameters['target_performance_measure'])
                    trainingDataPerformances = all_measures_performances
                    trainingDataPerformancesStandardDeviation = all_measures_standardDeviations
                    optimal_model_parameter = mnb_alpha

                else:
                    # The best model is already computed and best parameter is already determined
                    mnb_alpha = optimal_model_parameter
                model = OneVsRestClassifier(MultinomialNB(alpha=mnb_alpha)).partial_fit(train_tfidf.todense(),
                                                                                        train_target,
                                                                                        classes=np.unique(train_target))
                trainedModel = model
                trainedModelName = 'One_vs_Rest_Multinomial_Naive_Bayes_Alpha=' + str(
                    mnb_alpha) + '_' + gui_parameters['target_performance_measure'].replace(' (default)', '')
                es_conn.update(index=index_name_gloabals, id=id,
                               body={"doc": {"trainingDataPerformance": list(trainingDataPerformances),
                                             "trainingDataPerformancesStandardDeviations": list(
                                                 trainingDataPerformancesStandardDeviation),
                                             "optimal_model_parameter": optimal_model_parameter,
                                             "trainedModelName": trainedModelName}})

            elif str_model_name == 'One vs Rest(Logistic Regression)':
                # Get optimal alpha for the model
                lrl2_alpha = -1
                if automatic_mode == False:
                    lrl2_alpha, max_performance_measure, standard_deviation, all_measures_performances, all_measures_standardDeviations = getOptimalParameterForOVRLR_alpha(
                        train_tfidf.toarray(), train_data, train_target, gui_parameters['target_performance_measure'])
                    trainingDataPerformances = all_measures_performances
                    trainingDataPerformancesStandardDeviation = all_measures_standardDeviations
                    optimal_model_parameter = lrl2_alpha

                else:
                    lrl2_alpha = optimal_model_parameter

                random_state = np.random.RandomState(seed=87654)

                # output of the model is dependent on the interaction between alpha and the number of epochs (n_iter)
                # When alpha is very small, n_iter must be large to compensate for the slower learning rate
                # When using SGD, the partial_fit method has to be applied on different batches of the training data, and we need to epoch multiple times
                model = OneVsRestClassifier(
                    SGDClassifier(loss='log', penalty='l2', alpha=lrl2_alpha, class_weight=class_weight,
                                  random_state=random_state))

                train_tfidf_dense = train_tfidf.toarray()

                def batches(l, n):
                    for i in np.arange(0, len(l), n):
                        yield l[i:i + n]

                n_iter = 25
                np.random.seed(5647)
                shuffledRange = np.arange(len(train_data))
                for n in np.arange(n_iter):
                    np.random.shuffle(shuffledRange)
                    shuffled_train_tfidf = [train_tfidf_dense[i] for i in shuffledRange]
                    shuffled_train_target = [train_target[i] for i in shuffledRange]

                    # Training the model in 10 batches
                    for batch in batches(np.arange(len(shuffled_train_target)), 5):
                        model.partial_fit(shuffled_train_tfidf[batch[0]:batch[-1] + 1],
                                          shuffled_train_target[batch[0]:batch[-1] + 1],
                                          classes=np.unique(train_target))

                trainedModel = model
                trainedModelName = 'One_vs_Rest_Logistic_Regression_Alpha=' + str(
                    lrl2_alpha) + '_' + gui_parameters['target_performance_measure'].replace(' (default)', '')
                es_conn.update(index=index_name_gloabals, id=id,
                               body={"doc": {"trainingDataPerformance": list(trainingDataPerformances),
                                             "trainingDataPerformancesStandardDeviations": list(
                                                 trainingDataPerformancesStandardDeviation),
                                             "optimal_model_parameter": optimal_model_parameter,
                                             "trainedModelName": trainedModelName}})


            elif str_model_name == 'One vs Rest(Support vector Machine)':
                # Get optimal alpha for the model, performance of 5-fold CV, and standard deviation of performance

                if automatic_mode == False:
                    svm_alpha, max_performance_measure, standard_deviation, all_measures_performances, all_measures_standardDeviations = getOptimalParameterForOVRSVM_alpha(
                        train_tfidf.toarray(), train_data, train_target, gui_parameters['target_performance_measure'])
                    trainingDataPerformances = all_measures_performances
                    trainingDataPerformancesStandardDeviation = all_measures_standardDeviations
                    optimal_model_parameter = svm_alpha

                else:
                    svm_alpha = optimal_model_parameter

                random_state = np.random.RandomState(seed=87654)

                # output of the model is dependent on the interaction between alpha and the number of epochs (n_iter)
                # When alpha is very small, n_iter must be large to compensate for the slower learning rate
                # When using SGD, the partial_fit method has to be applied on different batches of the training data,
                # and we need to epoch multiple times
                model = OneVsRestClassifier(
                    SGDClassifier(loss='hinge', penalty='l2', alpha=svm_alpha, class_weight=class_weight,
                                  random_state=random_state))

                train_tfidf_dense = train_tfidf.toarray()

                def batches(l, n):
                    for i in np.arange(0, len(l), n):
                        yield l[i:i + n]

                n_iter = 25
                np.random.seed(5647)
                shuffledRange = np.arange(len(train_data))
                for n in np.arange(n_iter):
                    np.random.shuffle(shuffledRange)
                    shuffled_train_tfidf = [train_tfidf_dense[i] for i in shuffledRange]
                    shuffled_train_target = [train_target[i] for i in shuffledRange]

                    # Training the model in 10 batches
                    for batch in batches(np.arange(len(shuffled_train_target)), 5):
                        model.partial_fit(shuffled_train_tfidf[batch[0]:batch[-1] + 1],
                                          shuffled_train_target[batch[0]:batch[-1] + 1],
                                          classes=np.unique(train_target))

                trainedModel = model
                # needs to be clarified#########################################################################################################
                trainedModelName = 'One_vs_Rest_Support_Vector_Machine_Alpha=' + str(
                    svm_alpha) + '_' + gui_parameters['target_performance_measure'].replace(' (default)', '')
                #################################################################################################################################

                es_conn.update(index=index_name_gloabals, id=id,
                               body={"doc": {"trainingDataPerformance": list(trainingDataPerformances),
                                             "trainingDataPerformancesStandardDeviations": list(
                                                 trainingDataPerformancesStandardDeviation),
                                             "optimal_model_parameter": optimal_model_parameter,
                                             "trainedModelName": trainedModelName}})

            else:
                # Get optimal alpha for the model, performance of 5-fold CV, and standard deviation of performance
                if str_model_name == 'BERT':
                    dl_model_type = 'bert'
                    dl_model_name = 'bert-base-cased'
                elif str_model_name == 'RoBERTa':
                    dl_model_type = 'roberta'
                    dl_model_name = 'roberta-base'
                elif str_model_name == 'XLNet':
                    dl_model_type = 'xlnet'
                    dl_model_name = 'xlnet-base-cased'

                if automatic_mode == False:
                    # Need to set the output directory, where the temporary results for the model training will be stored
                    torch.cuda.empty_cache()
                    DEEP_LEARNING_OUTPUT_DIR = './DeepLearningOutputs/' + userName + '/'
                    if not os.path.exists(DEEP_LEARNING_OUTPUT_DIR):
                        os.makedirs(DEEP_LEARNING_OUTPUT_DIR)
                    print(123)
                    cudaExist = torch.cuda.is_available()
                    deeperror = {
                        'username': userName,
                        'error': cudaExist,
                    }
                    index_name = 'deeplearningerror'
                    if es_conn.indices.exists(index_name):
                        es_conn.create(index=index_name, doc_type=TYPE_NAME_USER, body=deeperror,
                                       id=uuid.uuid4())
                    else:
                        save_response = es_conn.indices.create(index=index_name, ignore=400)
                        es_conn.create(index=index_name, doc_type=TYPE_NAME_USER, body=deeperror,
                                       id=uuid.uuid4())
                    optimal_parameters, max_performance_measure, standard_deviation, all_measures_performances, all_measures_standardDeviations = getOptimalParametersForDeepLearning(
                        dl_model_type, dl_model_name, training_data, gui_parameters['target_performance_measure'],
                        DEEP_LEARNING_OUTPUT_DIR)
                    print(345)
                    trainingDataPerformances = all_measures_performances
                    trainingDataPerformancesStandardDeviation = all_measures_standardDeviations
                    optimal_model_parameter = str(optimal_parameters[0]) + '_' + str(optimal_parameters[1])
                    batchSize = optimal_parameters[0]
                    maxSequenceLength = optimal_parameters[1]

                random_state = np.random.RandomState(seed=87654)

                # output of the model is dependent on the interaction between alpha and the number of epochs (n_iter)
                # When alpha is very small, n_iter must be large to compensate for the slower learning rate
                # When using SGD, the partial_fit method has to be applied on different batches of the training data,
                # and we need to epoch multiple times

                total_labels = len(Counter(train_target))

                # Need to set the output directory, where the temporary results for the model training will be stored
                torch.cuda.empty_cache()
                DEEP_LEARNING_OUTPUT_DIR = './DeepLearningOutputs/' + userName + '/'
                if not os.path.exists(DEEP_LEARNING_OUTPUT_DIR):
                    os.makedirs(DEEP_LEARNING_OUTPUT_DIR)

                # Create a TransformerModel
                trainedModel = DeepLearningModel(dl_model_type, dl_model_name, batchSize, maxSequenceLength,
                                                 num_epochs=30, random_state=4987, output_dir=DEEP_LEARNING_OUTPUT_DIR)

                trainedModel.fit(training_data)

                trainedModelName = 'BERT_BatchSize=' + str(batchSize) + '_MaxSequenceLength=' + str(
                    maxSequenceLength)
                # needs to be clarified#########################################################################################################
                trainedModelName = 'BERT_BatchSize=' + str(batchSize) + '_MaxSequenceLength=' + str(
                    maxSequenceLength) + '_' + gui_parameters['target_performance_measure'].replace(' (default)', '')
                #################################################################################################################################

                es_conn.update(index=index_name_gloabals, id=id,
                               body={"doc": {"trainingDataPerformance": list(trainingDataPerformances),
                                             "trainingDataPerformancesStandardDeviations": list(
                                                 trainingDataPerformancesStandardDeviation),
                                             "optimal_model_parameter": optimal_model_parameter,
                                             "trainedModelName": trainedModelName}})

            # Prediction Phase
            trainedModel = pickle.dumps(trainedModel)
            s3.put_object(Body=trainedModel, Bucket=AWS_STORAGE_BUCKET_NAME,
                          Key=key_value + 'trainedModel.pkl')
            # The code for patent and journal testing data is different because it required different preprocessing
            if gui_parameters['testing_data_type'] == 'Patent':
                file_test_open = testing_data

                if 'Deep Learning' not in str_model_name:
                    file_test_open = file_test_open.split('\n')  # split by new line
                    file_test_open = list(filter(None, file_test_open))  # delete empty lines

                    # Now, the first line is header, so remove the first line
                    file_test_open = file_test_open[1:]
                    progress_text = progress_text + "\nPreprocessing unlabeled data..."
                    # Build the stop words
                    stops = stopwords

                    aux_stops = './static/AuxStops.txt'

                    aux_stops = open(aux_stops, 'r').read()
                    aux_stops = re.sub("[^a-zA-Z ]", "   ", aux_stops)  # remove non-alphanumeric
                    aux_stops = " ".join(aux_stops.split())  # split by any whitespace and rejoin w/ space
                    aux_stops = aux_stops.split(' ')
                    aux_stops = list(filter(None, aux_stops))

                    # append auxiliary stops
                    stops = stops + aux_stops

                    # Bulid stopterm dictionary
                    stopterms = {}
                    for stop in stops:
                        if stop in stopterms:
                            stopterms[stop] += 1
                        else:
                            stopterms[stop] = 1

                    (file_test_proc, file_test_stem, temp) = preprocess_collection(file_test_open, stopterms, False,
                                                                                   progress_text)
                    file_test_proc = list(filter(None, file_test_proc))
                    es_conn.update(index=index_name_gloabals, id=id,
                                   body={"doc": {"file_test_proc": file_test_proc}})
                    file_test = list(filter(None, file_test_stem))

                    title_test = [doc.split('\t')[1] for doc in file_test]
                    abstract_test = [doc.split('\t')[2] for doc in file_test]
                    claim_test = [doc.split('\t')[3] for doc in file_test]
                    test_data = [' '.join(doc) for doc in zip(title_test, abstract_test, claim_test)]

                    progressbar_value += 5

                    progress_text = progress_text + "\nMaking predictions on unlabeled data..."
                    es_conn.update(index=index_name_gloabals, id=id,
                                   body={
                                       "doc": {"progress_text": progress_text, "progressbar_value": progressbar_value}})
                    # convert text data to tfidf
                    test_data_tfidf = tfidf_vect.transform(test_data)
                    model_predicted = model.predict(test_data_tfidf.todense())
                    model_predicted = model_predicted.astype('U128')

                else:
                    progressbar_value += 5

                    progress_text = progress_text + "\nMaking predictions on unlabeled data..."
                    es_conn.update(index=index_name_gloabals, id=id,
                                   body={
                                       "doc": {"progress_text": progress_text, "progressbar_value": progressbar_value}})

                    model_predicted = model.predict(testing_data)
                    model_predicted = model_predicted.astype('U128')

            elif gui_parameters['testing_data_type'] == 'Journal':
                file_test_open = testing_data

                if 'Deep Learning' not in str_model_name:
                    file_test_open = file_test_open.split('\n')  # split by new line
                    file_test_open = list(filter(None, file_test_open))  # delete empty lines

                    # Now, the first line is header, so remove the first line
                    file_test_open = file_test_open[1:]

                    progress_text = progress_text + "\nPreprocessing unlabeled data..."

                    # Remove the duplicated document based on "title"
                    file_test_open = dedup_collection_journal(file_test_open, 1, 2)

                    # preprocessing for scoupus data
                    file_test_proc = preprocess_collection_journal(file_test_open)

                    user_defined_stopwords = []

                    test_data = [' '.join([stop_and_stem_journal(doc.split('\t')[1], user_defined_stopwords)
                                              , stop_and_stem_journal(doc.split('\t')[2], user_defined_stopwords)
                                           ]) for doc in file_test_proc]

                    progressbar_value += 5

                    progress_text = progress_text + "\nMaking predictions on unlabeled data..."

                    # convert text data to tfidf
                    test_data_tfidf = tfidf_vect.transform(test_data)
                    model_predicted = model.predict(test_data_tfidf.todense())
                    model_predicted = model_predicted.astype('U128')

                else:
                    progressbar_value += 5

                    progress_text = progress_text + "\nMaking predictions on unlabeled data..."
                    es_conn.update(index=index_name_gloabals, id=id,
                                   body={
                                       "doc": {"progress_text": progress_text, "progressbar_value": progressbar_value}})

                    model_predicted = model.predict(testing_data)
                    model_predicted = model_predicted.astype('U128')
                # end journal test data preprocessing
            model_predicted_training = model.predict(train_tfidf.todense())
            model_predicted_training = model_predicted_training.astype('U128')
            model_isotonic_calibration = None
            model_sigmoid_calibration = None
            es_conn.update(index=index_name_gloabals, id=id,
                           body={"doc": {"progress_text": progress_text, "progressbar_value": progressbar_value,
                                         "model_predicted": json.dumps(model_predicted.tolist()),
                                         "model_predicted_training": json.dumps(model_predicted_training.tolist()),
                                         "file_test_open": file_test_open, "file_test_proc": file_test_proc}})
            model_prob = None
            # The 'predict_proba 'function cannot be used for SGD with hinge loss,
            # hence we need calibrate the probability estimates for SGD with hinge loss
            if str_model_name == 'Support vector Machine' or str_model_name == 'One vs Rest(Support vector Machine)':
                # calibrate probabilities that will be used by the excel sheet
                if len(train_target) > 500:
                    model_isotonic_calibration = CalibratedClassifierCV(model, cv="prefit",
                                                                        method='isotonic')
                    model_isotonic_calibration.fit(train_tfidf.todense(), train_target)

                    model_isotonic_calibrations = pickle.dumps(model_isotonic_calibration)
                    s3.put_object(Body=model_isotonic_calibrations, Bucket=AWS_STORAGE_BUCKET_NAME,
                                  Key=key_value + 'model_isotonic_calibration.pkl')

                    model_prob = model_isotonic_calibration.predict_proba(test_data_tfidf.todense())
                    model_prob_all_training = model_isotonic_calibration.predict_proba(
                        train_tfidf.todense())
                else:
                    model_sigmoid_calibration = CalibratedClassifierCV(model, cv="prefit",
                                                                       method='sigmoid')

                    model_sigmoid_calibrations = pickle.dumps(model_sigmoid_calibration)
                    s3.put_object(Body=model_sigmoid_calibrations, Bucket=AWS_STORAGE_BUCKET_NAME,
                                  Key=key_value + 'model_sigmoid_calibration.pkl')

                    model_sigmoid_calibration.fit(train_tfidf.todense(), train_target)
                    model_prob = model_sigmoid_calibration.predict_proba(test_data_tfidf.todense())
                    model_prob_all_training = model_sigmoid_calibration.predict_proba(
                        train_tfidf.todense())
            else:
                if 'Deep Learning' in str_model_name:
                    model_prob = model.predict_proba(testing_data)
                    model_prob_all_training = model.predict_proba(training_data)
                else:
                    model_prob = model.predict_proba(test_data_tfidf.todense())
                    model_prob_all_training = model.predict_proba(train_tfidf.todense())

            # classify the patent with rel. threshold < th to "others" class
            model_prob_all = copy.copy(model_prob)
            model_prob[model_prob < relevanceThreshold] = 0.0
            model_prob_new = np.sum(model_prob, axis=1)
            model_predicted[model_prob_new == 0] = 'others'

            testingDataNumInstances = len(model_predicted)

            es_conn.update(index=index_name_gloabals, id=id,
                           body={
                               "doc": {"testingDataNumInstances": testingDataNumInstances,
                                       "model_prob_all_training": json.dumps(model_prob_all_training.tolist()),
                                       "model_prob_all": json.dumps(model_prob_all.tolist())}})

            # Update the treeview with the distribution of instances in the training data
            testingDataStatistics = []
            for label in set(model_predicted):
                distribution = str(
                    np.round(model_predicted.tolist().count(label) * 100.0 / len(model_predicted) * 1.0, 2)) + '%'
                testingDataStatistics.append(
                    [label, model_predicted.tolist().count(label), distribution])

            progressbar_value += 10

            progress_text = progress_text + "\nSaving results..."

            excel_status_code = 200

            testLabels = labels

            progressbar_value += 10

            programRunEndTime = datetime.now()

            timeDifference = relativedelta(programRunEndTime, programRunStartTime)

            programRunStartTimeLabel = "Program run took %d days %d hours %d minutes %d seconds." % (
                timeDifference.days, timeDifference.hours, timeDifference.minutes, timeDifference.seconds)
            progressbarlabel_text = programRunStartTimeLabel

            # textboxProgress.config(state=NORMAL)
            progress_text = progress_text + \
                            "\nPatent classification completed!\nPlease check all the results in the file: " + re.sub(
                '.txt', '_Threshold_Analysis.xlsx', gui_parameters['testing_file_name']) + \
                            "\n" + programRunStartTimeLabel

            progressbar_value += 5

            final_progress_value = 200
            es_conn.update(index=index_name_gloabals, id=id,
                           body={
                               "doc": {"testingDataStatistics": json.dumps(testingDataStatistics),
                                       'progressbar_value': progressbar_value,
                                       "progress_text": progress_text, "excel_status_code": excel_status_code,
                                       "final_progress_value": final_progress_value,
                                       "progressbarlabel_text": progressbarlabel_text}})
            if testingDataType == 'Patent':

                # Need to be clarified
                Prediction_Results_patent_FileName = re.sub('.txt', '_Prediction_Results.txt', testingFileName);
                # with open(Prediction_Results_patent_FileName, 'w') as fout:
                fout = ''
                for doc, model_i in zip(file_test_proc, model_predicted):
                    doc = doc.strip('\r').strip('\n')
                    doc = doc.replace('"', ' ')
                    fout += '\t'.join([doc, model_i.strip('\t').strip('\r').strip('\n')]) + '\n'
                    if model_i == 'others':
                        file_others.append(doc)
                        responsesupervised = HttpResponse(file_others, content_type='text/plain')
                        responsesupervised['Content-Disposition'] = 'attachment; filename=' + Prediction_Results_patent_FileName
                    else:
                        file_classified.append(doc)
                        responsesupervised = HttpResponse(file_classified, content_type='text/plain')
                        responsesupervised['Content-Disposition'] = 'attachment; filename=' + Prediction_Results_patent_FileName
                        model_classified.append(model_i)

            elif testingDataType == 'Journal':
                # Need to be clarified
                Prediction_Results_journal_FileName = re.sub('.txt', '_Prediction_Results.txt',
                                                             testingFileName);
                fout = ''
                # with open(Prediction_Results_journal_FileName, 'w') as fout:
                for doc, model_i in zip(file_test_proc, model_predicted):
                    doc = doc.strip('\r').strip('\n')
                    doc = doc.replace('"', ' ')
                    fout += '\t'.join([doc, model_i.strip('\t').strip('\r').strip('\n')]) + '\n'

                    if model_i == 'others':
                        file_others.append(doc)
                        responsesupervised = HttpResponse(file_others, content_type='text/plain')
                        responsesupervised['Content-Disposition'] = 'attachment; filename=' + Prediction_Results_journal_FileName
                    else:
                        file_classified.append(doc)
                        responsesupervised = HttpResponse(file_classified, content_type='text/plain')
                        responsesupervised['Content-Disposition'] = 'attachment; filename=' + Prediction_Results_journal_FileName
                        model_classified.append(model_i)
                # model_predicted_index = [labels.index(cls) for cls in model_predicted.tolist()]
                # model_classified_index = [labels.index(cls) for cls in model_classified]
        return HttpResponse('successfully executed')

    except Exception as e:
        errorString = 'Error while running the program please contact IP Group Analytics team.'
        final_progress_value = 200
        es_conn.update(index=index_name_gloabals, id=id,
                       body={
                           "doc": {"final_progress_value": final_progress_value,
                                   "errorString": errorString}})
        print(sys.exc_info())
        return HttpResponse("Error running the program {}".format(sys.exc_info()[-1].tb_lineno) + str(e))


def preprocess_collection(file_open, stopterms, printInfo, progress_text):
    # apollo4.globals.progress_text
    master = []  # list for used application numbers
    repeat = []  # list for duplicate application numbers
    file_temp = []  # updated collection file
    counter = 0
    num_docs = len(file_open)
    for index, doc in enumerate(file_open, start=1):
        try:
            apn = doc.split("\t")
            apn = apn[4].upper()
            if apn not in master:
                file_temp.append(doc)
                master.append(apn)
                counter = counter + 1
            elif apn in master:
                repeat.append(apn)
        except Exception as e:
            progress_text = progress_text + "*" * 50 + "\n" + "ERROR: The document number %d in the file could not be processed" % index + "\n" + "-" * 50
            continue

    # step through collection docs (ie lines)
    file_proc = []
    file_stem = []
    file_stop_words_removed = []

    design_count = 0  # counter for design cases
    utility_count = 0  # counter for utility cases

    for file_index, file_line in enumerate(file_temp, start=1):
        file_line = file_line.split("\t")  # split by tab
        # take correct col number docs only
        try:
            no = str(file_index)
            file_t = file_line[1]  # title
            file_a = file_line[2]  # abstract
            file_c = file_line[3]  # claims
            apn = file_line[4].lower()
            apd = file_line[5]
            asgn = file_line[6].lower()
            if len(file_line) > 7:
                upc = file_line[7].lower()
            if len(file_line) > 8:
                label = file_line[8].lower()  # solve the issue if label has tab
        except Exception as e:
            progress_text = progress_text + "*" * 50 + "\n" + "\n" + "ERROR: The document number %d in the file could not be processed" % file_index + "\n" + "-" * 50

        if apn.startswith("us2"):
            # filter out design cases
            progress_text = progress_text + "*" * 50 + "\n" + "Design patent found! App_No: %r\tUPC: %r" % (
                apn, upc) + '\n' + "-" * 50
            design_count = design_count + 1

        elif apn.startswith("us"):
            # filter out non-apn lines (ie not patent data)
            utility_count = utility_count + 1

            # stop and stem title, abstract, claim
            file_t_stem = stop_and_stem(file_t, stopterms)
            file_a_stem = stop_and_stem(file_a, stopterms)
            file_c_stem = stop_and_stem(file_c, stopterms)

            # remove stopwords from the title, abstract, claim
            file_t_stop = remove_stopwords(file_t, stopterms)
            file_a_stop = remove_stopwords(file_a, stopterms)
            file_c_stop = remove_stopwords(file_c, stopterms)

            # Output the orginal clean version of utility patent
            file_new_line = '\t'.join(file_line)
            file_proc.append(file_new_line)

            # Output the preprocessed version of utility patent
            if len(file_line) > 7:
                proc_doc = [no, file_t_stem, file_a_stem, file_c_stem,
                            apd, apn, asgn, upc]
                proc_doc_stop = [no, file_t_stop, file_a_stop, file_c_stop,
                                 apd, apn, asgn, upc]

            else:
                proc_doc = [no, file_t_stem, file_a_stem, file_c_stem,
                            apd, apn, asgn]
                proc_doc_stop = [no, file_t_stop, file_a_stop, file_c_stop,
                                 apd, apn, asgn]

            if len(file_line) > 8:  # solve the issue if label has tab
                proc_doc.append(label)
                proc_doc_stop.append(label)

            proc_doc = '\t'.join(proc_doc)
            proc_doc_stop = '\t'.join(proc_doc_stop)

            file_stem.append(proc_doc)
            file_stop_words_removed.append(proc_doc_stop)

    output = (file_proc, file_stem, file_stop_words_removed)
    return output


def dedup_collection_journal(file_open, uid, abstract_id):
    try:
        new_file_list = []
        new_item_list = []
        for doc in file_open:
            item = '\t'.join(doc.split('\t')[uid: abstract_id + 1])
            if item not in new_item_list:
                new_item_list.append(item)
                new_file_list.append(doc)
        return new_file_list
    except Exception as e:
        return HttpResponse(
            "Error running the program.")


def preprocess_collection_journal(file_open):
    # Sentence Level Preprocessing
    try:
        file_proc = []
        for file_index, file_line in enumerate(file_open, start=1):
            doc = file_line.split('\t')
            abstract = doc[2]
            # remove copyright info
            try:
                copyright_str = [u'\ufffd', 'copyright', 'ieee', 'acm', 'rights reserved', 'press',
                                 'all rights reserved']
                sent_list = list(filter(None, abstract.split('. ')))
                new_sent_list = [sent for sent in sent_list
                                 if all(x not in sent.decode('utf-8').lower() for x in copyright_str)]
                new_abstract = '. '.join(new_sent_list)
                doc[2] = new_abstract
            except:
                pass
            file_proc.append('\t'.join(doc))
        return file_proc
    except Exception as e:
        return HttpResponse(
            "Error running the program.")


def stop_and_stem_journal(file_t, user_defined_stopwords):
    '''
    Input: file_t: a text string, stopterms: a dictionary of stop terms
    Output: file_stem: a list of stopped and stemmed terms
    '''
    try:
        stopterms = build_stopterms_journal(user_defined_stopwords)
        # remove the patent specific terms
        file_t = file_t.lower()  # lowercase all
        file_t = re.sub("[^a-zA-Z0-9 ]", "   ", file_t)  # remove non-alphanumeric
        file_t = re.sub("\s[0-9]+$", '', file_t)
        file_t = re.sub("\s[0-9]+\s", ' ', file_t)
        file_t = re.sub("^[0-9]+\s", '', file_t)
        file_t = re.sub("androids*", "antroid", file_t)
        file_t = re.sub("andes", "antes", file_t)
        file_t = re.sub("and[0-9a-z]+", "", file_t)
        file_t = re.sub("antroid", "android", file_t)
        file_t = re.sub("antes", "andes", file_t)
        file_t = re.sub("including[0-9a-z]+", "", file_t)
        file_t = re.sub("wherein[0-9a-z]+", "", file_t)
        file_t = re.sub("comprising[0-9a-z]+", "", file_t)
        formula_chk0 = re.findall(" formula | math ", file_t)
        formula_chk1 = re.findall(" tail ", file_t)
        formula_chk2 = re.findall(" lead ", file_t)
        if len(formula_chk0) > 0 and len(formula_chk1) > 0 and len(formula_chk2) > 0:
            file_t = re.sub(" formula | math ", " ", file_t)
            file_t = re.sub(" tail ", " ", file_t)
            file_t = re.sub(" lead ", " ", file_t)
        file_t = " ".join(file_t.split())  # split by any whitespace and rejoin w/ space
        file_t = file_t.split(" ")  # split by space

        # remove the stop terms in the text
        file_stop = []  # initialize list
        for term in file_t:
            if term not in stopterms:
                file_stop.append(term)

        # stem using porter algorithm
        file_stem = []  # initialize list
        for term in file_stop:
            try:
                term = wn().lemmatize(term)
            except:
                pass
            term = str(term)
            file_stem.append(term)
        file_stem = ' '.join(file_stem)
        return file_stem
    except Exception as e:
        return HttpResponse(
            "Error running the program.")


def build_stopterms_journal(user_defined_stopwords):
    # Build stop terms
    try:
        stops = stopwords

        aux_stops = './static/AuxStops-Journal.txt'

        aux_stops = open(aux_stops, 'r').read()
        aux_stops = re.sub("[^a-zA-Z ]", " ", aux_stops)  # remove non-alphanumeric
        aux_stops = (" ".join(aux_stops.split())).split()  # remove white space
        aux_stops = list(filter(None, aux_stops))
        stops = stops + aux_stops

        if len(user_defined_stopwords) > 0:
            user_defined_stops = (" ".join(user_defined_stopwords.split(','))).split()
            # Add user defined stop words to the stop words list
            stops = stops + user_defined_stops

            # Bulid stopterm dictionary
        stopterms = {}
        for stop in stops:
            if stop in stopterms:
                stopterms[stop] += 1
            else:
                stopterms[stop] = 1
        return stopterms
    except Exception as e:
        return HttpResponse(
            "Error running the program. ")


def stop_and_stem(file_t, stopterms):
    '''
    Input: file_t: a text string, stopterms: a dictionary of stop terms
    Output: file_stem: a list of stopped and stemmed terms
    '''
    # remove the patent specific terms
    try:
        file_t = file_t.lower()  # lowercase all
        file_t = re.sub("[^a-zA-Z0-9 ]", "   ", file_t)  # remove non-alphanumeric
        file_t = re.sub("\s[0-9]+$", '', file_t)
        file_t = re.sub("\s[0-9]+\s", ' ', file_t)
        file_t = re.sub("^[0-9]+\s", '', file_t)
        file_t = re.sub("androids*", "antroid", file_t)
        file_t = re.sub("andes", "antes", file_t)
        file_t = re.sub("and[0-9a-z]+", "", file_t)
        file_t = re.sub("antroid", "android", file_t)
        file_t = re.sub("antes", "andes", file_t)
        file_t = re.sub("including[0-9a-z]+", "", file_t)
        file_t = re.sub("wherein[0-9a-z]+", "", file_t)
        file_t = re.sub("comprising[0-9a-z]+", "", file_t)
        formula_chk0 = re.findall(" formula | math ", file_t)
        formula_chk1 = re.findall(" tail ", file_t)
        formula_chk2 = re.findall(" lead ", file_t)
        if len(formula_chk0) > 0 and len(formula_chk1) > 0 and len(formula_chk2) > 0:
            file_t = re.sub(" formula | math ", " ", file_t)
            file_t = re.sub(" tail ", " ", file_t)
            file_t = re.sub(" lead ", " ", file_t)
        file_t = " ".join(file_t.split())  # split by any whitespace and rejoin w/ space
        file_t = file_t.split(" ")  # split by space

        # remove the stop terms in the text
        file_stop = []  # initialize list
        for term in file_t:
            if term not in stopterms:
                file_stop.append(term)

        # stem using porter algorithm
        file_stem = []  # initialize list
        for term in file_stop:
            try:
                term = wn().lemmatize(term)
            except:
                pass
            term = str(term)
            file_stem.append(term)
        file_stem = ' '.join(file_stem)
        return file_stem
    except Exception as e:
        return HttpResponse(
            "Error running the program.")


def remove_stopwords(file_t, stopterms):
    '''
    Input: file_t: a text string, stopterms: a dictionary of stop terms
    Output: file_stem: a list of stopped and stemmed terms
    '''
    # remove the patent specific terms
    try:
        file_t = file_t.lower()  # lowercase all
        file_t = re.sub("[^a-zA-Z0-9 ]", "   ", file_t)  # remove non-alphanumeric
        file_t = re.sub("\s[0-9]+$", '', file_t)
        file_t = re.sub("\s[0-9]+\s", ' ', file_t)
        file_t = re.sub("^[0-9]+\s", '', file_t)
        file_t = re.sub("androids*", "antroid", file_t)
        file_t = re.sub("andes", "antes", file_t)
        file_t = re.sub("and[0-9a-z]+", "", file_t)
        file_t = re.sub("antroid", "android", file_t)
        file_t = re.sub("antes", "andes", file_t)
        file_t = re.sub("including[0-9a-z]+", "", file_t)
        file_t = re.sub("wherein[0-9a-z]+", "", file_t)
        file_t = re.sub("comprising[0-9a-z]+", "", file_t)
        formula_chk0 = re.findall(" formula | math ", file_t)
        formula_chk1 = re.findall(" tail ", file_t)
        formula_chk2 = re.findall(" lead ", file_t)
        if len(formula_chk0) > 0 and len(formula_chk1) > 0 and len(formula_chk2) > 0:
            file_t = re.sub(" formula | math ", " ", file_t)
            file_t = re.sub(" tail ", " ", file_t)
            file_t = re.sub(" lead ", " ", file_t)
        file_t = " ".join(file_t.split())  # split by any whitespace and rejoin w/ space
        file_t = file_t.split(" ")  # split by space

        # remove the stop terms in the text
        file_stop = []  # initialize list
        for term in file_t:
            if term not in stopterms:
                file_stop.append(term)
        file_stop = ' '.join(file_stop)
        return file_stop
    except Exception as e:
        return HttpResponse(
            "Error running the program.")


def remove_stopwords_journal(file_t, user_defined_stopwords):
    '''
    Input: file_t: a text string, stopterms: a dictionary of stop terms
    Output: file_stem: a list of stopped and stemmed terms
    '''
    try:
        stopterms = build_stopterms_journal(user_defined_stopwords)
        # remove the patent specific terms
        file_t = file_t.lower()  # lowercase all
        file_t = re.sub("[^a-zA-Z0-9 ]", "   ", file_t)  # remove non-alphanumeric
        file_t = re.sub("\s[0-9]+$", '', file_t)
        file_t = re.sub("\s[0-9]+\s", ' ', file_t)
        file_t = re.sub("^[0-9]+\s", '', file_t)
        file_t = re.sub("androids*", "antroid", file_t)
        file_t = re.sub("andes", "antes", file_t)
        file_t = re.sub("and[0-9a-z]+", "", file_t)
        file_t = re.sub("antroid", "android", file_t)
        file_t = re.sub("antes", "andes", file_t)
        file_t = re.sub("including[0-9a-z]+", "", file_t)
        file_t = re.sub("wherein[0-9a-z]+", "", file_t)
        file_t = re.sub("comprising[0-9a-z]+", "", file_t)
        formula_chk0 = re.findall(" formula | math ", file_t)
        formula_chk1 = re.findall(" tail ", file_t)
        formula_chk2 = re.findall(" lead ", file_t)
        if len(formula_chk0) > 0 and len(formula_chk1) > 0 and len(formula_chk2) > 0:
            file_t = re.sub(" formula | math ", " ", file_t)
            file_t = re.sub(" tail ", " ", file_t)
            file_t = re.sub(" lead ", " ", file_t)
        file_t = " ".join(file_t.split())  # split by any whitespace and rejoin w/ space
        file_t = file_t.split(" ")  # split by space

        # remove the stop terms in the text
        file_stop = []  # initialize list
        for term in file_t:
            if term not in stopterms:
                file_stop.append(term)

        return ' '.join(file_stop)
    except Exception as e:
        return HttpResponse(
            "Error running the program. ")


def get_topic_list(model, feature_names, n_top_words):
    try:
        topic_list = []
        n_top_words = int(n_top_words)
        for topic_idx, topic in enumerate(model.components_):
            topic_list.append(" | ".join([feature_names[i] for i in topic.argsort()[:-n_top_words - 1:-1]]))
        return topic_list
    except Exception as e:
        return HttpResponse(
            "Error running the program.")


############# runDocumentClassifierSupervised ends#########################

############# saving and retreiving model starts#########################

@csrf_exempt
def save_both_existing_model(request):
    modelSaveStartTime = datetime.now()
    s3 = boto3.client('s3')
    INDEX_NAME = 'savemodelsupervised'
    trainingDataPatents = 'training_data_patents.txt'
    trainingDataJournals = 'training_data_journals.txt'
    trainedModel_file = "trainedModel.pkl"
    topicModelingFrame_file = 'topicModelingFrame.pkl'
    tfidf_vect_file = "tfidf_vect.pkl"
    user_defined_stopwords_file = "user_defined_stopwords.pkl"

    try:
        if request.method == 'POST':
            username = request.user.username;
            save_model_Name = request.POST.getlist('input')[0]
            saveName = json.loads(save_model_Name);

            if saveName['existingSaveModel']:
                saveProjectName = saveName['exisitingProjectName']
                save_project_description = saveName['exisitingProjectDescription']
                saveProjectValidation = saveName['saveProject']
                saveModelDesc = saveName['existingModelDesc']
                targetPerformanceMeasure = saveName['target_performance_measure']
                modelName = saveName['trainedModelName']
            elif saveName['newSaveModel']:
                saveProjectName = saveName['newProjectName']
                save_project_description = saveName['newProjectDescription']
                saveProjectValidation = saveName['saveProject']
                saveModelDesc = saveName['newModelDesc']
                targetPerformanceMeasure = saveName['target_performance_measure']
                modelName = saveName['trainedModelName']

            # index_name_patent = 'patentorjournal'
            # query = {"query": {"bool": {"must": {"match": {"username.keyword": username}}}}}
            # res = es_conn.search(index=index_name_patent, body=query)
            # patentOrJournalTrainingData = res['hits']['hits'][0]['_source']['patentOrJournalTrainingData']

            if saveName['learningType'] == 'supervised':
                INDEX_NAME = 'savemodelsupervised'
                index_name_gloabals = 'apolloglobals'
                query = {"query": {"bool": {"must": {"match": {"username.keyword": username}}}}}
                res = es_conn.search(index=index_name_gloabals, body=query)
                id = res['hits']['hits'][0]['_id']
                data = res['hits']['hits'][0]['_source']
                user_defined_stopwords = data['user_defined_stopwords']
                trainingDataPerformances = data['trainingDataPerformance']
                trainingDataPerformancesStandardDeviation = data['trainingDataPerformancesStandardDeviations']
                trainingDataStatistics = data['trainingDataStatistics']
                trainingDataType = data['trainingDataType']
                trainedModelName = saveModelDesc + '_' + modelName
                model_data = {
                    'saved_date': datetime.now(),
                    'trainingDataNumInstances': data['trainingDataNumInstances'],  # Training examples
                    'trainingDataNumClasses': data['trainingDataNumClasses'],  # training classes
                    'trainingDataTables': trainingDataStatistics,  # trainig data table
                    'trainingDataPerformances': trainingDataPerformances,  # model evaluation before +/-
                    'trainingDataPerformancesStandardDeviation': trainingDataPerformancesStandardDeviation,
                    # model evaluation after +/-
                    'str_parameter_name': data['str_parameter_name'],  # Hyper parameter left before =
                    'trainedModelName': trainedModelName,
                    'optimal_model_parameter': data['optimal_model_parameter'],  # Hyper parameter right after =
                    'learningType': 'supervised'
                }

            elif saveName['learningType'] == 'unsupervised':
                INDEX_NAME = 'savemodelunsupervised'
                index_name_gloabals = 'apolloglobalsunsupervised'
                query = {"query": {"bool": {"must": {"match": {"username.keyword": username}}}}}
                res = es_conn.search(index=index_name_gloabals, body=query)
                id = res['hits']['hits'][0]['_id']
                data = res['hits']['hits'][0]['_source']
                user_defined_stopwords = data['user_defined_stopwords']
                clusterTopicsAndCounts = data['clusterTopicsAndCounts']
                numberOfClusters = data['numberOfClusters']
                number_of_top_words = data['number_of_top_words']
                testingDataType = data['training_data_type']
                trainedModelName = saveModelDesc + '_K_Means_Clustering_NumClusters=' + str(
                    numberOfClusters) + '_TopWords=' + str(
                    number_of_top_words)

                model_data = {
                    'saved_date': datetime.now(),
                    'clusterTopicsAndCounts': json.dumps(clusterTopicsAndCounts),
                    'trainedModelName': trainedModelName,
                    'learningType': 'unsupervised'
                }

            save_model_data = {
                'username': username,
                'saveProjectName': saveProjectName,
                'save_project_description': save_project_description,
                'model_data': model_data,

            }

            # save_response = es_conn.indices.create(index=INDEX_NAME, ignore=400)

            if saveName['newSaveModel']:
                if saveProjectValidation:
                    save_response = es_conn.create(index=INDEX_NAME, doc_type=TYPE_NAME_USER, body=save_model_data,
                                                   id=uuid.uuid4())
            elif saveName['existingSaveModel']:
                # this is for the append of model to the existing projectName
                if saveProjectValidation:
                    save_response = es_conn.create(index=INDEX_NAME, doc_type=TYPE_NAME_USER, body=save_model_data,
                                                   id=uuid.uuid4())
                    # this is for the replace of model to the existing projectName
                # else:
                # res = es_conn.update(index=INDEX_NAME, id=hits_id, body={"doc": {"model_data": model_data}})

                # res = es_conn.delete_by_query(index=INDEX_NAME,body=query
            save_response_validation = save_response['_shards']['successful']
            if save_response_validation > 0:
                es_conn.update(index=index_name_gloabals, id=id,
                               body={"doc": {"saved_project_status": 200}})

            append_text_to_history_file = "\n"
            user_defined_stopwords = pickle.dumps(user_defined_stopwords)
            if saveName['learningType'] == 'supervised':
                finalTrainingData = request.FILES.getlist('trainFile')
                training_data = Process_All_Files(finalTrainingData)
                key_value = 'runDocumentClassifier/'
                key_value += username + '/'
                tfidf_vectorizer = s3.get_object(Bucket=AWS_STORAGE_BUCKET_NAME,
                                                 Key=key_value + 'tfidf_vect.pkl')

                tfidf_vect = tfidf_vectorizer['Body'].read()
                trainedModel = s3.get_object(Bucket=AWS_STORAGE_BUCKET_NAME,
                                             Key=key_value + 'trainedModel.pkl')
                trainedModel = trainedModel['Body'].read()

                key_value = 'classificationprojects/'
                key_value += saveProjectName + '/supervised/'
                key_value += trainedModelName + '/'

                programRunStartTime = datetime.now()
                programRunStartTimeLabel = 'Progress: Program run started at ' + programRunStartTime.strftime(
                    "%I:%M%p on %B %d, %Y") + ' (UTC time). '

                append_text_to_history_file += '-' * 100 + '\n'
                append_text_to_history_file += 'username: ' + username + '\n'
                append_text_to_history_file += "Program run started at " + programRunStartTime.strftime(
                    "%I:%M%p on %B %d, %Y") + " (UTC time).\n"
                append_text_to_history_file += '-' * 100
                append_text_to_history_file += '\nProject name: ' + saveProjectName + '\n'
                append_text_to_history_file += 'Project description: ' + save_project_description + '\n'
                #
                if trainingDataType == 'Patent':
                    append_text_to_history_file += 'Supervised learning model ' + trainedModelName + ' was trained on the PATENT training data file \n'
                elif trainingDataType == 'Journal':
                    append_text_to_history_file += 'Supervised learning model ' + trainedModelName + ' was trained on the JOURNAL training data file \n'
                # write the number of instances and classes for tracking purposes
                numInstancesInTrainingData = 0
                stringToDisplayTrainingDataStats = '{:<40s}{:>20s}{:>20s}'.format('Class', '# Examples',
                                                                                  'Class %') + '\n'

                for entry in json.loads(trainingDataStatistics):
                    stringToDisplayTrainingDataStats += '{:<40s}{:>20s}{:>20s}'.format(str(entry[0]), str(entry[1]),
                                                                                       str(entry[2])) + '\n'
                    numInstancesInTrainingData += int(entry[1])

                append_text_to_history_file += 'Total number of documents in the training data: ' + str(
                    numInstancesInTrainingData) + '\n'
                append_text_to_history_file += 'Total number of classes in the training data: ' + str(
                    len(trainingDataStatistics)) + '\n'
                append_text_to_history_file += 'The model parameters were optimized for \'' + targetPerformanceMeasure + '\'.' + '\n'
                append_text_to_history_file += '5-fold Cross Validation Performance: ' + '\n'

                perfMeasuresStr = ['Accuracy:', 'AUC:', 'Precision:', 'Recall:', 'F1:']
                for i in range(len(trainingDataPerformances)):
                    stringToWrite = '{:<10s}{:>10.2f}{:>4s}{:>10.2f}{:>1s}'.format(perfMeasuresStr[i],
                                                                                   trainingDataPerformances[
                                                                                       i] * 100.0, '% +/- ',
                                                                                   trainingDataPerformancesStandardDeviation[
                                                                                       i] * 100.0, '%')
                    append_text_to_history_file += stringToWrite + '\n'

                append_text_to_history_file += programRunStartTimeLabel + '.' + '\n'

                s3.put_object(Body=tfidf_vect, Bucket=AWS_STORAGE_BUCKET_NAME,
                              Key=key_value + tfidf_vect_file)
                s3.put_object(Body=user_defined_stopwords, Bucket=AWS_STORAGE_BUCKET_NAME,
                              Key=key_value + user_defined_stopwords_file)
                s3.put_object(Body=trainedModel, Bucket=AWS_STORAGE_BUCKET_NAME,
                              Key=key_value + trainedModel_file)
                if trainingDataType == 'Patent':
                    s3.put_object(Body=training_data, Bucket=AWS_STORAGE_BUCKET_NAME,
                                  Key=key_value + trainingDataPatents)
                elif trainingDataType == 'Journal':
                    s3.put_object(Body=training_data, Bucket=AWS_STORAGE_BUCKET_NAME,
                                  Key=key_value + trainingDataJournals)
                if 'Support_Vector_Machine' in trainedModelName:
                    try:
                        model_sigmoid_calibrations = s3.get_object(Bucket=AWS_STORAGE_BUCKET_NAME,
                                                                   Key=key_value + 'model_sigmoid_calibration.pkl')

                        model_sigmoid_calibration = model_sigmoid_calibrations['Body'].read()
                        model_sigmoid_calibration = pickle.loads(model_sigmoid_calibration)

                    except botocore.exceptions.ClientError as e:
                        if e.response['Error']['Code'] == "404":
                            # There is no existing patent training data, so no need to deduplicate based on previous data.
                            model_isotonic_calibrations = s3.get_object(Bucket=AWS_STORAGE_BUCKET_NAME,
                                                                        Key=key_value + 'model_isotonic_calibration.pkl')

                            model_isotonic_calibration = model_isotonic_calibrations['Body'].read()
                            model_sigmoid_calibration = pickle.loads(model_isotonic_calibration)
                        else:
                            model_isotonic_calibrations = s3.get_object(Bucket=AWS_STORAGE_BUCKET_NAME,
                                                                        Key=key_value + 'model_isotonic_calibration.pkl')

                            model_isotonic_calibration = model_isotonic_calibrations['Body'].read()
                            model_sigmoid_calibration = pickle.loads(model_isotonic_calibration)
                            pass
                    if model_isotonic_calibration != None:
                        model_isotonic_calibration = pickle.dumps(model_isotonic_calibration)
                        s3.put_object(Body=model_isotonic_calibration, Bucket=AWS_STORAGE_BUCKET_NAME,
                                      Key=key_value + 'model_isotonic_calibration.pkl')
                    if model_sigmoid_calibration != None:
                        model_sigmoid_calibration = pickle.dumps(model_sigmoid_calibration)
                        s3.put_object(Body=model_sigmoid_calibration, Bucket=AWS_STORAGE_BUCKET_NAME,
                                      Key=key_value + 'model_sigmoid_calibration.pkl')

            elif saveName['learningType'] == 'unsupervised':
                finalTrainingData = request.FILES.getlist('trainFile')
                training_data = Process_All_Files(finalTrainingData)
                key_value = 'classificationprojects/'
                key_value += saveProjectName + '/unsupervised/'
                key_value += trainedModelName + '/'
                programRunStartTime = datetime.now()
                programRunStartTimeLabel = 'Progress: Program run started at ' + programRunStartTime.strftime(
                    "%I:%M%p on %B %d, %Y") + '. '

                append_text_to_history_file += '-' * 100 + '\n'
                append_text_to_history_file += 'username: ' + username + '\n'
                append_text_to_history_file += "Program run started at " + programRunStartTime.strftime(
                    "%I:%M%p on %B %d, %Y") + " (UTC time).\n"
                append_text_to_history_file += '-' * 100
                append_text_to_history_file += '\nProject name: ' + saveProjectName + '\n'
                append_text_to_history_file += 'Project description: ' + save_project_description + '\n'
                #
                if testingDataType == 'Patent':
                    append_text_to_history_file += 'Unsupervised learning model ' + trainedModelName + ' was trained on the PATENT training data file \n'
                elif testingDataType == 'Journal':
                    append_text_to_history_file += 'Unsupervised learning model ' + trainedModelName + ' was trained on the JOURNAL training data file \n'
                # write the number of instances and classes for tracking purposes
                append_text_to_history_file += programRunStartTimeLabel + '.' + '\n'
                key_value = 'runDocumentClassifierUnsupervised/'
                key_value += username + '/'
                tfidf_vectorizer = s3.get_object(Bucket=AWS_STORAGE_BUCKET_NAME,
                                                 Key=key_value + 'tfidf_vect.pkl')

                tfidf_vect = tfidf_vectorizer['Body'].read()
                trainedModel = s3.get_object(Bucket=AWS_STORAGE_BUCKET_NAME,
                                             Key=key_value + 'trainedModel.pkl')

                trainedModel = trainedModel['Body'].read()
                topicModelingFrame = s3.get_object(Bucket=AWS_STORAGE_BUCKET_NAME,
                                                   Key=key_value + 'topicModelingFrame.pkl')

                topicModelingFrame = topicModelingFrame['Body'].read()
                key_value = 'classificationprojects/'
                key_value += saveProjectName + '/unsupervised/'
                key_value += trainedModelName + '/'
                s3.put_object(Body=tfidf_vect, Bucket=AWS_STORAGE_BUCKET_NAME,
                              Key=key_value + tfidf_vect_file)
                s3.put_object(Body=user_defined_stopwords, Bucket=AWS_STORAGE_BUCKET_NAME,
                              Key=key_value + user_defined_stopwords_file)
                s3.put_object(Body=trainedModel, Bucket=AWS_STORAGE_BUCKET_NAME,
                              Key=key_value + trainedModel_file)
                s3.put_object(Body=topicModelingFrame, Bucket=AWS_STORAGE_BUCKET_NAME,
                              Key=key_value + topicModelingFrame_file)
                if testingDataType == 'Patent':
                    s3.put_object(Body=training_data, Bucket=AWS_STORAGE_BUCKET_NAME,
                                  Key=key_value + trainingDataPatents)
                elif testingDataType == 'Journal':
                    s3.put_object(Body=training_data, Bucket=AWS_STORAGE_BUCKET_NAME,
                                  Key=key_value + trainingDataJournals)

            modelSaveEndTime = datetime.now()
            timeDifference = relativedelta(modelSaveEndTime, modelSaveStartTime)
            modelSavingTimeLabel = "Saving the model took %d hours %d minutes %d seconds." % (
                timeDifference.hours, timeDifference.minutes, timeDifference.seconds)
            append_text_to_history_file += modelSavingTimeLabel + '\n' + '*' * 95 + '\n'

            try:
                response2 = s3.get_object(Bucket=AWS_STORAGE_BUCKET_NAME,
                                          Key=key_value + 'history.txt')
                history_file_old_text = response2['Body'].read().decode('utf-8')
                append_text_to_history_file = history_file_old_text + append_text_to_history_file

                s3.put_object(Body=append_text_to_history_file, Bucket=AWS_STORAGE_BUCKET_NAME,
                              Key=key_value + 'history.txt')

            except botocore.exceptions.ClientError as e:
                if e.response['Error']['Code'] == "404":
                    # There is no existing history file, so create a new history file and write the history into that file in S3.
                    s3.put_object(Body=append_text_to_history_file, Bucket=AWS_STORAGE_BUCKET_NAME,
                                  Key=key_value + 'history.txt')
                elif e.response['Error']['Code'] == "NoSuchKey":
                    pass
                else:
                    pass
            res = {
                'message': 'saved successfully'
            }
        return JsonResponse(res)
    except Exception as e:
        return HttpResponse(
            "Error running the program.")


@csrf_exempt
def save_both_validation(request):
    try:
        if request.method == 'POST':
            userName = request.user.username;
            save_model_Name = request.body.decode('utf-8');
            saveName = json.loads(save_model_Name);
            if saveName['existingSaveModel']:
                saveProjectName = saveName['exisitingProjectName']
                saveModelDesc = saveName['existingModelDesc']
                trainedModelName = saveName['trainedModelName']
            elif saveName['newSaveModel']:
                saveProjectName = saveName['newProjectName']
                saveModelDesc = saveName['newModelDesc']
                trainedModelName = saveName['trainedModelName']
            if saveName['learningType'] == 'supervised':
                INDEX_NAME = 'savemodelsupervised'
            elif saveName['learningType'] == 'unsupervised':
                INDEX_NAME = 'savemodelunsupervised'
            if saveName['newSaveModel']:
                query = {"query": {"bool": {"must": {"match": {"saveProjectName.keyword": saveProjectName}}}}}
                res = es_conn.search(index=INDEX_NAME, body=query)
                res_hits = res['hits']
                res_hits_hits = res_hits['hits']
                hits_length = len(res_hits_hits)
                if hits_length == 0:
                    saveProject = True
                else:
                    saveProject = False
                    # res = es_conn.delete_by_query(index=INDEX_NAME,body=query
            elif saveName['existingSaveModel']:
                modelName = saveModelDesc + '_' + trainedModelName
                query = {"query": {
                    "bool": {
                        "must": [{"match": {"model_data.trainedModelName.keyword": modelName}},
                                 {"match": {"saveProjectName.keyword": saveProjectName}}]}}}
                res = es_conn.search(index=INDEX_NAME, body=query)
                res_hits = res['hits']
                res_hits_hits = res_hits['hits']
                hits_length = len(res_hits_hits)
                model_data_count = 0
                # this is for the append of model to the existing projectName
                if hits_length == 0:
                    saveProject = True
                    # this is for the replace of model to the existing projectName
                else:
                    saveProject = False
                    # res = es_conn.delete_by_query(index=INDEX_NAME,body=query
        return HttpResponse(saveProject)
    except Exception as e:
        return HttpResponse(
            "Error running the program.")


@csrf_exempt
def retrieve_existing_Project_name(request):
    try:
        if request.method == 'POST':
            save_model_Name = request.body.decode('utf-8');
            saveName = json.loads(save_model_Name);
            if saveName['learningType'] == 'supervised':
                INDEX_NAME = 'savemodelsupervised'
            elif saveName['learningType'] == 'unsupervised':
                INDEX_NAME = 'savemodelunsupervised'

            query = {"query": {"match_all": {}}}
            res = es_conn.search(index=INDEX_NAME, body=query)
        return JsonResponse(res)
    except Exception as e:
        return HttpResponse(
            "Error running the program.")


@csrf_exempt
def retreieve_Model_for_seleted(request):
    try:
        if request.method == 'POST':
            save_model_Name = request.body.decode('utf-8');
            saveName = json.loads(save_model_Name);
            if saveName['learningType'] == 'supervised':
                saveProjectName = saveName['exisitingProjectName']
                INDEX_NAME = 'savemodelsupervised'
            elif saveName['learningType'] == 'unsupervised':
                saveProjectName = saveName['exisitingProjectName']
                INDEX_NAME = 'savemodelunsupervised'

            query = {"query": {"bool": {"must": {"match": {"saveProjectName.keyword": saveProjectName}}}}}
            res = es_conn.search(index=INDEX_NAME, body=query)
        return JsonResponse(res)
    except Exception as e:
        return HttpResponse(
            "Error running the program.")


############# saving and retreiving model ends#########################
@csrf_exempt
def runUnsupervisedSaving(request):
    try:
        if request.method == 'POST':
            userName = request.user.username;
            update_fields = {
                'username': userName,
                'progressbar_maximum': 0,  # progress bar max_value
                'progressbar_value': 0,  # progress bar value
                'progressbarlabel_text': '',  # progress field
                'progress_text': '',  # progress text
                'clusterTopicsAndCounts': '',
                'final_progress_value': '',
                'current_tab': 0,
                'errorString': '',
                'numberOfClusters': 0,
                'number_of_top_words': 0,
                "topicModelingFrame": '',
                "training_data_type": '',
                "testing_data_type": ''
            }
            index_name_gloabals = 'apolloglobalsunsupervised'
            query = {"query": {"bool": {"must": {"match": {"username.keyword": userName}}}}}
            # es_conn.indices.create(index=index_name_gloabals)
            es_conn.delete_by_query(index=index_name_gloabals, body=query)
            es_conn.create(index=index_name_gloabals, doc_type='_doc', body=update_fields, id=uuid.uuid4())
        return HttpResponse('sucess')
    except Exception as e:
        return HttpResponse(
            "Error running the program.")


@csrf_exempt
def userRunModelTrackUSL(request):
    try:
        if request.method == 'POST':
            try:
                userName = request.user.username;
                userTrack = request.body.decode('utf-8');
                userTrack = json.loads(userTrack);
                trainingFileName = userTrack['training_file_name']
                trainingDataType = userTrack['training_data_type']
                str_model_name = userTrack['model']
                additional_stopwords = userTrack['additional_stopwords']
                number_of_top_words = userTrack['number_of_top_words']
                number_of_clusters = userTrack['number_of_clusters']
            except Exception as e:
                print('parsing went wrong', e)

            time = datetime.now()
            time = time.strftime("%I:%M%p on %B %d, %Y")

            update_fields = {
                'username': userName,
                'trainingFileName': trainingFileName,
                'trainingDataType': trainingDataType,
                'str_model_name': str_model_name,
                'additional_stopwords': additional_stopwords,
                'number_of_top_words': number_of_top_words,
                'number_of_clusters': number_of_clusters,
                'time': time + ' UTC  Time'
            }
            index_name = 'userrunmodeltrackusl'
            if es_conn.indices.exists(index_name):
                es_conn.create(index=index_name, doc_type='_doc', body=update_fields, id=uuid.uuid4())
            else:
                save_response = es_conn.indices.create(index=index_name, ignore=400)
                es_conn.create(index=index_name, doc_type='_doc', body=update_fields, id=uuid.uuid4())

        return HttpResponse('sucess')
    except Exception as e:
        return HttpResponse(
            "Error running the program.")


############## runDocumentClassifierUnsupervised starts################################
@csrf_exempt
def runDocumentClassifierUnsupervised(request):
    current_tab = 2
    global response

    try:
        # Assuming that the client-side has already selected the options before running the program.
        # Assuming that the request from the client side will have all the fields necessary for running the program.
        if request.method == "GET":
            return response
        elif request.method == "POST":
            try:
                userName = request.user.username;
                gui_parameters = request.POST.getlist('inputData')[0]
                gui_parameters = json.loads(gui_parameters);
                finalTrainingData = request.FILES.getlist('trainFile')
                training_data = Process_All_Files(finalTrainingData)
            except Exception as e:
                print('parsing went wrong', e)

            index_name_gloabals = 'apolloglobalsunsupervised'
            query = {"query": {"bool": {"must": {"match": {"username.keyword": userName}}}}}
            res = es_conn.search(index=index_name_gloabals, body=query)
            id = res['hits']['hits'][0]['_id']
            # Set maximum value for progressbar. Reserve 20 for preprocessing and 30 for clustering. Additionally, for each cluster, reserve 5 for topic modeling
            progressbar_maximum = 50 + 5 * int(gui_parameters['number_of_clusters'])
            progressbar_value = 0

            # Set the text in progressbarlabel
            programRunStartTime = datetime.now()
            programRunStartTimeLabel = 'Progress: Program run started at ' + programRunStartTime.strftime(
                "%I:%M%p on %B %d, %Y") + ' (UTC time). '
            progress_text = ''
            progressbarlabel_text = programRunStartTimeLabel

            progress_text = progress_text + '-' * 75 + '\n' + "Program run started at " + programRunStartTime.strftime(
                "%I:%M%p on %B %d, %Y") + " (UTC time).\n" + \
                            '-' * 75 + '\n' + "Starting document classification process..."

            # output_folder = os.path.dirname(selectedDataFilePath)

            # The preprocessing is different for patent and journal data

            numberOfClusters = gui_parameters['number_of_clusters']
            number_of_top_words = gui_parameters['number_of_top_words']
            es_conn.update(index=index_name_gloabals, id=id,
                           body={"doc": {"progressbar_value": 0, 'progressbarlabel_text': progressbarlabel_text,
                                         'progress_text': progress_text, "progressbar_maximum": progressbar_maximum,
                                         "numberOfClusters": numberOfClusters,
                                         "number_of_top_words": number_of_top_words,
                                         "current_tab": current_tab}})
            # index_name = 'trainingfiledata'
            key_value = 'runDocumentClassifierUnsupervised/'
            key_value += userName + '/'
            s3 = boto3.client('s3')
            if gui_parameters['training_data_type'] == 'Patent':
                # query = {"query": {"bool": {"must": {"match": {"username.keyword": userName}}}}}
                # res = es_conn.search(index=index_name, body=query)
                # training_data = res['hits']['hits'][0]['_source']['training_data']

                # Read the paper data from the file
                # [type, title, abstract, author, affiliation, year]
                file_data_open = training_data
                file_data_open = file_data_open.split('\n')  # split by new line
                file_data_open = list(filter(None, file_data_open))  # delete empty lines

                # Now, the first line is a header, so remove it
                file_data_open = file_data_open[1:]

                progressbar_value = 5

                progress_text = progress_text + "\nFound " + str(
                    len(file_data_open)) + " documents! \nPreprocessing documents...\n"

                # Build the stop words
                stops = stopwords

                aux_stops = './static/AuxStops.txt'

                aux_stops = open(aux_stops, 'r').read()
                aux_stops = re.sub("[^a-zA-Z ]", "   ", aux_stops)  # remove non-alphanumeric
                aux_stops = " ".join(aux_stops.split())  # split by any whitespace and rejoin w/ space
                aux_stops = aux_stops.split(' ')
                aux_stops = list(filter(None, aux_stops))

                # append auxiliary stops
                stops = stops + aux_stops

                # Append user-provided stop words to the stop words list
                user_defined_stopwords = (" ".join(gui_parameters['additional_stopwords'].lower().split(','))).split()
                es_conn.update(index=index_name_gloabals, id=id,
                               body={"doc": {"progressbar_value": progressbar_value, 'progress_text': progress_text,
                                             "user_defined_stopwords": user_defined_stopwords,
                                             "training_data_type": gui_parameters['training_data_type']}})
                stops = stops + user_defined_stopwords

                # Bulid stopterm dictionary
                stopterms = {}
                for stop in stops:
                    if stop in stopterms:
                        stopterms[stop] += 1
                    else:
                        stopterms[stop] = 1

                (file_data_proc, file_data_stem, file_data_stop_words_removed) = preprocess_collection(file_data_open,
                                                                                                       stopterms, True,
                                                                                                       progress_text)
                file_data_proc = list(filter(None, file_data_proc))
                file_data = list(filter(None, file_data_stem))
                file_data_stop_words_removed = list(filter(None, file_data_stop_words_removed))

                title_data = [doc.split('\t')[1] for doc in file_data]
                abstract_data = [doc.split('\t')[2] for doc in file_data]
                claim_data = [doc.split('\t')[3] for doc in file_data]
                unlabeled_data = [' '.join(doc) for doc in zip(title_data, abstract_data, claim_data)]

                title_data_stop_words_removed = [doc.split('\t')[1] for doc in file_data_stop_words_removed]
                abstract_data_stop_words_removed = [doc.split('\t')[2] for doc in file_data_stop_words_removed]
                claim_data_stop_words_removed = [doc.split('\t')[3] for doc in file_data_stop_words_removed]
                unlabeled_data_stop_words_removed = [' '.join(doc) for doc in
                                                     zip(title_data_stop_words_removed,
                                                         abstract_data_stop_words_removed,
                                                         claim_data_stop_words_removed)]

                progressbar_value += 5

                progress_text = progress_text + "Removed duplicates and preprocessed " + str(
                    len(unlabeled_data)) + " documents." + \
                                "\nStarting unsupervised learning..."

                # Changed the n-grams to (1,5) in the line below
                tfidf_vect = TfidfVectorizer(analyzer='word',
                                             ngram_range=(1, 5),
                                             min_df=2,
                                             max_df=0.8,
                                             max_features=200000,
                                             stop_words='english',
                                             use_idf=True)

                # tf-idf with params
                unlabeled_data_tfidf = tfidf_vect.fit_transform(unlabeled_data)
                tfidfVectorizer = tfidf_vect

                tfidf = pickle.dumps(tfidfVectorizer)
                s3.put_object(Body=tfidf, Bucket=AWS_STORAGE_BUCKET_NAME,
                              Key=key_value + 'tfidf_vect.pkl')

                # Increment progress bar value by 10 for vectorization
                progressbar_value += 10

                # Model and model parameters
                model = gui_parameters['model']
                number_of_clusters = gui_parameters['number_of_clusters']
                es_conn.update(index=index_name_gloabals, id=id,
                               body={"doc": {"progressbar_value": progressbar_value, 'progress_text': progress_text,
                                             "model": model}})

                if model == 'K-means Clustering':
                    # determine the best model based on evaluating several models
                    automatic_mode = True
                    progress_text = progress_text + "\nPerforming clustering on the data using K-means clustering..."

                    # Added additonal checks to avoid errors and exceptions
                    # 1. Check that the number of documents in OTHERS category is at least greater than 2
                    if len(unlabeled_data) <= 2:
                        progress_text = progress_text + "\nThere are only " + str(len(
                            unlabeled_data)) + " documents in OTHERS category. Cannot perform clustering on less than or equal to 2 documents."
                        es_conn.update(index=index_name_gloabals, id=id,
                                       body={"doc": {'progress_text': progress_text}})
                        return HttpResponse("Error running the program.")

                    # 2. Check that the number of clusters specified is greater than or equal to the number of documents categorized as OTHERS
                    if len(unlabeled_data) < int(numberOfClusters):
                        progress_text = progress_text + "\nThe number of documents in OTHERS category is less than the number of clusters provided in the GUI. Cannot perform clustering in this case."
                        es_conn.update(index=index_name_gloabals, id=id,
                                       body={"doc": {'progress_text': progress_text}})
                        return HttpResponse("Error running the program.")

                    output_filename = re.sub('.txt', '_Results_Topic_Modeling.txt',
                                             gui_parameters['training_file_name'])
                    es_conn.update(index=index_name_gloabals, id=id,
                                   body={"doc": {'progress_text': progress_text}})

                    # step2: topic extraction and clustering
                    if len(unlabeled_data) > 2 and len(unlabeled_data) >= int(numberOfClusters):

                        # Added random_state for fixing results of clustering
                        random_state_kmeans = np.random.RandomState(seed=6758)
                        km = MiniBatchKMeans(n_clusters=int(numberOfClusters), random_state=random_state_kmeans)
                        km.partial_fit(unlabeled_data_tfidf)

                        trainedModel = km
                        trainedModelName = 'K_Means_Clustering_NumClusters=' + str(
                            numberOfClusters) + '_TopWords=' + str(
                            number_of_top_words)

                        clusters = km.labels_.tolist()

                        # start checking code below:

                        test_patent_others = {'content': unlabeled_data, 'file_others': file_data_proc,
                                              'content_stop_words_removed': unlabeled_data_stop_words_removed,
                                              'cluster': clusters}

                        frame = pd.DataFrame(test_patent_others, index=[clusters],
                                             columns=['content', 'file_others', 'content_stop_words_removed',
                                                      'cluster'])

                        topicModelingFrame = frame

                        clusterTopicsAndCounts = []

                        clustering_successful = False

                        # Increment progrss bar by 30, for clustering complete
                        progressbar_value += 30
                        trainedModel = pickle.dumps(trainedModel)
                        s3.put_object(Body=trainedModel, Bucket=AWS_STORAGE_BUCKET_NAME,
                                      Key=key_value + 'trainedModel.pkl')
                        topicModelingFrame = pickle.dumps(topicModelingFrame)
                        s3.put_object(Body=topicModelingFrame, Bucket=AWS_STORAGE_BUCKET_NAME,
                                      Key=key_value + 'topicModelingFrame.pkl')
                        es_conn.update(index=index_name_gloabals, id=id,
                                       body={"doc": {'progressbar_value': progressbar_value,
                                                     "trainedModelName": trainedModelName}})

                        # output lda other topics
                        # fout_others = {}
                        # fout_others = open(output_filename)
                        # with open(output_filename, 'w') as fout_others:
                        fout_others = ''
                        for no in range(int(numberOfClusters)):
                            try:
                                # sometimes, there is no document in the group, so handle that case with try and except
                                patent_group = frame.groupby(frame['cluster']).get_group(no)
                            except:
                                # continue, because there is no document in this cluster. Move on to the topic modeling for next cluster
                                continue

                            patent_tac = patent_group.ix[:, 0].tolist()
                            patent_org = patent_group.ix[:, 1].tolist()
                            patent_tac_stop_words_removed = patent_group.ix[:, 2].tolist()

                            lda_tf_vect = TfidfVectorizer(max_df=0.8, min_df=1,
                                                          max_features=200000,
                                                          ngram_range=(1, 5),
                                                          use_idf=True,
                                                          stop_words='english')
                            tf = None
                            try:
                                tf = lda_tf_vect.fit_transform(patent_tac_stop_words_removed)

                            except Exception as e:
                                lda_tf_vect = TfidfVectorizer(max_df=1.0, min_df=1,
                                                              max_features=200000,
                                                              ngram_range=(1, 5),
                                                              use_idf=True,
                                                              stop_words='english')
                                tf = lda_tf_vect.fit_transform(patent_tac_stop_words_removed)

                            # LDA Model
                            lda = LatentDirichletAllocation(n_components=1, max_iter=20,
                                                            learning_method='online',
                                                            learning_offset=50,
                                                            random_state=0).fit(tf)

                            lda_feature_names = lda_tf_vect.get_feature_names()

                            lda_topics = get_topic_list(lda, lda_feature_names, number_of_top_words)

                            clusterTopicsAndCounts.append([len(patent_tac), lda_topics[0]])

                            doc_topic = lda.transform(tf)
                            doc_topic_index = doc_topic.argmax(axis=1)

                            for doc, doc_topic_i in zip(patent_org, doc_topic_index):
                                fout_others += '\t'.join([doc.strip('\r').strip('\n'),
                                                          lda_topics[doc_topic_i].strip('\t').strip('\r').strip(
                                                              '\n')]) + '\n'

                            clustering_successful = True

                            # For each cluster's topic modeling, increment progressbar by 5
                            progressbar_value += 5

                        if clustering_successful == True:
                            progress_text = progress_text + '\nTopic extraction and clustering completed.'
                            # Load the topic modeling results in the treeview
                            # Compute the columns: (i) # instances, (ii) Topics extracted
                            # populateTreeviewWithTopicModelingResults(apollo4.globals.clusterTopicsAndCounts)

                        progress_text = progress_text + '\nPlease click on Download File button to download the file.'


                    elif len(unlabeled_data) > 0 and (
                            len(unlabeled_data) <= 2 or len(unlabeled_data) >= int(numberOfClusters)):
                        progress_text = progress_text + "\nTopic extraction could not be performed."

                    else:
                        progress_text = progress_text + "\nNo patent paper found for topic extraction."

                    programRunEndTime = datetime.now()

                    timeDifference = relativedelta(programRunEndTime, programRunStartTime)

                    programRunStartTimeLabel = "\nProgram run took %d days %d hours %d minutes %d seconds." % (
                        timeDifference.days, timeDifference.hours, timeDifference.minutes, timeDifference.seconds)
                    progress_text = progress_text + programRunStartTimeLabel

                es_conn.update(index=index_name_gloabals, id=id,
                               body={"doc": {'progressbar_value': progressbar_value, "progress_text": progress_text,
                                             "clusterTopicsAndCounts": clusterTopicsAndCounts}})
                response = HttpResponse(content=fout_others,
                                        content_type='text/plain')
                response['Content-Disposition'] = 'attachment; filename=' + output_filename
                final_progress_value = 200
                es_conn.update(index=index_name_gloabals, id=id,
                               body={"doc": {"final_progress_value": final_progress_value}})

                return HttpResponse('successfully patents executed patent')
            # FOR JOURNAL DATA
            elif gui_parameters['training_data_type'] == 'Journal':
                # query = {"query": {"bool": {"must": {"match": {"username.keyword": userName}}}}}
                # res = es_conn.search(index=index_name, body=query)
                # training_data = res['hits']['hits'][0]['_source']['training_data']

                file_data_open = training_data
                file_data_open = file_data_open.split('\n')  # split by new line
                file_data_open = list(filter(None, file_data_open))  # delete empty lines

                # Now, the first line is a header, so remove it
                file_data_open = file_data_open[1:]

                progressbar_value = 5

                progress_text = progress_text + "Found " + str(
                    len(file_data_open)) + " documents! \nPreprocessing documents...\n"

                # Remove the duplicated documents based on "title"
                file_data_open = dedup_collection_journal(file_data_open, 1, 2)

                # Preprocessing for scoupus data
                file_data_open = preprocess_collection_journal(file_data_open)

                user_defined_stopwords = gui_parameters['additional_stopwords'].lower()
                es_conn.update(index=index_name_gloabals, id=id,
                               body={"doc": {"progressbar_value": progressbar_value, 'progress_text': progress_text,
                                             "user_defined_stopwords": user_defined_stopwords,
                                             "training_data_type": gui_parameters['training_data_type']}})

                unlabeled_data = [' '.join([stop_and_stem_journal(doc.split('\t')[1], user_defined_stopwords)
                                               , stop_and_stem_journal(doc.split('\t')[2], user_defined_stopwords)
                                            ]) for doc in file_data_open]

                unlabeled_data_stop_words_removed = [
                    ' '.join([remove_stopwords_journal(doc.split('\t')[1], user_defined_stopwords)
                                 , remove_stopwords_journal(doc.split('\t')[2], user_defined_stopwords)
                              ]) for doc in file_data_open]

                progressbar_value += 5

                progress_text = progress_text + "Removed duplicates and preprocessed " + str(
                    len(unlabeled_data)) + " documents."

                # Changed the n-grams to (1,5) in the line below
                tfidf_vect = TfidfVectorizer(analyzer='word',
                                             ngram_range=(1, 5),
                                             min_df=2,
                                             max_df=0.8,
                                             max_features=200000,
                                             stop_words='english',
                                             use_idf=True)

                # tf-idf with params
                unlabeled_data_tfidf = tfidf_vect.fit_transform(unlabeled_data)
                tfidfVectorizer = tfidf_vect
                tfidf = pickle.dumps(tfidfVectorizer)
                s3.put_object(Body=tfidf, Bucket=AWS_STORAGE_BUCKET_NAME,
                              Key=key_value + 'tfidf_vect.pkl')
                # Increment progress bar value by 10 for vectorization
                progressbar_value += 10

                # Model and model parameters
                model = gui_parameters['model']
                es_conn.update(index=index_name_gloabals, id=id,
                               body={"doc": {"progressbar_value": progressbar_value, 'progress_text': progress_text,
                                             "model": model}})
                if model == 'K-means Clustering':
                    # determine the best model based on evaluating several models
                    automatic_mode = True
                    progress_text = progress_text + "\nPerforming clustering on the data using K-means clustering..."

                    # Added additonal checks to avoid errors and exceptions
                    # 1. Check that the number of documents in OTHERS category is at least greater than 2
                    if len(unlabeled_data) <= 2:
                        progress_text = progress_text + "\nThere are only " + str(len(
                            unlabeled_data)) + " documents in OTHERS category. Cannot perform clustering on less than or equal to 2 documents."
                        es_conn.update(index=index_name_gloabals, id=id,
                                       body={"doc": {'progress_text': progress_text}})
                        return HttpResponse(
                            "Error running the program. Please contact the IP Group Analytics Team")

                    # 2. Check that the number of clusters specified is greater than or equal to the number of documents categorized as OTHERS
                    if len(unlabeled_data) < int(numberOfClusters):
                        progress_text = progress_text + "\nThe number of documents in OTHERS category is less than the number of clusters provided in the GUI. Cannot perform clustering in this case."
                        es_conn.update(index=index_name_gloabals, id=id,
                                       body={"doc": {'progress_text': progress_text}})
                        return HttpResponse(
                            "Error running the program.")

                    output_filename = re.sub('.txt', '_Results_Topic_Modeling.txt',
                                             gui_parameters['training_file_name'])
                    es_conn.update(index=index_name_gloabals, id=id,
                                   body={"doc": {'progress_text': progress_text}})

                    # 3: topic extraction and clustering
                    if len(unlabeled_data) > 2 and len(unlabeled_data) >= int(numberOfClusters):

                        # Added random_state for fixing results of clustering
                        random_state_kmeans = np.random.RandomState(seed=6758)
                        km = MiniBatchKMeans(n_clusters=int(numberOfClusters), random_state=random_state_kmeans)
                        km.partial_fit(unlabeled_data_tfidf)

                        trainedModel = km
                        trainedModelName = 'K_Means_Clustering_NumClusters=' + str(
                            numberOfClusters) + '_TopWords=' + str(
                            number_of_top_words)

                        clusters = km.labels_.tolist()

                        test_patent_others = {'content': unlabeled_data, 'file_others': file_data_open,
                                              'content_stop_words_removed': unlabeled_data_stop_words_removed,
                                              'cluster': clusters}

                        frame = pd.DataFrame(test_patent_others, index=[clusters],
                                             columns=['content', 'file_others', 'content_stop_words_removed',
                                                      'cluster'])

                        topicModelingFrame = frame

                        clusterTopicsAndCounts = []

                        # Increment progress bar by 30, for clustering complete
                        progressbar_value += 30
                        trainedModel = pickle.dumps(trainedModel)
                        s3.put_object(Body=trainedModel, Bucket=AWS_STORAGE_BUCKET_NAME,
                                      Key=key_value + 'trainedModel.pkl')
                        topicModelingFrame = pickle.dumps(topicModelingFrame)
                        s3.put_object(Body=topicModelingFrame, Bucket=AWS_STORAGE_BUCKET_NAME,
                                      Key=key_value + 'topicModelingFrame.pkl')
                        es_conn.update(index=index_name_gloabals, id=id,
                                       body={"doc": {'progressbar_value': progressbar_value,
                                                     "trainedModelName": trainedModelName}})
                        # output lda other topics
                        # with open(output_filename, 'w') as fout_others:
                        fout_others = ''
                        for no in range(int(numberOfClusters)):
                            patent_group = frame.groupby(frame['cluster']).get_group(no)
                            patent_tac = patent_group.ix[:, 0].tolist()
                            patent_org = patent_group.ix[:, 1].tolist()

                            clustering_successful = False
                            try:
                                lda_tf_vect = TfidfVectorizer(max_df=0.8, min_df=1,
                                                              max_features=200000,
                                                              ngram_range=(1, 5),
                                                              use_idf=True,
                                                              stop_words='english')
                                tf = lda_tf_vect.fit_transform(patent_tac)

                                # LDA Model
                                lda = LatentDirichletAllocation(n_components=1, max_iter=20,
                                                                learning_method='online',
                                                                learning_offset=50,
                                                                random_state=0).fit(tf)

                                lda_feature_names = lda_tf_vect.get_feature_names()
                                lda_topics = get_topic_list(lda, lda_feature_names, number_of_top_words)
                                clusterTopicsAndCounts.append([len(patent_tac), lda_topics[0]])
                                doc_topic = lda.transform(tf)
                                doc_topic_index = doc_topic.argmax(axis=1)
                                for doc, doc_topic_i in zip(patent_org, doc_topic_index):
                                    fout_others += '\t'.join(
                                        [doc, lda_topics[doc_topic_i].strip('\r').strip('\n')]) + '\n'

                                clustering_successful = True

                            except:
                                progress_text = progress_text + '\nERROR: Unexpected error.'
                                es_conn.update(index=index_name_gloabals, id=id,
                                               body={"doc": {'progress_text': progress_text}})
                                return HttpResponse(
                                    "Error running the program.")

                            # For each cluster's topic modeling, increment progressbar by 5
                            progressbar_value += 5

                        if clustering_successful == True:
                            progress_text = progress_text + '\nTopic extraction and clustering completed.'

                        progress_text = progress_text + "\nPlease download the " + output_filename + " file and check all the results in the file."


                    elif len(unlabeled_data) > 0 and (
                            len(unlabeled_data) <= 2 or len(unlabeled_data) >= int(numberOfClusters)):
                        progress_text = progress_text + "\nTopic extraction could not be performed."
                        es_conn.update(index=index_name_gloabals, id=id,
                                       body={"doc": {'progress_text': progress_text}})

                    else:
                        progress_text = progress_text + "\nNo journal paper found for topic extraction."
                        es_conn.update(index=index_name_gloabals, id=id,
                                       body={"doc": {'progress_text': progress_text}})

                programRunEndTime = datetime.now()

                timeDifference = relativedelta(programRunEndTime, programRunStartTime)

                programRunStartTimeLabel = "Program run took %d days %d hours %d minutes %d seconds." % (
                    timeDifference.days, timeDifference.hours, timeDifference.minutes, timeDifference.seconds)
                progressbarlabel_text = programRunStartTimeLabel

                response = HttpResponse(content=fout_others,
                                        content_type='text/plain')
                response['Content-Disposition'] = 'attachment; filename=' + output_filename
                final_progress_value = 200
                es_conn.update(index=index_name_gloabals, id=id,
                               body={"doc": {'progress_text': progress_text, "progressbar_value": progressbar_value,
                                             "clusterTopicsAndCounts": clusterTopicsAndCounts,
                                             "progressbarlabel_text": progressbarlabel_text,
                                             "final_progress_value": final_progress_value}})

                return HttpResponse('successfully executed')
    except Exception as e:
        errorString = 'Error while running the program please contact IP Group Analytics team.'
        final_progress_value = 200
        es_conn.update(index=index_name_gloabals, id=id,
                       body={"doc": {'errorString': errorString,
                                     "final_progress_value": final_progress_value}})
        return HttpResponse(
            "Error running the program.")


@csrf_exempt
def fetch_update_unsupervised(request):
    # dumps_clusterTopicsAndCounts = list(apollo4.globals.clusterTopicsAndCounts)

    try:
        if request.method == 'POST':
            userName = request.user.username;
            index_name_gloabals = 'apolloglobalsunsupervised'
            query = {"query": {"bool": {"must": {"match": {"username.keyword": userName}}}}}
            res = es_conn.search(index=index_name_gloabals, body=query)
            data = res['hits']['hits'][0]['_source']

        return JsonResponse({'data': data})
    except Exception as e:
        return HttpResponse(
            "Error running the program.")


############## runDocumentClassifierUnsupervised ends ################################

############################# patent scoring functions starts ###########################
def remove_tab(text):
    try:
        if not text:
            return text  # return empty string if string is empty
        text = text.replace('\n', ' ').replace('\t', ' ').strip()  # remove line break and tab
        return text
    except Exception as e:
        return HttpResponse(
            "Error running the program. Please contact the IP Group Analytics Team () to resolve the issue. Please provide the error details below in your email. \nPlease provide all the steps to reproduce this issue. \n" + "-" * 40 + "\n" + str(
                e) + "\n" + "-" * 40)


def preprocess_collection_PatentScoring(file_open, stopterms, inventionDisclosure=False):
    try:
        progress_text = ''
        if inventionDisclosure:
            # Preprocess the text file containing text of invention disclosure
            # stop and stem title, abstract, claim
            file_stem = stop_and_stem(file_open[0], stopterms)
            file_proc = remove_tab(file_open[0])
            output = ([file_proc], [file_stem])
            progress_text += '\nStopwords removed and terms stemmed in the keyword search file.'
        else:
            # preprocess patent data as usual
            # de-duplication
            apn_cur = []  # list of current used application numbers
            file_dedup = []  # the collection of deduped patents
            counter = 0  # num of deduped patents
            num_docs = len(file_open)
            for index, doc in enumerate(file_open, start=1):
                try:
                    # print "de-duplication document %d of %d" % (index,num_docs)
                    apn = doc.split("\t")[4]
                    if apn not in apn_cur:
                        file_dedup.append(doc)
                        apn_cur.append(apn)
                        counter = counter + 1
                except Exception as e:
                    #            print "ERROR: document %d could not be processed" % index
                    print(e)

            # step through the deduped patent collection
            file_proc = []
            file_stem = []
            design_count = 0  # counter for design cases
            utility_count = 0  # counter for utility cases

            for doc_no, doc in enumerate(file_dedup, start=1):
                #        print "pre-processing document %d of %d" % (doc_no,len(file_open))
                fields = doc.split("\t")
                try:
                    no = str(doc_no)
                    file_t = remove_tab(fields[1])  # title
                    file_a = remove_tab(fields[2])  # abstract
                    file_c = remove_tab(fields[3])  # claims
                    apn = fields[4].lower()
                    apd = fields[5]
                    asgn = fields[6].lower()
                #            upc = fields[7].lower()
                except Exception as e:
                    print(e)

                if apn.startswith("us2"):
                    # filter out design cases
                    progress_text += 'Design patent found! App_No: %r' % (apn)
                    design_count = design_count + 1
                    continue

                if apn.startswith("us"):
                    utility_count = utility_count + 1

                    # stop and stem title, abstract, claim
                    file_t_stem = stop_and_stem(file_t, stopterms)
                    file_a_stem = stop_and_stem(file_a, stopterms)
                    file_c_stem = stop_and_stem(file_c, stopterms)

                    # Output the orginal clean version of utility patent
                    file_proc.append([no, file_t, file_a, file_c, apn, apd, asgn])
                    # Output the stemmed version of utility patent
                    file_stem.append([no, file_t_stem, file_a_stem, file_c_stem, apn, apd, asgn])

            output = (file_proc, file_stem)
            # progress_text += '\nstopwords removed, terms stemmed, documents de-duplicated, design removed'
            # progress_text += '\n%d unique documents out of %d total' % (counter, num_docs)
            # progress_text += '\n%d design documents out of %d total' % (design_count, num_docs)
            # progress_text += '\n%d utility documents out of %d total' % (utility_count, num_docs)
        return output
    except Exception as e:
        return HttpResponse(
            "Error running the program." + str(e))


# Note that the function below will be used to validate both Samsung Patent data file and Non-Samsung patent data file that the user will upload in GUI.
@csrf_exempt
def patentScoringGlobals(request):
    try:
        if request.method == 'POST':
            userName = request.user.username;
            update_fields = {
                'username': userName,
                'progressbar_maximum': 0,  # progress bar max_value
                'progress_value': 0,  # progress bar value
                'progressbarlabel_text': '',  # progress field
                'progress_text': '',  # progress text
                'final_progress_value': '',
                'current_tab': 0,
                'errorString': '',
                "training_data_type": '',
                "testing_data_type": ''
            }
            index_name_gloabals = 'apolloglobalspatentscoring'
            query = {"query": {"bool": {"must": {"match": {"username.keyword": userName}}}}}
            # es_conn.indices.create(index=index_name_gloabals)
            es_conn.delete_by_query(index=index_name_gloabals, body=query)
            es_conn.create(index=index_name_gloabals, doc_type='_doc', body=update_fields, id=uuid.uuid4())
        return HttpResponse('sucess')
    except Exception as e:
        return HttpResponse(
            "Error running the program." + str(e))


@csrf_exempt
def userRunModelTrackPS(request):
    try:
        if request.method == 'POST':
            try:
                userName = request.user.username;
                userTrack = request.body.decode('utf-8');
                userTrack = json.loads(userTrack);
                filename_NonSamsung_Patents = userTrack['filename_NonSamsung_Patents']
                training_data_type = userTrack['training_data_type']
                testing_data_type = userTrack['testing_data_type']
                keywords = userTrack['keywords']
                searchType = userTrack['searchType']
            except Exception as e:
                print('parsing went wrong', e)

            time = datetime.now()
            time = time.strftime("%I:%M%p on %B %d, %Y")

            update_fields = {
                'username': userName,
                'filename_NonSamsung_Patents': filename_NonSamsung_Patents,
                'training_data_type': training_data_type,
                'testing_data_type': testing_data_type,
                'keywords': keywords,
                'searchType': searchType,
                'time': time + ' UTC  Time'
            }
            index_name = 'userrunmodeltrackps'
            if es_conn.indices.exists(index_name):
                es_conn.create(index=index_name, doc_type='_doc', body=update_fields, id=uuid.uuid4())
            else:
                save_response = es_conn.indices.create(index=index_name, ignore=400)
                es_conn.create(index=index_name, doc_type='_doc', body=update_fields, id=uuid.uuid4())

        return HttpResponse('sucess')
    except Exception as e:
        return HttpResponse(
            "Error running the program." + str(e))


@csrf_exempt
def patentScoringData(request):
    superVisedResponse = {}
    global response
    try:
        if request.method == "POST":
            response = None
            get_value = request.body
            get_value = get_value.decode("utf-8")

            if "identification number\ttitle\tabstract\tclaims\tapplication number\tapplication date\tcurrent assignee\tupc" in get_value.lower():
                checkPatentData = 'Patent'
            elif "keywords" in get_value.lower():
                checkPatentData = 'Keywords'
            elif 'nasca' in get_value.lower():
                checkPatentData = 'NASCA File Error.'
            else:
                checkPatentData = 'Patent File Error.'

            return HttpResponse(checkPatentData)
    except Exception as e:
        return HttpResponse(
            "Error running the program.")


@csrf_exempt
def computeSimilarityBetweenSamsungAndNonSamsungPatents(request):
    global response

    if response != None:
        return response
    else:
        try:
            if request.method:
                username = request.user.username;
                get_value = request.body
                get_value = get_value.decode("utf-8")
                get_value = json.loads(get_value)
                output_fext = splitext(get_value['filename_NonSamsung_Patents'])[1]
                output_response_name = splitext(get_value['filename_NonSamsung_Patents'])[0]
                training_data_type = get_value['training_data_type'],
                testing_data_type = get_value['testing_data_type'],
                searchType = get_value['searchType'],
                keywords = get_value['keywords'],
                current_tab = 7
                index_name_gloabals = 'apolloglobalspatentscoring'
                query = {"query": {"bool": {"must": {"match": {"username.keyword": username}}}}}
                res = es_conn.search(index=index_name_gloabals, body=query)
                id = res['hits']['hits'][0]['_id']
                progress_text = ""
                progress_value = 0
                progressbar_maximum = 200
                programRunStartTime = datetime.now()
                programRunStartTimeLabel = 'Progress: Program run started at ' + programRunStartTime.strftime(
                    "%I:%M%p on %B %d, %Y") + ' (UTC time). '

                progressbarlabel_text = programRunStartTimeLabel
                progress_text += progress_text + '-' * 75 + '\n' + "Program run started at " + programRunStartTime.strftime(
                    "%I:%M%p on %B %d, %Y") + " (UTC time).\n" + \
                                 '-' * 75 + '\n' + "Starting document classification process..."
                # filename_NonSamsung_Patents
                es_conn.update(index=index_name_gloabals, id=id,
                               body={"doc": {'progressbarlabel_text': progressbarlabel_text,
                                             'progress_text': progress_text, "progressbar_maximum": progressbar_maximum,
                                             "current_tab": current_tab, "training_data_type": training_data_type,
                                             "testing_data_type": testing_data_type}})
                # Build the stop words
                stops = stopwords

                aux_stops = './static/AuxStops.txt'

                aux_stops = open(aux_stops, 'r').read()
                aux_stops = re.sub("[^a-zA-Z ]", "   ", aux_stops)  # remove non-alphanumeric
                aux_stops = " ".join(aux_stops.split())  # split by any whitespace and rejoin w/ space
                aux_stops = aux_stops.split(' ')
                aux_stops = list(filter(None, aux_stops))

                # append auxiliary stops
                stops = stops + aux_stops

                # Bulid stopterm dictionary
                stopterms = {}
                for stop in stops:
                    if stop in stopterms:
                        stopterms[stop] += 1
                    else:
                        stopterms[stop] = 1

                query = {"query": {"bool": {"must": {"match": {"username.keyword": username}}}}}
                res = es_conn.search(index='testingfiledata', body=query)
                testing_data = res['hits']['hits'][0]['_source']['testing_data']
                file_NonSamsungPatents_open = testing_data
                file_NonSamsungPatents_open = file_NonSamsungPatents_open.split('\n')  # split by new line
                file_NonSamsungPatents_open = list(filter(None, file_NonSamsungPatents_open))  # delete empty lines
                file_NonSamsungPatents_open = file_NonSamsungPatents_open[1:]
                progress_value += 5
                progress_text += '\nPreprocessing related patents file...'
                es_conn.update(index=index_name_gloabals, id=id,
                               body={"doc": {"progress_value": progress_value,
                                             'progress_text': progress_text}})
                (file_NonSamsungPatents_proc, file_NonSamsungPatents_stem) = preprocess_collection_PatentScoring(
                    file_NonSamsungPatents_open, stopterms, inventionDisclosure=False)
                file_NonSamsungPatents_proc = list(
                    filter(None, file_NonSamsungPatents_proc))  # patent list after dedup and remove the design patents
                file_NonSamsungPatents_stem = list(filter(None, file_NonSamsungPatents_stem))

                # patent list after stem

                # Check whether input file contains search keywords, instead of patent list
                keywordSearch = False
                if searchType[0].lower() == 'keywords':
                    keywordSearch = True
                    file_SamsungPatents_open = [" ".join(keywords[0].lower().split(','))]
                else:
                    res = es_conn.search(index='trainingfiledata', body=query)
                    training_data = res['hits']['hits'][0]['_source']['training_data']
                    file_SamsungPatents_open = training_data
                    file_SamsungPatents_open = file_SamsungPatents_open.split('\n')  # split by new line
                    file_SamsungPatents_open = list(filter(None, file_SamsungPatents_open))  # delete empty lines
                    file_SamsungPatents_open = file_SamsungPatents_open[1:]

                progress_value += 5
                if keywordSearch:
                    progress_text += '\nPreprocessing search keywords file...'
                else:
                    progress_text += '\nPreprocessing input patents file...'
                es_conn.update(index=index_name_gloabals, id=id,
                               body={"doc": {"progress_value": progress_value,
                                             'progress_text': progress_text}})
                (file_SamsungPatents_proc, file_SamsungPatents_stem) = preprocess_collection_PatentScoring(
                    file_SamsungPatents_open, stopterms, inventionDisclosure=keywordSearch)
                file_SamsungPatents_proc = list(
                    filter(None, file_SamsungPatents_proc))  # patent list after dedup and remove the design patents
                file_SamsungPatents_stem = list(filter(None, file_SamsungPatents_stem))  # patent list after stem
                progress_value += 15
                es_conn.update(index=index_name_gloabals, id=id,
                               body={"doc": {"progress_value": progress_value,
                                             'progress_text': progress_text}})
                # combine title, abstract, claim
                if keywordSearch:
                    tac_SamsungPatents = [file_SamsungPatents_stem]
                    progressbar_maximum = 100 + len(file_NonSamsungPatents_proc)
                else:
                    tac_SamsungPatents = [' '.join([doc[1], doc[2], doc[3]]) for doc in file_SamsungPatents_stem]
                    progressbar_maximum = 100 + len(file_NonSamsungPatents_proc) * len(tac_SamsungPatents)

                tac_NonSamsungPatents = [' '.join([doc[1], doc[2], doc[3]]) for doc in file_NonSamsungPatents_stem]

                num_SamsungPatents = len(tac_SamsungPatents)
                num_NonSamsungPatents = len(tac_NonSamsungPatents)

                # debug print
                if keywordSearch:
                    docsALL = tac_SamsungPatents[0] + tac_NonSamsungPatents
                else:
                    docsALL = tac_SamsungPatents + tac_NonSamsungPatents
                doc_terms = []
                for doc in docsALL:
                    doc_terms = doc_terms + doc.split(' ')

                # LSA
                # calculate the tf-idf for the doc-term matrix
                cvTFIDF = TfidfVectorizer(norm='l1')
                matrix_TFIDF = cvTFIDF.fit_transform(docsALL)
                (num_docs, num_features) = np.shape(matrix_TFIDF)
                progress_value += 25
                es_conn.update(index=index_name_gloabals, id=id,
                               body={"doc": {"progress_value": progress_value}})
                if num_features > 100:
                    num_features = 100
                # only keep the top 100 features
                svd = TruncatedSVD(n_components=num_features)

                # dimension reduction to low rank matrix
                matrix_TFIDF_SVD = svd.fit_transform(matrix_TFIDF)

                matrix_SamsungPatents = matrix_TFIDF_SVD[0:num_SamsungPatents, :]
                matrix_NonSamsungPatents = matrix_TFIDF_SVD[num_SamsungPatents:, :]

                cos_d_all = cosine_similarity(matrix_SamsungPatents, matrix_NonSamsungPatents)

                # zipObj = ZipFile(splitext(get_value['filename_NonSamsung_Patents'])[0] + '.zip', 'w')
                zipFileName = splitext(get_value['filename_NonSamsung_Patents'])[0] + '.zip'

                zipIO = io.BytesIO()

                zipObj = ZipFile(zipIO, 'w')
                progress_value += 25
                es_conn.update(index=index_name_gloabals, id=id,
                               body={"doc": {"progress_value": progress_value}})

                if keywordSearch:
                    # output the ranked top 5 patent list
                    output_file = 'TopFiveKeywordSearchResults.txt'
                    cos_d = cos_d_all[0, :]
                    file_NonSamsungPatents_sort = (np.array(file_NonSamsungPatents_proc)[np.argsort(cos_d)]).tolist()[
                                                  ::-1]  # to list of list
                    cos_d_sort = (np.sort(cos_d)).tolist()[::-1]  # to list

                    tempString = ''
                    numberOfPatentsToDisplay = 5
                    patentCounter = 0
                    for doc, score in zip(file_NonSamsungPatents_sort, cos_d_sort):
                        progress_value += 1
                        if patentCounter < numberOfPatentsToDisplay:
                            patentCounter += 1
                            samsungPatent_apd2 = datetime.strptime(doc[5], '%m/%d/%Y')
                            tempString += '\t'.join(doc + ["{:.2%}".format(score)]) + '\n'
                        else:
                            break

                    zipObj.writestr(output_file, tempString)
                else:
                    # output the ranked reference patent list of ith samsung patent
                    for samsungPatent_idx, samsungPatent_doc in enumerate(file_SamsungPatents_proc):
                        samsungPatent_apn = samsungPatent_doc[4].replace('/', '-')
                        samsungPatent_apd = datetime.strptime(samsungPatent_doc[5], '%m/%d/%Y')
                        output_file = samsungPatent_apn + '_rank' + output_fext
                        cos_d = cos_d_all[samsungPatent_idx, :]
                        file_NonSamsungPatents_sort = (np.array(file_NonSamsungPatents_proc)[
                            np.argsort(cos_d)]).tolist()[
                                                      ::-1]  # to list of list
                        cos_d_sort = (np.sort(cos_d)).tolist()[::-1]  # to list

                        tempString = ''
                        for doc, score in zip(file_NonSamsungPatents_sort, cos_d_sort):
                            samsungPatent_apd2 = datetime.strptime(doc[5], '%m/%d/%Y')
                            if (samsungPatent_apd <= samsungPatent_apd2 and 'samsung' not in doc[6].lower()):
                                tempString += '\t'.join(doc + ["{:.2%}".format(score)]) + '\n'
                            progress_value += 1

                        zipObj.writestr(output_file, tempString)

                zipObj.close()
                progress_text += '\nPatent scoring completed sucessfully.'
                progress_value += 25
                final_progress_value = 200
                response = HttpResponse(zipIO.getvalue(), content_type='application/x-zip-compressed')
                response['Content-Disposition'] = 'attachment; filename=' + zipFileName
                es_conn.update(index=index_name_gloabals, id=id,
                               body={"doc": {"progress_text": progress_text, "progress_value": progress_value,
                                             "final_progress_value": final_progress_value}})

            return HttpResponse('success')
        except Exception as e:
            return HttpResponse(
                "Error running the program." + str(e))


@csrf_exempt
def fetch_update_patentscoring(request):
    try:
        if request.method == 'POST':
            userName = request.user.username;
            index_name_gloabals = 'apolloglobalspatentscoring'
            query = {"query": {"bool": {"must": {"match": {"username.keyword": userName}}}}}
            res = es_conn.search(index=index_name_gloabals, body=query)
            data = res['hits']['hits'][0]['_source']

        return JsonResponse({'data': data})
    except Exception as e:
        return HttpResponse(
            "Error running the program." + str(e))


############################# patent scoring functions ends###########################
@csrf_exempt
def userRunModelTrackEM(request):
    try:
        if request.method == 'POST':
            try:
                userName = request.user.username;
                userTrack = request.body.decode('utf-8');
                userTrack = json.loads(userTrack);
                saveProjectName = userTrack['saveProjectName']
                trainedModelName = userTrack['trainedModelName']
                testing_file_name = userTrack['testing_file_name']
                testing_data_type = userTrack['testing_data_type']
            except Exception as e:
                print('parsing went wrong', e)

            time = datetime.now()
            time = time.strftime("%I:%M%p on %B %d, %Y")

            update_fields = {
                'username': userName,
                'saveProjectName': saveProjectName,
                'trainedModelName': trainedModelName,
                'testing_file_name': testing_file_name,
                'testing_data_type': testing_data_type,
                'time': time + ' UTC  Time'
            }
            index_name = 'userrunmodeltrackem'
            if es_conn.indices.exists(index_name):
                es_conn.create(index=index_name, doc_type='_doc', body=update_fields, id=uuid.uuid4())
            else:
                save_response = es_conn.indices.create(index=index_name, ignore=400)
                es_conn.create(index=index_name, doc_type='_doc', body=update_fields, id=uuid.uuid4())

        return HttpResponse('sucess')
    except Exception as e:
        return HttpResponse(
            "Error running the program." + str(e))


@csrf_exempt
def makePredictionsForSupervisedLearning(request):
    global responseTrain

    try:
        if request.method == "GET":
            return responseTrain
        elif request.method == "POST":
            userName = request.user.username;
            gui_parameters = request.POST.getlist('inputData')[0]
            get_value = json.loads(gui_parameters);
            finaltestingData = request.FILES.getlist('testFile')
            # training_data = request.FILES.getlist('file').read().decode("ISO-8859-1")
            testing_data = Process_All_Files(finaltestingData)
            saveProjectName = get_value['saveProjectName']
            trainedModelName = get_value['trainedModelName']
            testingFileName = get_value['testing_file_name']
            testingDataType = get_value['testing_data_type']
            progressbar_maximum = 120
            progressbar_value = 0

            # 1. supervised learning

            # Set the text in progressbarlabel
            programRunStartTime = datetime.now()
            programRunStartTimeLabel = 'Progress: Program run started at ' + programRunStartTime.strftime(
                "%I:%M%p on %B %d, %Y") + '. '

            progressbarlabel_text = programRunStartTimeLabel
            progress_text = ''
            progress_text = progress_text + '-' * 75 + '\n' + \
                            "Program run started at " + programRunStartTime.strftime(
                "%I:%M%p on %B %d, %Y") + " (UTC time)." + '\n' + '-' * 75 + '\n'
            index_name_gloabals = 'apolloglobals'
            query = {"query": {"bool": {"must": {"match": {"username.keyword": userName}}}}}
            res = es_conn.search(index=index_name_gloabals, body=query)
            id = res['hits']['hits'][0]['_id']
            es_conn.update(index=index_name_gloabals, id=id,
                           body={"doc": {"progressbar_maximum": progressbar_maximum, "current_tab": 3,
                                         "progressbar_value": progressbar_value,
                                         "testingFileName": testingFileName, "trainedModelName": trainedModelName,
                                         "testingDataType": testingDataType, "progress_text": progress_text
                               , "progressbarlabel_text": progressbarlabel_text}})
            # historyFilename = ''
            # if os.name == 'nt':
            #    historyFilename = 'C:/Users/manali.s/Desktop/Classification_Projects/'
            # else:
            #    historyFilename = '/data/Classification_Projects/'

            # historyFilename += comboboxAvailableProjects.get() + '/history.txt'
            key_value = 'classificationprojects/'
            key_value += saveProjectName + '/supervised/'
            key_value += trainedModelName + '/'

            s3 = boto3.client('s3')
            response1 = s3.get_object(Bucket=AWS_STORAGE_BUCKET_NAME,
                                      Key=key_value + 'tfidf_vect.pkl')
            response2 = s3.get_object(Bucket=AWS_STORAGE_BUCKET_NAME,
                                      Key=key_value + 'trainedModel.pkl')
            # for line in contents.splitlines():

            progressbar_value += 5

            # Load the tfidf vectorizer pickle file that was previously saved in S3
            tfidfVectorizer = response1['Body'].read()
            tfidf_vect = pickle.loads(tfidfVectorizer)

            # 1. Load the model
            model_file = response2['Body'].read()
            model = pickle.loads(model_file)

            progressbar_value += 5

            # Prediction Phase

            # The code for patent and journal testing data is different because it required different preprocessing
            es_conn.update(index=index_name_gloabals, id=id,
                           body={"doc": {"progressbar_value": progressbar_value}})

            if testingDataType == 'Patent':
                # Read the test patent data
                file_test_open = testing_data
                file_test_open = file_test_open.split('\n')  # split by new line
                file_test_open = list(filter(None, file_test_open))  # delete empty lines

                # Now, the first line is header, so remove the first line
                file_test_open = file_test_open[1:]

                progress_text = progress_text + "Preprocessing unlabeled data..."

                progressbar_value += 5

                # Build the stop words
                stops = stopwords

                aux_stops = './static/AuxStops.txt'

                aux_stops = open(aux_stops, 'r').read()
                aux_stops = re.sub("[^a-zA-Z ]", "   ", aux_stops)  # remove non-alphanumeric
                aux_stops = " ".join(aux_stops.split())  # split by any whitespace and rejoin w/ space
                aux_stops = aux_stops.split(' ')
                aux_stops = list(filter(None, aux_stops))

                # Bulid stopterm dictionary
                stopterms = {}
                for stop in stops:
                    if stop in stopterms:
                        stopterms[stop] += 1
                    else:
                        stopterms[stop] = 1

                (file_test_proc, file_test_stem, temp) = preprocess_collection(file_test_open, stopterms, False,
                                                                               progress_text)
                file_test_proc = list(filter(None, file_test_proc))
                file_test = list(filter(None, file_test_stem))

                title_test = [doc.split('\t')[1] for doc in file_test]
                abstract_test = [doc.split('\t')[2] for doc in file_test]
                claim_test = [doc.split('\t')[3] for doc in file_test]
                test_data = [' '.join(doc) for doc in zip(title_test, abstract_test, claim_test)]

                progressbar_value += 15
                progress_text = progress_text + "\nMaking predictions on unlabeled data..."

                # convert text data to tfidf
                test_data_tfidf = tfidf_vect.transform(test_data)
                model_predicted = model.predict(test_data_tfidf.todense())
                model_predicted = model_predicted.astype('U128')
                es_conn.update(index=index_name_gloabals, id=id,
                               body={"doc": {"progressbar_value": progressbar_value, "progress_text": progress_text,
                                             "model_predicted": json.dumps(model_predicted.tolist())}})

            elif testingDataType == 'Journal':
                file_test_open = testing_data
                file_test_open = file_test_open.split('\n')  # split by new line
                file_test_open = list(filter(None, file_test_open))  # delete empty lines

                # Now, the first line is header, so remove the first line
                file_test_open = file_test_open[1:]

                progress_text = progress_text + "Preprocessing unlabeled data..."
                progressbar_value += 10

                # Remove the duplicated document based on "title"
                file_test_open = dedup_collection_journal(file_test_open, 1, 2)

                # preprocessing for scoupus data
                file_test_proc = preprocess_collection_journal(file_test_open)

                user_defined_stopwords = []

                test_data = [' '.join([stop_and_stem_journal(doc.split('\t')[1], user_defined_stopwords)
                                          , stop_and_stem_journal(doc.split('\t')[2], user_defined_stopwords)
                                       ]) for doc in file_test_proc]

                progressbar_value += 20
                progress_text = progress_text + "\nMaking predictions on unlabeled data..."

                # convert text data to tfidf
                test_data_tfidf = tfidf_vect.transform(test_data)
                model_predicted = model.predict(test_data_tfidf.todense())
                model_predicted = model_predicted.astype('U128')
                # end journal test data preprocessing
                es_conn.update(index=index_name_gloabals, id=id,
                               body={"doc": {"progressbar_value": progressbar_value, "progress_text": progress_text,
                                             "model_predicted": json.dumps(model_predicted.tolist())}})

            model_prob = None
            # The 'predict_proba 'function cannot be used for SGD with hinge loss, hence we need calibrate the probability estimates for SGD with hinge loss
            if 'Support_Vector_Machine' in trainedModelName:
                model_isotonic_calibration = None
                model_sigmoid_calibration = None

                # Note that for Support Vector machine model we save the calibration model. So, we need to load the calibration model from S3 in this case.

                # Load the model to calibrate probabilities that will be used by the excel sheet
                # The calibration model that is savied could be either 'model_sigmoid_calibration' or 'model_isotonic_calibration'
                # We need to check which one of these models was saved for SVM.

                try:
                    s3.Object(AWS_STORAGE_BUCKET_NAME, key_value + 'model_sigmoid_calibration.pkl').load()
                except botocore.exceptions.ClientError as e:
                    if e.response['Error']['Code'] == "404":
                        # The object does not exist.
                        model_sigmoid_calibration = None
                    elif e.response['Error']['Code'] == "NoSuchKey":
                        pass
                    else:
                        # Something else has gone wrong.
                        final_progress_value = 200
                        errorString = 'Error running retreiving data from s3. Please contact the IP Group Analytics Team'
                        es_conn.update(index=index_name_gloabals, id=id,
                                       body={"doc": {"final_progress_value": final_progress_value,
                                                     "errorString": errorString}})
                        return JsonResponse({'finalResponse': 'Error'})
                else:
                    # The object does exist.
                    response3 = s3.get_object(Bucket=AWS_STORAGE_BUCKET_NAME,
                                              Key=key_value + 'model_sigmoid_calibration.pkl')
                    sigmoid_calibration_file = response3['Body'].read()
                    model_sigmoid_calibration = pickle.loads(sigmoid_calibration_file)

                try:
                    s3.Object(AWS_STORAGE_BUCKET_NAME, key_value + 'model_isotonic_calibration.pkl').load()
                except botocore.exceptions.ClientError as e:
                    if e.response['Error']['Code'] == "404":
                        # The object does not exist.
                        model_isotonic_calibration = None
                    elif e.response['Error']['Code'] == "NoSuchKey":
                        pass
                    else:
                        # Something else has gone wrong.
                        final_progress_value = 200
                        errorString = 'Error running retreiving data from s3. Please contact the IP Group Analytics Team'
                        es_conn.update(index=index_name_gloabals, id=id,
                                       body={"doc": {"final_progress_value": final_progress_value,
                                                     "errorString": errorString}})
                        return JsonResponse({'finalResponse': 'Error'})
                else:
                    # The object does exist.
                    response4 = s3.get_object(Bucket=AWS_STORAGE_BUCKET_NAME,
                                              Key=key_value + 'model_isotonic_calibration.pkl')
                    isotonic_calibration_file = response4['Body'].read()
                    model_isotonic_calibration = pickle.loads(isotonic_calibration_file)

                if model_isotonic_calibration != None:
                    model_prob = model_isotonic_calibration.predict_proba(test_data_tfidf.todense())

                elif model_sigmoid_calibration != None:
                    model_prob = model_sigmoid_calibration.predict_proba(test_data_tfidf.todense())

                else:
                    final_progress_value = 200
                    errorString = 'Error running retreiving data from s3. Please contact the IP Group Analytics Team'
                    es_conn.update(index=index_name_gloabals, id=id,
                                   body={"doc": {"final_progress_value": final_progress_value,
                                                 "errorString": errorString}})

                    return JsonResponse({'finalResponse': 'Error'})
            else:
                if 'Deep Learning' in trainedModelName:
                    model_prob = model.predict_proba(testing_data)
                else:
                    model_prob = model.predict_proba(test_data_tfidf.todense())

            progressbar_value += 20

            # classify the patent with rel. threshold < th to "others" class
            model_prob_all = copy.copy(model_prob)
            # model_prob[model_prob < 0.0] = 0.0
            model_prob_new = np.sum(model_prob, axis=1)
            model_predicted[model_prob_new == 0] = 'others'

            trainingDataNumInstances = len(model_predicted)
            trainingDataNumClasses = len(set(model_predicted))
            es_conn.update(index=index_name_gloabals, id=id,
                           body={"doc": {"progressbar_value": progressbar_value,
                                         "trainingDataNumInstances": trainingDataNumInstances,
                                         "trainingDataNumClasses": trainingDataNumClasses}})

            # Update the treeview with the distribution of instances in the training data
            testingstats = []
            for label in set(model_predicted):
                distribution = str(
                    np.round(model_predicted.tolist().count(label) * 100.0 / len(model_predicted) * 1.0, 2)) + '%'
                testingstats.append([label, model_predicted.tolist().count(label), distribution])

            trainingDataStatistics = testingstats
            labels = model.classes_
            labels = sorted(list(set(labels)))
            labels.append(u'others')

            progressbar_value += 20

            # Save the results to an excel worksheet
            progress_text = progress_text + "\nSaving results..."

            if testingDataType == 'Patent':
                workbook = load_workbook('./static/template_patent.xlsx')

            elif testingDataType == 'Journal':

                workbook = load_workbook('./static/template_journal.xlsx')

            resultsSheet = workbook["Results"]
            rawDataSheet = workbook["Raw_Data"]

            # Write results for predicted probabilities and class names to the sheet containing raw data
            # In excel, the indices start from 1, instead of 0
            if testingDataType == 'Patent':
                # Now, the first line is header, so remove the first line
                for row_index in np.arange(len(file_test_proc)):
                    doc = file_test_proc[row_index].split('\t')
                    for column_index in np.arange(8):
                        resultsSheet.cell(row=row_index + 2, column=column_index + 1).value = doc[column_index].strip(
                            '\"')
                    resultsSheet.cell(row=row_index + 2, column=9).value = model_predicted[row_index].strip('\"')

            elif testingDataType == 'Journal':
                for row_index in np.arange(len(file_test_open)):
                    doc = file_test_open[row_index].split('\t')
                    for column_index in np.arange(6):
                        resultsSheet.cell(row=row_index + 2, column=column_index + 1).value = doc[column_index].strip(
                            '\"')
                    resultsSheet.cell(row=row_index + 2, column=7).value = model_predicted[row_index].strip('\"')

            # In the Raw_Data sheet, write the class names starting from column B

            column_header_index = 2
            for cls in model.classes_:
                rawDataSheet.cell(row=1, column=column_header_index).value = cls.strip('\r')
                column_header_index += 1

            # Wirte all the probabilities for each class assgined by the model in the Raw_Data sheet
            for row_index in np.arange(len(model_prob_all)):
                for column_index in np.arange(len(model_prob_all[row_index])):
                    # The first column in template excel file is formula for 'OTHERS',
                    # hence start writing the probability values from second column in the excel sheet
                    rawDataSheet.cell(row=row_index + 2, column=column_index + 2).value = model_prob_all[
                        row_index][column_index]

            # workbook.save(re.sub('.txt', '_Threshold_Analysis.xlsx',gui_parameters['testing_file_name']))
            #
            thresholdAnalysisResultFileName = re.sub('.txt', '_Threshold_Analysis.xlsx', testingFileName)
            # response1 = HttpResponse(content_type='application/ms-excel')
            responseTrain = HttpResponse(content=save_virtual_workbook(workbook),
                                         content_type='application/vnd.ms-excel')
            responseTrain['Content-Disposition'] = 'attachment; filename=' + thresholdAnalysisResultFileName

            progressbar_value += 25
            programRunEndTime = datetime.now()

            timeDifference = relativedelta(programRunEndTime, programRunStartTime)

            programRunStartTimeLabel = "Program run took %d days %d hours %d minutes %d seconds." % (
                timeDifference.days, timeDifference.hours, timeDifference.minutes, timeDifference.seconds)
            progressbarlabel_text = programRunStartTimeLabel
            es_conn.update(index=index_name_gloabals, id=id,
                           body={"doc": {"progressbar_value": progressbar_value,
                                         "progressbarlabel_text": progressbarlabel_text,
                                         "trainingDataStatistics": trainingDataStatistics,
                                         "progress_text": progress_text}})

            progressbar_value += 5
            final_progress_value = 200
            es_conn.update(index=index_name_gloabals, id=id,
                           body={"doc": {"progressbar_value": progressbar_value,
                                         "progress_text": progress_text,
                                         "final_progress_value": final_progress_value}})
        return HttpResponse('successfully executed')
    except Exception as e:
        final_progress_value = 200
        errorString = errorString
        es_conn.update(index=index_name_gloabals, id=id,
                       body={"doc": {"errorString": errorString,
                                     "final_progress_value": final_progress_value}})
        return JsonResponse({'finalResponse': 'Error'})


@csrf_exempt
def userRunModelTrackIL(request):
    try:
        if request.method == 'POST':
            try:
                userName = request.user.username;
                userTrack = request.body.decode('utf-8');
                userTrack = json.loads(userTrack);
                saveProjectName = userTrack['saveProjectName']
                trainedModelName = userTrack['trainedModelName']
                training_file_name = userTrack['training_file_name']
                training_data_type = userTrack['training_data_type']
            except Exception as e:
                print('parsing went wrong', e)

            time = datetime.now()
            time = time.strftime("%I:%M%p on %B %d, %Y")

            update_fields = {
                'username': userName,
                'saveProjectName': saveProjectName,
                'trainedModelName': trainedModelName,
                'training_file_name': training_file_name,
                'training_data_type': training_data_type,
                'time': time + ' UTC  Time'
            }
            index_name = 'userrunmodeltrackil'
            if es_conn.indices.exists(index_name):
                es_conn.create(index=index_name, doc_type='_doc', body=update_fields, id=uuid.uuid4())
            else:
                save_response = es_conn.indices.create(index=index_name, ignore=400)
                es_conn.create(index=index_name, doc_type='_doc', body=update_fields, id=uuid.uuid4())

        return HttpResponse('sucess')
    except Exception as e:
        return HttpResponse(
            "Error running the program." + str(e))


@csrf_exempt
def makePredictionsForUnsupervisedLearning(request):
    global response
    try:
        if request.method == 'GET':
            return response
        elif request.method == "POST":
            userName = request.user.username;
            gui_parameters = request.POST.getlist('inputData')[0]
            get_value = json.loads(gui_parameters);
            finaltestingData = request.FILES.getlist('testFile')
            # training_data = request.FILES.getlist('file').read().decode("ISO-8859-1")
            testing_data = Process_All_Files(finaltestingData)
            saveProjectName = get_value['saveProjectName']
            trainedModelName = get_value['trainedModelName']
            testingFileName = get_value['testing_file_name']
            testingDataType = get_value['testing_data_type']
            current_tab = 4

            programRunStartTime = datetime.now()
            programRunStartTimeLabel = 'Progress: Program run started at ' + programRunStartTime.strftime(
                "%I:%M%p on %B %d, %Y") + ' (UTC time). '

            progressbarlabel_text = programRunStartTimeLabel

            progress_text = ''
            progress_text = progress_text + '-' * 75 + '\n' + "Program run started at " + programRunStartTime.strftime(
                "%I:%M%p on %B %d, %Y") + " (UTC time)." + \
                            '\n' + '-' * 75 + '\n' + "Starting unsupervised learning process...\n"
            # trainedModelName = 'K-means clustering'
            numberOfClusters = int(re.search('NumClusters=(.*)_TopWords', trainedModelName).group(1))
            numberOfClusters = numberOfClusters

            number_of_top_words = int(trainedModelName.split("TopWords=")[1].replace('/', ''))

            progressbar_maximum = 60 + 10 * numberOfClusters

            progressbar_value = 10
            index_name_gloabals = 'apolloglobalsunsupervised'
            query = {"query": {"bool": {"must": {"match": {"username.keyword": userName}}}}}
            res = es_conn.search(index=index_name_gloabals, body=query)
            id = res['hits']['hits'][0]['_id']
            es_conn.update(index=index_name_gloabals, id=id,
                           body={"doc": {"progressbar_maximum": progressbar_maximum, "current_tab": current_tab,
                                         "progressbar_value": progressbar_value,
                                         "testingFileName": testingFileName, "trainedModelName": trainedModelName,
                                         "testingDataType": testingDataType, "progress_text": progress_text
                               , "progressbarlabel_text": progressbarlabel_text, "numberOfClusters": numberOfClusters,
                                         "number_of_top_words": number_of_top_words}})

            # historyFilename = ''
            # if os.name == 'nt':
            #    historyFilename = 'C:/Users/manali.s/Desktop/Classification_Projects/'
            # else:
            #    historyFilename = '/data/Classification_Projects/'

            # historyFilename += comboboxAvailableProjects.get() + '/history.txt'

            # Set directory for the output folder
            # output_folder = os.path.dirname(testing_file_name)

            # Read the paper data from the file
            # [type, title, abstract, author, affiliation, year]
            # The code for patent and journal testing data is different because it required different preprocessing

            user_defined_stopwords = []

            # Load tfidf vectorizer
            key_value = 'classificationprojects/'
            key_value += saveProjectName + '/unsupervised/'
            key_value += trainedModelName + '/'

            s3 = boto3.client('s3')
            response1 = s3.get_object(Bucket=AWS_STORAGE_BUCKET_NAME,
                                      Key=key_value + 'tfidf_vect.pkl')
            # for line in contents.splitlines():
            # Load the tfidf vectorizer pickle file that was previously saved in S3
            tfidfVectorizer = response1['Body'].read()
            tfidf_vect = pickle.loads(tfidfVectorizer)
            progressbar_value += 10
            es_conn.update(index=index_name_gloabals, id=id,
                           body={"doc": {"progressbar_value": progressbar_value}})

            if testingDataType == 'Patent':

                # Read the test patent data
                file_test_open = testing_data
                file_test_open = file_test_open.split('\n')  # split by new line
                file_test_open = list(filter(None, file_test_open))  # delete empty lines

                # Now, the first line is header, so remove the first line
                file_test_open = file_test_open[1:]

                progress_text = progress_text + "Preprocessing unlabeled data...\n"
                es_conn.update(index=index_name_gloabals, id=id,
                               body={"doc": {"progress_text": progress_text}})

                # Build the stop words
                stops = stopwords

                aux_stops = './static/AuxStops.txt'
                aux_stops = open(aux_stops, 'r').read()
                aux_stops = re.sub("[^a-zA-Z ]", "   ", aux_stops)  # remove non-alphanumeric
                aux_stops = " ".join(aux_stops.split())  # split by any whitespace and rejoin w/ space
                aux_stops = aux_stops.split(' ')
                aux_stops = list(filter(None, aux_stops))

                # append auxiliary stops
                stops = stops + aux_stops

                # Bulid stopterm dictionary
                stopterms = {}
                for stop in stops:
                    if stop in stopterms:
                        stopterms[stop] += 1
                    else:
                        stopterms[stop] = 1

                (file_test_proc, file_test_stem, file_test_stop_words_removed) = preprocess_collection(file_test_open,
                                                                                                       stopterms, False,
                                                                                                       progress_text)
                file_test_proc = list(filter(None, file_test_proc))
                file_test = list(filter(None, file_test_stem))

                title_test = [doc.split('\t')[1] for doc in file_test]
                abstract_test = [doc.split('\t')[2] for doc in file_test]
                claim_test = [doc.split('\t')[3] for doc in file_test]
                test_data = [' '.join(doc) for doc in zip(title_test, abstract_test, claim_test)]

                title_test_stop_words_removed = [doc.split('\t')[1] for doc in file_test_stop_words_removed]
                abstract_test_stop_words_removed = [doc.split('\t')[2] for doc in file_test_stop_words_removed]
                claim_test_stop_words_removed = [doc.split('\t')[3] for doc in file_test_stop_words_removed]
                unlabeled_test_stop_words_removed = [' '.join(doc) for doc in
                                                     zip(title_test_stop_words_removed,
                                                         abstract_test_stop_words_removed,
                                                         claim_test_stop_words_removed)]

                # convert text data to tfidf
                test_data_tfidf = tfidf_vect.transform(test_data)

                progress_text = progress_text + "\nRemoved duplicates and preprocessed " + str(
                    len(test_data)) + " documents."
                progressbar_value += 20

                # Load the model from pickle file
                response2 = s3.get_object(Bucket=AWS_STORAGE_BUCKET_NAME,
                                          Key=key_value + 'trainedModel.pkl')
                # 1. Load the model
                model_file = response2['Body'].read()
                model = pickle.loads(model_file)

                # determine the best model based on evaluating several models
                automatic_mode = True
                progress_text = progress_text + "\nPerforming clustering on the data..."

                output_filename = re.sub('.txt', '_Results_Topic_Modeling.txt', testingFileName)
                es_conn.update(index=index_name_gloabals, id=id,
                               body={"doc": {"progress_text": progress_text, "progressbar_value": progressbar_value}})

                # Predict clusters of documents
                clusters = model.predict(test_data_tfidf)

                test_patent_others = {'content': test_data, 'file_others': file_test_proc,
                                      'content_stop_words_removed': unlabeled_test_stop_words_removed,
                                      'cluster': clusters}

                # 1. Load the previous data frame to run topic modeling again on each set of document clusters

                response3 = s3.get_object(Bucket=AWS_STORAGE_BUCKET_NAME,
                                          Key=key_value + 'topicModelingFrame.pkl')
                topicModelingFrame = response3['Body'].read()
                oldFrame = pickle.loads(topicModelingFrame)

                # The new documents are clustered using the old model, however, the topics have to be re-computed for each cluster based on new documents.

                frame = pd.DataFrame(test_patent_others, index=[clusters],
                                     columns=['content', 'file_others', 'content_stop_words_removed', 'cluster'])

                # Append old frame to new frame and recompute the topics for clusters.
                # Alternatively, check if IP Managers would like to see the new documents belonging to the same clusters; but this might not be the expected behavior.
                frame = frame.append(oldFrame)

                clustering_successful = False
                # output lda other topics

                # Need to extract number of top words from path to save the model

                progressbar_value += 10
                es_conn.update(index=index_name_gloabals, id=id,
                               body={"doc": {"progressbar_value": progressbar_value}})
                # with open(output_filename, 'w+') as fout_others:
                for no in range(numberOfClusters):
                    try:
                        # sometimes, there is no document in the group, so handle that case with try and except
                        patent_group = frame.groupby(frame['cluster']).get_group(no)
                    except:
                        # continue, because there is no document in this cluster. Move on to the topic modeling for next cluster
                        continue

                    patent_tac = patent_group.ix[:, 0].tolist()
                    patent_org = patent_group.ix[:, 1].tolist()

                    lda_tf_vect = TfidfVectorizer(max_df=0.8, min_df=1,
                                                  max_features=200000,
                                                  ngram_range=(1, 5),
                                                  use_idf=True,
                                                  stop_words='english')
                    tf = None
                    try:
                        tf = lda_tf_vect.fit_transform(patent_tac)

                    except Exception as e:
                        lda_tf_vect = TfidfVectorizer(max_df=1.0, min_df=1,
                                                      max_features=200000,
                                                      ngram_range=(1, 5),
                                                      use_idf=True,
                                                      stop_words='english')
                        tf = lda_tf_vect.fit_transform(patent_tac)

                    # LDA Model
                    lda = LatentDirichletAllocation(n_components=1, max_iter=20,
                                                    learning_method='online',
                                                    learning_offset=50,
                                                    random_state=0).fit(tf)

                    lda_feature_names = lda_tf_vect.get_feature_names()

                    lda_topics = get_topic_list(lda, lda_feature_names, number_of_top_words)
                    clusterTopicsAndCounts = []
                    clusterTopicsAndCounts.append([len(patent_tac), lda_topics[0]])

                    doc_topic = lda.transform(tf)
                    doc_topic_index = doc_topic.argmax(axis=1)
                    fout_others = ''
                    for doc, doc_topic_i in zip(patent_org, doc_topic_index):
                        fout_others += '\t'.join(
                            [doc.strip('\r').strip('\n'), lda_topics[doc_topic_i].strip('\r').strip('\n')]) + '\n'

                progressbar_value += 10;

                progress_text = progress_text + '\nTopic extraction and clustering completed.'

                programRunEndTime = datetime.now();

                timeDifference = relativedelta(programRunEndTime, programRunStartTime);

                programRunStartTimeLabel = "Program run took %d days %d hours %d minutes %d seconds." % (
                    timeDifference.days, timeDifference.hours, timeDifference.minutes, timeDifference.seconds);
                progressbarlabel_text = programRunStartTimeLabel;

                progress_text = progress_text + programRunStartTimeLabel + '\n' + '*' * 95 + '\n';
                progressbar_value += 10;
                final_progress_value = 200
                es_conn.update(index=index_name_gloabals, id=id,
                               body={
                                   "doc": {"progress_text": progress_text, "progressbar_value": progressbar_value,
                                           "progressbarlabel_text": progressbarlabel_text,
                                           "final_progress_value": final_progress_value}})
                response = HttpResponse(content=fout_others, content_type='text/plain');
                response['Content-Disposition'] = 'attachment; filename=' + output_filename

                return HttpResponse('successfully executed')

            elif testingDataType == 'Journal':

                file_test_open = testing_data
                file_test_open = file_test_open.split('\n')  # split by new line
                file_test_open = list(filter(None, file_test_open))  # delete empty lines

                # Now, the first line is header, so remove the first line
                file_test_open = file_test_open[1:]

                progress_text = progress_text + "\nPreprocessing unlabeled data..."
                es_conn.update(index=index_name_gloabals, id=id,
                               body={
                                   "doc": {"progress_text": progress_text}})

                # Remove the duplicated document based on "title"
                file_test_open = dedup_collection_journal(file_test_open, 1, 2)

                # preprocessing for scoupus data
                file_test_proc = preprocess_collection_journal(file_test_open)

                test_data = []
                unlabeled_test_stop_words_removed = []

                for doc in file_test_open:
                    stop_and_stem_document_title, stop_document_title = stop_and_stem_journal_2(doc.split('\t')[1],
                                                                                                user_defined_stopwords)
                    stop_and_stem_document_abstract, stop_document_abstract = stop_and_stem_journal_2(
                        doc.split('\t')[2],
                        user_defined_stopwords)
                    test_data.append(' '.join([stop_and_stem_document_title, stop_and_stem_document_abstract]))
                    unlabeled_test_stop_words_removed.append(' '.join([stop_document_title, stop_document_abstract]))

                # convert text data to tfidf
                test_data_tfidf = tfidf_vect.transform(test_data)

                progress_text = progress_text + "\nRemoved duplicates and preprocessed " + str(
                    len(test_data)) + " documents."
                progressbar_value += 20
                es_conn.update(index=index_name_gloabals, id=id,
                               body={
                                   "doc": {"progress_text": progress_text, "progressbar_value": progressbar_value}})
                # Load the model from pickle file
                response2 = s3.get_object(Bucket=AWS_STORAGE_BUCKET_NAME,
                                          Key=key_value + 'trainedModel.pkl')
                # 1. Load the model
                model_file = response2['Body'].read()
                model = pickle.loads(model_file)

                # determine the best model based on evaluating several models
                automatic_mode = True
                progress_text = progress_text + "\nPerforming clustering on the data..."
                es_conn.update(index=index_name_gloabals, id=id,
                               body={
                                   "doc": {"progress_text": progress_text}})

                output_filename = re.sub('.txt', '_Results_Topic_Modeling.txt', testingFileName)

                # Predict clusters of documents
                clusters = model.predict(test_data_tfidf)

                test_patent_others = {'content': test_data, 'file_others': file_test_proc,
                                      'content_stop_words_removed': unlabeled_test_stop_words_removed,
                                      'cluster': clusters}

                # 1. Load the previous data frame to run topic modeling again on each set of document clusters
                response3 = s3.get_object(Bucket=AWS_STORAGE_BUCKET_NAME,
                                          Key=key_value + 'topicModelingFrame.pkl')
                topicModelingFrame = response3['Body'].read()
                oldFrame = pickle.loads(topicModelingFrame)

                # The new documents are clustered using the old model, however, the topics have to be re-computed for each cluster based on new documents.

                frame = pd.DataFrame(test_patent_others, index=[clusters],
                                     columns=['content', 'file_others', 'content_stop_words_removed', 'cluster'])

                # Append old frame to new frame and recompute the topics for clusters.
                # Alternatively, check if IP Managers would like to see the new documents belonging to the same clusters; but this might not be the expected behavior.
                frame = frame.append(oldFrame)

                clustering_successful = False
                # output lda other topics

                # Need to extract number of top words from path to save the model

                progressbar_value += 10
                es_conn.update(index=index_name_gloabals, id=id,
                               body={
                                   "doc": {"progressbar_value": progressbar_value}})

                # with open(output_filename, 'w+') as fout_others:
                for no in range(numberOfClusters):
                    try:
                        # sometimes, there is no document in the group, so handle that case with try and except
                        patent_group = frame.groupby(frame['cluster']).get_group(no)
                    except:
                        # continue, because there is no document in this cluster. Move on to the topic modeling for next cluster
                        continue

                    patent_tac = patent_group.ix[:, 0].tolist()
                    patent_org = patent_group.ix[:, 1].tolist()

                    lda_tf_vect = TfidfVectorizer(max_df=0.8, min_df=1,
                                                  max_features=200000,
                                                  ngram_range=(1, 5),
                                                  use_idf=True,
                                                  stop_words='english')
                    tf = None
                    try:
                        tf = lda_tf_vect.fit_transform(patent_tac)

                    except Exception as e:
                        lda_tf_vect = TfidfVectorizer(max_df=1.0, min_df=1,
                                                      max_features=200000,
                                                      ngram_range=(1, 5),
                                                      use_idf=True,
                                                      stop_words='english')
                        tf = lda_tf_vect.fit_transform(patent_tac)

                    # LDA Model
                    lda = LatentDirichletAllocation(n_components=1, max_iter=20,
                                                    learning_method='online',
                                                    learning_offset=50,
                                                    random_state=0).fit(tf)

                    lda_feature_names = lda_tf_vect.get_feature_names()

                    lda_topics = get_topic_list(lda, lda_feature_names, number_of_top_words)
                    clusterTopicsAndCounts = []
                    clusterTopicsAndCounts.append([len(patent_tac), lda_topics[0]])

                    doc_topic = lda.transform(tf)
                    doc_topic_index = doc_topic.argmax(axis=1)
                    fout_others = ''
                    for doc, doc_topic_i in zip(patent_org, doc_topic_index):
                        fout_others += '\t'.join(
                            [doc.strip('\r').strip('\n'), lda_topics[doc_topic_i].strip('\r').strip('\n')]) + '\n'

                    progressbar_value += 10

                progress_text = progress_text + '\nTopic extraction and clustering completed.'
                es_conn.update(index=index_name_gloabals, id=id,
                               body={
                                   "doc": {"progress_text": progress_text, "progressbar_value": progressbar_value}})

                # Load the topic modeling results in the treeview
                # Compute the columns: (i) # instances, (ii) Topics extracted
                # populateOutputTreeviewWithTopicModelingResults(clusterTopicsAndCounts)

                programRunEndTime = datetime.now()

                timeDifference = relativedelta(programRunEndTime, programRunStartTime)

                programRunStartTimeLabel = "Program run took %d days %d hours %d minutes %d seconds." % (
                    timeDifference.days, timeDifference.hours, timeDifference.minutes, timeDifference.seconds)
                progressbarlabel_text = programRunStartTimeLabel

                progress_text = progress_text + programRunStartTimeLabel + '\n' + '*' * 95 + '\n'

                progressbar_value += 10
                final_progress_value = 200
                es_conn.update(index=index_name_gloabals, id=id,
                               body={
                                   "doc": {"progress_text": progress_text, "progressbar_value": progressbar_value,
                                           "progressbarlabel_text": progressbarlabel_text,
                                           "final_progress_value": final_progress_value}})
                response = HttpResponse(content=fout_others,
                                        content_type='text/plain')
                response['Content-Disposition'] = 'attachment; filename=' + output_filename

                return HttpResponse('successfully executed')
    except Exception as e:
        final_progress_value = 200
        es_conn.update(index=index_name_gloabals, id=id,
                       body={
                           "doc": {"final_progress_value": final_progress_value}})
        return JsonResponse({'finalResponse': 'Error'})


def stop_and_stem_journal_2(file_t, user_defined_stopwords):
    '''
    Input: file_t: a text string, stopterms: a dictionary of stop terms
    Output: file_stem: a list of stopped and stemmed terms
    '''
    try:
        stopterms = build_stopterms_journal(user_defined_stopwords)
        # remove the patent specific terms
        file_t = file_t.lower()  # lowercase all
        file_t = re.sub("[^a-zA-Z0-9 ]", "   ", file_t)  # remove non-alphanumeric
        file_t = re.sub("\s[0-9]+$", '', file_t)
        file_t = re.sub("\s[0-9]+\s", ' ', file_t)
        file_t = re.sub("^[0-9]+\s", '', file_t)
        file_t = re.sub("androids*", "antroid", file_t)
        file_t = re.sub("andes", "antes", file_t)
        file_t = re.sub("and[0-9a-z]+", "", file_t)
        file_t = re.sub("antroid", "android", file_t)
        file_t = re.sub("antes", "andes", file_t)
        file_t = re.sub("including[0-9a-z]+", "", file_t)
        file_t = re.sub("wherein[0-9a-z]+", "", file_t)
        file_t = re.sub("comprising[0-9a-z]+", "", file_t)
        formula_chk0 = re.findall(" formula | math ", file_t)
        formula_chk1 = re.findall(" tail ", file_t)
        formula_chk2 = re.findall(" lead ", file_t)
        if len(formula_chk0) > 0 and len(formula_chk1) > 0 and len(formula_chk2) > 0:
            file_t = re.sub(" formula | math ", " ", file_t)
            file_t = re.sub(" tail ", " ", file_t)
            file_t = re.sub(" lead ", " ", file_t)
        file_t = " ".join(file_t.split())  # split by any whitespace and rejoin w/ space
        file_t = file_t.split(" ")  # split by space

        # remove the stop terms in the text
        file_stop = []  # initialize list
        for term in file_t:
            if term not in stopterms:
                file_stop.append(term)

        # stem using porter algorithm
        file_stem = []  # initialize list
        for term in file_stop:
            try:
                term = wn().lemmatize(term)
            except:
                pass
            term = str(term)
            file_stem.append(term)
        file_stem = ' '.join(file_stem)
        file_stop = ' '.join(file_stop)
        return file_stem, file_stop
    except Exception as e:
        return HttpResponse(
            "Error running the program." + str(e))


@csrf_exempt
def incrementalsupervisedlearning(request):
    # Assuming that the training data uploaded by the user is available in the variable

    progress_text = ""
    progressbarlabel_text = ""
    global training_data

    try:
        # Assuming that the client-side has already selected the options before running the program.
        # Assuming that the request from the client side will have all the fields necessary for running the program.
        if request.method == "POST":
            gui_parameters = request.POST.getlist('inputData')[0]
            gui_parameters = json.loads(gui_parameters);
            finalTrainingData = request.FILES.getlist('trainFile')
            # training_data = request.FILES.getlist('file').read().decode("ISO-8859-1")
            training_data = Process_All_Files(finalTrainingData)
            trainingFileName = gui_parameters['training_file_name']
            trainingDataType = gui_parameters['training_data_type']
            selectedProjectName = gui_parameters['saveProjectName']
            selectedModelName = gui_parameters['trainedModelName']
            current_tab = 5
            username = request.user.username;
            progressbar_maximum = 200
            progressbar_value = 0

            # Set the text in progressbarlabel
            programRunStartTime = datetime.now()
            programRunStartTimeLabel = 'Progress: Program run started at ' + programRunStartTime.strftime(
                "%I:%M%p on %B %d, %Y") + ' (UTC time). '

            progressbarlabel_text = programRunStartTimeLabel
            progress_text = progress_text + '-' * 75 + '\n' + "Program run started at " + programRunStartTime.strftime(
                "%I:%M%p on %B %d, %Y") + " (UTC time).\n" + '-' * 75 + '\n' + "Starting incremental learning process..."

            index_name_gloabals = 'apolloglobals'
            query = {"query": {"bool": {"must": {"match": {"username.keyword": username}}}}}
            res = es_conn.search(index=index_name_gloabals, body=query)
            id = res['hits']['hits'][0]['_id']
            es_conn.update(index=index_name_gloabals, id=id,
                           body={"doc": {"progressbar_maximum": progressbar_maximum, "current_tab": current_tab,
                                         "progressbar_value": progressbar_value,
                                         "trainingFileName": trainingFileName, "trainedModelName": selectedModelName,
                                         "trainingDataType": trainingDataType, "progress_text": progress_text
                               , "progressbarlabel_text": progressbarlabel_text}})
            historyFilename = 'history.txt'
            append_text_to_history_file = ""

            # The code for loading and pre-processing the data is different for patent and journal data
            if trainingDataType == 'Patent':
                file_sample_open = training_data
                file_sample_open = file_sample_open.split('\n')  # split by new line
                file_sample_open = list(filter(None, file_sample_open))  # delete empty lines

                # Now, the first line is header, so remove the first line
                file_sample_open = file_sample_open[1:]

                progress_text = progress_text + "\nFound " + str(
                    len(file_sample_open)) + " patents! \nPreprocessing documents...\n"

                # Set value of progressbar to 10 once the training dataset is loaded
                progressbar_value += 10
                es_conn.update(index=index_name_gloabals, id=id,
                               body={"doc": {"progressbar_value": progressbar_value,
                                             "progress_text": progress_text}})

                # Build the stop words
                stops = stopwords

                aux_stops = './static/AuxStops.txt'

                aux_stops = open(aux_stops, 'r').read()
                aux_stops = re.sub("[^a-zA-Z ]", "   ", aux_stops)  # remove non-alphanumeric
                aux_stops = " ".join(aux_stops.split())  # split by any whitespace and rejoin w/ space
                aux_stops = aux_stops.split(' ')
                aux_stops = list(filter(None, aux_stops))

                # append auxiliary stops
                stops = stops + aux_stops

                # append user-provided stop words. This is not available for incremental learning framework, so keep it empty.
                user_defined_stopwords = []
                stops = stops + user_defined_stopwords

                # Bulid stopterm dictionary
                stopterms = {}
                for stop in stops:
                    if stop in stopterms:
                        stopterms[stop] += 1
                    else:
                        stopterms[stop] = 1

                # Preprocess the sample file
                # Also, we need to check the original data file for any duplicates in the new data

                # Load the latest training data for this model and use the application number column to deduplicate the documents from previous patents data

                s3 = boto3.client('s3')
                key_value = 'classificationprojects/'
                key_value += selectedProjectName + '/supervised/'
                key_value += selectedModelName + '/'

                try:
                    response = s3.get_object(Bucket=AWS_STORAGE_BUCKET_NAME,
                                             Key=key_value + 'training_data_patents.txt')  # Need to change to training_data_patent after impementation of save_both_exisitng_model

                    oldPatentsFile = response['Body'].read()
                    # The training data already exists, and we need to append the new data to this data as well.
                    file_old_training_data_open = oldPatentsFile.decode('utf-8')
                    training_data_IL = training_data + '\n' + file_old_training_data_open.split('\n', 1)[1]
                    file_old_training_data_open = file_old_training_data_open.split('\n')  # split by new line
                    file_old_training_data_open = list(filter(None, file_old_training_data_open))  # delete empty lines

                    # Now, the first line is header, so remove the first line
                    file_old_training_data_open = file_old_training_data_open[1:]

                    existing_application_numbers = [doc.split('\t')[4].upper() for doc in file_old_training_data_open]
                except botocore.exceptions.ClientError as e:
                    if e.response['Error']['Code'] == "404":
                        # There is no existing patent training data, so no need to deduplicate based on previous data.
                        existing_application_numbers = []
                    elif e.response['Error']['Code'] == "NoSuchKey":
                        existing_application_numbers = []
                        pass
                    else:
                        existing_application_numbers = []
                        pass

                # Remove the duplicated documents based on "Application number"
                (file_sample_proc, file_sample_stem,
                 file_test_stop_words_removed) = preprocess_collection_incremental_learning(username, file_sample_open,
                                                                                            stopterms,
                                                                                            existing_application_numbers,
                                                                                            True, progress_text)
                file_sample = list(filter(None, file_sample_stem))

                if len(file_sample) < 1:
                    progressbar_value = 180
                    # Simply display the error message below and exit.
                    final_progress_value = 200
                    errorString = 'The additional training data file does not contain any new Patents for training the model. \nCannot perform incremental learning in this case.'
                    es_conn.update(index=index_name_gloabals, id=id,
                                   body={"doc": {"final_progress_value": final_progress_value,
                                                 "errorString": errorString}})
                    return JsonResponse({'finalResponse': "Error"})

                title_samples = [doc.split('\t')[1].strip('\r').strip('\n') for doc in file_sample]
                abstract_samples = [doc.split('\t')[2].strip('\r').strip('\n') for doc in file_sample]
                claim_samples = [doc.split('\t')[3].strip('\r').strip('\n') for doc in file_sample]
                label_samples = [doc.split('\t')[8].lower().strip('\r').strip('\n') for doc in file_sample]
                labels = sorted(list(set(label_samples)))

                train_data = [' '.join(doc) for doc in zip(title_samples, abstract_samples, claim_samples)]
                train_target = label_samples

                # End patent training data

            elif trainingDataType == 'Journal':
                file_sample_open = training_data
                file_sample_open = file_sample_open.split('\n')  # split by new line
                file_sample_open = list(filter(None, file_sample_open))  # delete empty lines

                # Now, the first line is header, so remove the first line
                file_sample_open = file_sample_open[1:]

                progress_text = progress_text + "\nFound " + str(
                    len(file_sample_open)) + " documents! \nPreprocessing documents...\n"

                # Set value of progressbar to 10 once the training dataset is loaded
                progressbar_value = 10

                # Load the latest training data for this model and use the title and abstract columns to deduplicate the documents from previous journals data
                es_conn.update(index=index_name_gloabals, id=id,
                               body={"doc": {"progressbar_value": progressbar_value,
                                             "progress_text": progress_text}})
                # load_data_from_S3()

                s3 = boto3.client('s3')
                key_value = 'classificationprojects/'
                key_value += selectedProjectName + '/supervised/'
                key_value += selectedModelName + '/'

                try:
                    response = s3.get_object(Bucket=AWS_STORAGE_BUCKET_NAME,
                                             Key=key_value + 'training_data_journals.txt')
                    oldJournalsFile = response['Body'].read()
                    # The training data already exists, and we need to append the new data to this data as well.
                    # file_old_training_data_open = codecs.open(oldJournalsFile).read() # open file
                    file_old_training_data_open = oldJournalsFile.decode('utf-8')
                    training_data_IL = training_data + '\n' + file_old_training_data_open.split('\n', 1)[1]
                    file_old_training_data_open = file_old_training_data_open.split('\n')  # split by new line
                    file_old_training_data_open = list(filter(None, file_old_training_data_open))  # delete empty lines

                    # Now, the first line is header, so remove the first line
                    file_old_training_data_open = file_old_training_data_open[1:]
                    existing_title_and_abstracts = ['\t'.join([doc.split('\t')[1], doc.split('\t')[2]]) for doc in
                                                    file_old_training_data_open]

                except botocore.exceptions.ClientError as e:
                    if e.response['Error']['Code'] == "404":
                        # There is no existing patent training data, so no need to deduplicate based on previous data.
                        existing_title_and_abstracts = []
                    elif e.response['Error']['Code'] == "NoSuchKey":
                        existing_title_and_abstracts = []
                        pass
                    else:
                        existing_title_and_abstracts = []
                        pass

                        # Remove the duplicated documents based on "title"
                file_sample_open_journal = dedup_collection_journal_incremental_learning(file_sample_open,
                                                                                         existing_title_and_abstracts,
                                                                                         1, 2)

                if len(file_sample_open_journal) < 1:
                    progressbar_value = 180
                    final_progress_value = 200
                    errorString = 'The additional training data file does not contain any new journals for training the model. \nCannot perform incremental learning in this case.'
                    es_conn.update(index=index_name_gloabals, id=id,
                                   body={"doc": {"final_progress_value": final_progress_value,
                                                 "progressbar_value": progressbar_value,
                                                 "errorString": errorString}})
                    return JsonResponse({'finalResponse': "Error"})

                # Preprocessing for scoupus data
                file_sample_open_journal = preprocess_collection_journal(file_sample_open_journal)

                # Take the stopwords from the GUI and add them to the stopwords list. This is not available for incremental learning framework.
                user_defined_stopwords = []

                file_sample_data = [' '.join([stop_and_stem_journal(doc.split('\t')[1], user_defined_stopwords)
                                                 , stop_and_stem_journal(doc.split('\t')[2], user_defined_stopwords)
                                              ]) for doc in file_sample_open_journal]

                # Training Phase
                label_samples = [doc.split('\t')[-1].lower().strip('\r').strip('\n') for doc in
                                 file_sample_open_journal]
                labels = sorted(list(set(label_samples)))

                train_data = file_sample_data
                train_target = label_samples

                # end journal training data preprocessing

            progress_text = progress_text + "Removed duplicates and preprocessed " + str(
                len(train_data)) + " documents."

            # Increment the value of progressbar by 20 once the training dataset is loaded
            progressbar_value += 20
            es_conn.update(index=index_name_gloabals, id=id,
                           body={"doc": {"progressbar_value": progressbar_value,
                                         "progress_text": progress_text}})
            # Check if the training data contains any new class. If yes, then the model needs to be trained from scratch
            # Warn user to make sure that the new class wasn't added mistakenly

            # Get the classes from the trainingDataStatistics object loaded from S3 buket
            INDEX_NAME = 'savemodelsupervised'

            query = {"query": {"bool": {"must": [{"match": {"saveProjectName.keyword": selectedProjectName}},
                                                 {"match": {
                                                     "model_data.trainedModelName.keyword": selectedModelName}}]}}}
            res = es_conn.search(index=INDEX_NAME, body=query)
            trainingDataStatisticsFromS3 = res['hits']['hits'][0]['_source']['model_data']['trainingDataTables']
            # response = s3.get_object(Bucket=AWS_STORAGE_BUCKET_NAME,
            #                          Key=key_value + 'trainingDataStatistics.pkl')
            #
            # trainingDataStatisticsFromS3 = response['Body'].read()
            trainingDataStatistics = json.loads(trainingDataStatisticsFromS3)
            currentModelClasses = [data[0] for data in trainingDataStatistics]
            newclassString = ''
            if not set(train_target).issubset(set(currentModelClasses)):
                # Get a list of new classes
                newClassesList, oldClasses = set(train_target), set(currentModelClasses)
                newClasses = []
                for cls in newClassesList:
                    if cls not in oldClasses:
                        newclassString = newclassString + '\'' + cls.strip('\r').strip(
                            '\n') + '\' '
                        newClasses.append(cls)

                progressbar_value_IL = progressbar_value
                progressbar_value = 0
                train_data_IL = train_data
                train_target_IL = train_target
                es_conn.update(index=index_name_gloabals, id=id,
                               body={"doc": {"progressbar_value": progressbar_value_IL,
                                             "train_data": train_data_IL,
                                             "train_target": train_target_IL}})
                if 'accuracy' in selectedModelName:
                    targetPerformanceMeasure = 'accuracy'
                elif 'auc' in selectedModelName:
                    targetPerformanceMeasure = 'auc'
                elif 'macro_f1' in selectedModelName:
                    targetPerformanceMeasure = 'macro_f1'
                elif 'macro_precision' in selectedModelName:
                    targetPerformanceMeasure = 'macro_precision'
                elif 'macro_recall' in selectedModelName:
                    targetPerformanceMeasure = 'macro_recall'
                targetPerformanceMeasure_IL = targetPerformanceMeasure
                es_conn.update(index=index_name_gloabals, id=id,
                               body={"doc": {"targetPerformanceMeasure_IL": targetPerformanceMeasure_IL}})
                # Show the message below. If user presses "Yes", then call another function to retrain the model from scratch. Otherwise, do nothing further.
                return JsonResponse(
                    {'finalResponse': 'retrainModelFromScratch ', 'newclassString': newclassString})

            # If no new classes are found, continue with model training as usual.
            # For multinomial naive Bayes, incremental learning produces bad results. For now, the workaround is to retrain the model from scratch

            if 'Multinomial_Naive_Bayes' in selectedModelName:
                if 'accuracy' in selectedModelName:
                    targetPerformanceMeasure = 'accuracy'
                elif 'auc' in selectedModelName:
                    targetPerformanceMeasure = 'auc'
                elif 'macro_f1' in selectedModelName:
                    targetPerformanceMeasure = 'macro_f1'
                elif 'macro_precision' in selectedModelName:
                    targetPerformanceMeasure = 'macro_precision'
                elif 'macro_recall' in selectedModelName:
                    targetPerformanceMeasure = 'macro_recall'

                # F1, recall, precision
                successful = retrainModelFromScratch(train_data, train_target, selectedProjectName, selectedModelName,
                                                     targetPerformanceMeasure, programRunStartTime, False, username)

                if successful == True:
                    # Append the patent data to patent training file, and journal data to journal training file
                    if trainingDataType == 'Patent':
                        saveTrainingFileKey = key_value + 'training_data_patents.txt'
                    elif trainingDataType == 'Journal':
                        saveTrainingFileKey = key_value + 'training_data_journals.txt'

                    try:
                        response = s3.get_object(Bucket=AWS_STORAGE_BUCKET_NAME,
                                                 Key=saveTrainingFileKey)
                        oldFile = response['Body'].read()
                        # The training data already exists, and we need to append the new data to this data as well.
                        # file_old_training_data_open = codecs.open(oldFile).read() # open file
                        file_old_training_data_open = oldFile.decode('utf-8')
                        file_old_training_data_open += training_data
                        # write the file back to S3
                        s3.put_object(Body=file_old_training_data_open, Bucket=AWS_STORAGE_BUCKET_NAME,
                                      Key=saveTrainingFileKey)

                    except botocore.exceptions.ClientError as e:
                        if e.response['Error']['Code'] == "404":
                            # There is no existing patent (or journal) training data, so simply put the file into S3
                            s3.put_object(Body=training_data, Bucket=AWS_STORAGE_BUCKET_NAME,
                                          Key=saveTrainingFileKey)
                        elif e.response['Error']['Code'] == "NoSuchKey":
                            pass
                        else:
                            pass
                    final_progress_value = 200
                    es_conn.update(index=index_name_gloabals, id=id,
                                   body={"doc": {"final_progress_value": final_progress_value}})
                    return JsonResponse({'finalResponse': "Incremental learning finished successfully."})
                else:
                    progressbar_value = 100
                    final_progress_value = 200
                    errorString = 'retrainModelFromScratch failed. Please contact the IP Group Analytics Team.' + e
                    es_conn.update(index=index_name_gloabals, id=id,
                                   body={"doc": {"final_progress_value": final_progress_value,
                                                 "progressbar_value": progressbar_value,
                                                 "errorString": errorString}})
                    return JsonResponse({'finalResponse': 'Error'})

            # At this point, proceed with incremental learning for other models
            oldStatsForSelectedModels = trainingDataStatistics

            # Update the training data statistics
            # Add the number of instances for respective classes and recompute the class distribution

            newStatsForSelectedModel = []
            oldStatsForSelectedModelLabels = [i[0] for i in oldStatsForSelectedModels]
            oldStatsForSelectedModelCounts = [int(i[1]) for i in oldStatsForSelectedModels]
            oldStatsTotalInstances = np.sum(oldStatsForSelectedModelCounts)

            newStatsForSelectedModel = []
            for i in range(len(oldStatsForSelectedModelLabels)):
                label = oldStatsForSelectedModelLabels[i]
                distribution = str(np.round((oldStatsForSelectedModelCounts[i] + train_target.count(label)) * 100.0 / (
                        oldStatsTotalInstances + len(train_target)) * 1.0, 2)) + '%'
                newStatsForSelectedModel.append(
                    [label, oldStatsForSelectedModelCounts[i] + train_target.count(label), distribution])

            trainingDataStatistics = newStatsForSelectedModel

            # simply copy the file to the trainingData folder with the timestamp information
            # simply copy the file to the trainingData folder with the timestamp information
            # training_data = pickle.dumps(training_data)
            # s3.put_object(Body=training_data, Bucket=AWS_STORAGE_BUCKET_NAME,
            #               Key=key_value + '/trainingData/' + os.path.basename(
            #                   trainingFileName) + '_' + datetime.now().strftime('%Y-%m-%d %H-%M-%S') +
            #                   os.path.splitext(trainingFileName)[1])

            if trainingDataType == 'Patent':
                saveTrainingFileKey = key_value + 'training_data_patents.txt'
            elif trainingDataType == 'Journal':
                saveTrainingFileKey = key_value + 'training_data_journals.txt'

            # Append the patent data to patent training file, and journal data to journal training file
            # Load the data file from S3, append the data to the file, and then write the updated file back to S3.
            # First, check whether there is existing patent (or journal) training data file that already exists in S3. If exists, then append the existing file, else, create and save the new file.

            try:
                response = s3.get_object(Bucket=AWS_STORAGE_BUCKET_NAME,
                                         Key=saveTrainingFileKey)
                oldFile = response['Body'].read()
                # The training data already exists, and we need to append the new data to this data as well.
                # file_old_training_data_open = codecs.open(oldFile).read() # open file
                file_old_training_data_open = oldFile.decode('utf-8')

                # Remove the first header line
                file_old_training_data_open += training_data.split('\n', 1)[1]
                # write the file back to S3
                s3.put_object(Body=file_old_training_data_open, Bucket=AWS_STORAGE_BUCKET_NAME,
                              Key=saveTrainingFileKey)

            except botocore.exceptions.ClientError as e:
                if e.response['Error']['Code'] == "404":
                    # There is no existing patent (or journal) training data, so simply put the file into S3
                    s3.put_object(Body=training_data, Bucket=AWS_STORAGE_BUCKET_NAME,
                                  Key=saveTrainingFileKey)
                elif e.response['Error']['Code'] == "NoSuchKey":
                    pass
                else:
                    pass

            append_text_to_history_file += '-' * 100 + '\n' + 'username: ' + username + '\n'
            append_text_to_history_file += "Program run started at " + programRunStartTime.strftime(
                "%I:%M%p on %B %d, %Y") + " (UTC time).\n" + '-' * 100 + '\n'

            # Update the history based on whether the updated mdoel is supervised or unsupervised model.

            if trainingDataType == 'Patent':
                append_text_to_history_file += 'Supervised learning model ' + selectedModelName + ' was incrementally updated using the PATENT training data file: ' + trainingFileName + '.' + '\n'
                append_text_to_history_file += str(len(train_target)) + ' training examples were added to the model.\n'
                append_text_to_history_file += 'Total number of examples with which the model is fully trained is: ' + str(
                    oldStatsTotalInstances + len(train_target)) + '\n'
            elif trainingDataType == 'Journal':
                append_text_to_history_file += 'Supervised learning model ' + selectedModelName + ' was incrementally updated using the JOURNAL training data file: ' + trainingFileName + '.' + '\n'
                append_text_to_history_file += str(len(train_target)) + ' training examples were added to the model.\n'
                append_text_to_history_file += 'Total number of examples with which the model is fully trained is: ' + str(
                    oldStatsTotalInstances + len(train_target)) + '\n'

            # Write the new stats for the model to pickle file to save
            dump_trainingDataStatistics = pickle.dumps(trainingDataStatistics)
            s3.put_object(Body=dump_trainingDataStatistics, Bucket=AWS_STORAGE_BUCKET_NAME,
                          Key=key_value + 'trainingDataStats.pkl')

            progress_text = progress_text + "\nTraining Data Statistics have been updated."
            es_conn.update(index=index_name_gloabals, id=id,
                           body={"doc": {"progress_text": progress_text}})

            # Load the tfidf vectorizer that was previously saved
            response = s3.get_object(Bucket=AWS_STORAGE_BUCKET_NAME,
                                     Key=key_value + 'tfidf_vect.pkl')
            tfidfVectorizerFile = response['Body'].read()
            tfidf_vect = pickle.loads(tfidfVectorizerFile)

            ##tf-idf with params
            train_tfidf = tfidf_vect.transform(train_data)

            # Update model with this additional data

            # 1. Load the model

            response = s3.get_object(Bucket=AWS_STORAGE_BUCKET_NAME,
                                     Key=key_value + 'trainedModel.pkl')
            modelFile = response['Body'].read()

            model = pickle.loads(modelFile)

            # Need to update train_target for training, because the new data might not have all the classes on which the model was initially trained.
            train_target_original = model.classes_

            currentModelClasses = [data[0] for data in trainingDataStatistics]

            # Increment the value of progressbar by 10 once the training stats have ben updated
            progressbar_value += 10
            es_conn.update(index=index_name_gloabals, id=id,
                           body={"doc": {"progressbar_value": progressbar_value}})

            # Update the model using partial_fit.
            # The partial_fit method will be different for LR and SVM models

            if ('Support_Vector_Machine' in selectedModelName) or ('Logistic_Regression' in selectedModelName):
                def batches(l, n):
                    for i in np.arange(0, len(l), n):
                        yield l[i:i + n]

                train_tfidf_dense = train_tfidf.toarray()
                n_iter = 25
                np.random.seed(5647)
                shuffledRange = np.arange(len(train_data))
                for n in np.arange(n_iter):
                    np.random.shuffle(shuffledRange)
                    shuffled_train_tfidf = [train_tfidf_dense[i] for i in shuffledRange]
                    shuffled_train_target = [train_target[i] for i in shuffledRange]

                    # Training the model in 10 batches
                    for batch in batches(np.arange(len(shuffled_train_target)), 5):
                        model.partial_fit(shuffled_train_tfidf[batch[0]:batch[-1] + 1],
                                          shuffled_train_target[batch[0]:batch[-1] + 1], classes=train_target_original)

            else:
                if 'Deep_Learning' in selectedModelName:
                    if training_data_IL == None:
                        model.fit(training_data)
                    else:
                        model.fit(training_data_IL)
                else:
                    model.partial_fit(train_tfidf.todense(), train_target, classes=train_target_original)

            progress_text = progress_text + "\nThe model has been incrementally updated with additional training data."
            es_conn.update(index=index_name_gloabals, id=id,
                           body={"doc": {"progress_text": progress_text}})

            # If old training data exists, load the data and preprocess the data
            # Once loaded, append the data to train_data and train_target to train the model from scratch

            all_train_data = []

            progress_text = progress_text + "\nUpdating the five-fold cross validation performances..."
            es_conn.update(index=index_name_gloabals, id=id,
                           body={"doc": {"progress_text": progress_text}})

            try:
                response = s3.get_object(Bucket=AWS_STORAGE_BUCKET_NAME,
                                         Key=key_value + 'training_data_patents.txt')
                oldFile = response['Body'].read()
                # file_sample_open = codecs.open(oldFile) # open file
                file_sample_open = oldFile.decode('utf-8')
                file_sample_open = file_sample_open.split('\n')  # split by new line
                file_sample_open = list(filter(None, file_sample_open))  # delete empty lines

                # Now, the first line is header, so remove the first line
                file_sample_open = file_sample_open[1:]

                progress_text = progress_text + "\nFound existing " + str(
                    len(file_sample_open)) + " patents in prior training data files! \nPreprocessing documents...\n"
                es_conn.update(index=index_name_gloabals, id=id,
                               body={"doc": {"progress_text": progress_text}})
                # Build the stop words
                stops = stopwords

                aux_stops = './static/AuxStops.txt'

                aux_stops = open(aux_stops, 'r').read()
                aux_stops = re.sub("[^a-zA-Z ]", "   ", aux_stops)  # remove non-alphanumeric
                aux_stops = " ".join(aux_stops.split())  # split by any whitespace and rejoin w/ space
                aux_stops = aux_stops.split(' ')
                aux_stops = list(filter(None, aux_stops))

                # append auxiliary stops
                stops = stops + aux_stops

                # append user-provided stop words

                try:
                    response = s3.get_object(Bucket=AWS_STORAGE_BUCKET_NAME,
                                             Key=key_value + 'user_defined_stopwords.pkl')
                    user_defined_stopwords_file = response['Body'].read()
                    user_defined_stopwords = pickle.loads(user_defined_stopwords_file)
                    user_defined_stopwords = list(user_defined_stopwords)
                except botocore.exceptions.ClientError as e:
                    if e.response['Error']['Code'] == "404":
                        # There is no existing patent (or journal) training data, so do nothing
                        user_defined_stopwords = []
                    elif e.response['Error']['Code'] == "NoSuchKey":
                        user_defined_stopwords = []
                        pass
                    else:
                        pass

                stops = stops + user_defined_stopwords

                # Bulid stopterm dictionary
                stopterms = {}
                for stop in stops:
                    if stop in stopterms:
                        stopterms[stop] += 1
                    else:
                        stopterms[stop] = 1

                # Preprocess the sample file
                existing_application_numbers = []
                (file_sample_proc, file_sample_stem, temp) = preprocess_collection_incremental_learning(
                    username, file_sample_open, stopterms, existing_application_numbers, True, progress_text)
                file_sample = list(filter(None, file_sample_stem))

                title_samples = [doc.split('\t')[1] for doc in file_sample]
                abstract_samples = [doc.split('\t')[2] for doc in file_sample]
                claim_samples = [doc.split('\t')[3] for doc in file_sample]
                label_samples = [doc.split('\t')[8].lower() for doc in file_sample]
                labels = sorted(list(set(label_samples)))

                train_data_exisiting_patents = [' '.join(doc) for doc in
                                                zip(title_samples, abstract_samples, claim_samples)]
                train_target_existing_patents = label_samples

                # Append the existing data to new data
                all_train_data = train_data + train_data_exisiting_patents
                all_train_target = train_target + train_target_existing_patents

                # End patent training data

            except botocore.exceptions.ClientError as e:
                if e.response['Error']['Code'] == "404":
                    # There is no existing patent training data, so do nothing
                    pass
                elif e.response['Error']['Code'] == "NoSuchKey":
                    pass
                else:
                    pass

            try:
                response = s3.get_object(Bucket=AWS_STORAGE_BUCKET_NAME,
                                         Key=key_value + 'training_data_journals.txt')
                oldFile = response['Body'].read()
                # file_sample_open = codecs.open(oldFile) # open file
                file_sample_open = oldFile.decode('utf-8')
                file_sample_open = file_sample_open.split('\n')  # split by new line
                file_sample_open = list(filter(None, file_sample_open))  # delete empty lines

                # Now, the first line is header, so remove the first line
                file_sample_open = file_sample_open[1:]

                progress_text = progress_text + "\nFound " + str(
                    len(file_sample_open)) + " documents from prior training data files! \nPreprocessing documents...\n"
                es_conn.update(index=index_name_gloabals, id=id,
                               body={"doc": {"progress_text": progress_text}})
                # Remove the duplicated documents based on "title"
                file_sample_open = dedup_collection_journal(file_sample_open, 1, 2)

                # Preprocessing for scoupus data
                file_sample_open = preprocess_collection_journal(file_sample_open)

                # Take the stopwords from the GUI and add them to the stopwords list
                # append user-provided stop words
                try:
                    response = s3.get_object(Bucket=AWS_STORAGE_BUCKET_NAME,
                                             Key=key_value + 'user_defined_stopwords.pkl')
                    user_defined_stopwords_file = response['Body'].read()
                    user_defined_stopwords = pickle.loads(user_defined_stopwords_file)


                except botocore.exceptions.ClientError as e:
                    if e.response['Error']['Code'] == "404":
                        # There is no existing patent (or journal) training data, so do nothing
                        user_defined_stopwords = []
                    elif e.response['Error']['Code'] == "NoSuchKey":
                        pass
                    else:
                        pass

                if isinstance(user_defined_stopwords, str):
                    if user_defined_stopwords.strip('\r').strip('\n') == '':
                        user_defined_stopwords = []

                file_sample_data = [' '.join([stop_and_stem_journal(doc.split('\t')[1], user_defined_stopwords)
                                                 , stop_and_stem_journal(doc.split('\t')[2], user_defined_stopwords)
                                              ]) for doc in file_sample_open]

                # Training Phase
                label_samples = [doc.split('\t')[-1].lower() for doc in file_sample_open]
                labels = sorted(list(set(label_samples)))

                train_data_existing_journals = file_sample_data
                train_target_existing_journals = label_samples

                if len(all_train_data) == 0:
                    all_train_data = train_data + train_data_existing_journals
                    all_train_target = train_target + train_target_existing_journals
                else:
                    all_train_data = all_train_data + train_data_existing_journals
                    all_train_target = all_train_target + train_target_existing_journals

            except botocore.exceptions.ClientError as e:
                if e.response['Error']['Code'] == "404":
                    # There is no existing journal training data, so do nothing
                    pass
                elif e.response['Error']['Code'] == "NoSuchKey":
                    pass
                else:
                    pass

                # end journal training data preprocessing

            all_train_tfidf = tfidf_vect.transform(all_train_data)

            if 'Support_Vector_Machine' in selectedModelName:

                # Need to update the calibration model that will be used to output probabilities in the excel sheet
                model_isotonic_calibration = None
                model_sigmoid_calibration = None

                # calibrate probabilities that will be used by the excel sheet
                if len(all_train_target) > 500:
                    model_isotonic_calibration = CalibratedClassifierCV(model, cv="prefit", method='isotonic')
                    model_isotonic_calibration.fit(all_train_tfidf.todense(), all_train_target)

                else:
                    model_sigmoid_calibration = CalibratedClassifierCV(model, cv="prefit", method='sigmoid')
                    model_sigmoid_calibration.fit(all_train_tfidf.todense(), all_train_target)

                # Save the calibration models for SVM
                if model_isotonic_calibration != None:
                    # Save the isotonic calibration model
                    isotonic_model = pickle.dumps(model_isotonic_calibration)
                    s3.put_object(Body=isotonic_model, Bucket=AWS_STORAGE_BUCKET_NAME,
                                  Key=key_value + 'model_isotonic_calibration.pkl')

                    # Remove the sigmoid calibration model, because this model was trained on less than 500 training data examples
                    try:
                        response2 = s3.get_object(Bucket=AWS_STORAGE_BUCKET_NAME,
                                                  Key=key_value + 'model_sigmoid_calibration.pkl')

                        response2 = s3.delete_object(Bucket=AWS_STORAGE_BUCKET_NAME,
                                                     Key=key_value + 'model_sigmoid_calibration.pkl')

                    except botocore.exceptions.ClientError as e:
                        if e.response['Error']['Code'] == "404":
                            pass
                        elif e.response['Error']['Code'] == "NoSuchKey":
                            pass
                        else:
                            pass

                if model_sigmoid_calibration != None:
                    # Save the sigmoid calibration model
                    isotonic_model = pickle.dumps(model_sigmoid_calibration)
                    s3.put_object(Body=isotonic_model, Bucket=AWS_STORAGE_BUCKET_NAME,
                                  Key=key_value + 'model_sigmoid_calibration.pkl')

            # Save the model, model's CV performance, and CV performance STDEV
            model_dumps = pickle.dumps(model)
            s3.put_object(Body=model_dumps, Bucket=AWS_STORAGE_BUCKET_NAME,
                          Key=key_value + 'model.pkl')

            # Increment the value of progressbar by 20 once the model is incrementally updated with additional data
            progressbar_value += 20
            es_conn.update(index=index_name_gloabals, id=id,
                           body={"doc": {"progressbar_value": progressbar_value}})

            # Update model's CV performance and save to file
            # Load the fiveFoldModel.pkl file with the five models

            # We cannot save the five fold training models for CV, because of multiprocessing implementation, which does not allow saving models to Multiprocessing Queues.
            # The logic has been changed to recompute five fold cross validation performances.
            # response = s3.get_object(Bucket=AWS_STORAGE_BUCKET_NAME,
            #                         Key=key_value + 'fiveFoldModels.pkl')
            # fiveFoldModelsFile = response['Body'].read()
            # fiveFoldModels = pickle.loads(fiveFoldModelsFile)

            # Also need to get the test data, which would be the original data

            # All the data is already saved into respective pickle files. Load the training and testing data for the five folds from the file, and add the training and test datasets
            # to respective datasets and re-evaluate the model's 5-fold cross validation performance.

            # response = s3.get_object(Bucket=AWS_STORAGE_BUCKET_NAME,
            #                          Key=key_value + 'fiveFoldTestingDatasetTfidf.pkl')
            # fiveFoldTestingDatasetTfidfFile = response['Body'].read()
            # fiveFoldTestingDatasetTfidf = pickle.loads(fiveFoldTestingDatasetTfidfFile)
            #
            # response = s3.get_object(Bucket=AWS_STORAGE_BUCKET_NAME,
            #                          Key=key_value + 'fiveFoldTestingLabels.pkl')
            # fiveFoldTestingLabelsFile = response['Body'].read()
            # fiveFoldTestingLabels = pickle.loads(fiveFoldTestingLabelsFile)
            #
            # response = s3.get_object(Bucket=AWS_STORAGE_BUCKET_NAME,
            #                          Key=key_value + 'fiveFoldTrainingDatasetTfidf.pkl')
            # fiveFoldTrainingDatasetTfidfFile = response['Body'].read()
            # fiveFoldTrainingDatasetTfidf = pickle.loads(fiveFoldTrainingDatasetTfidfFile)
            #
            # response = s3.get_object(Bucket=AWS_STORAGE_BUCKET_NAME,
            #                          Key=key_value + 'fiveFoldTrainingLabels.pkl')
            # fiveFoldTrainingLabelsFile = response['Body'].read()
            # fiveFoldTrainingLabels = pickle.loads(fiveFoldTrainingLabelsFile)

            performances_all_measures_all_folds = []
            standard_deviations_all_measures_all_folds = []

            # performCV = False
            # if len(train_data) > 4:
            #     num_splits = 5
            #     performCV = True
            # elif len(train_data) > 1:
            #     num_splits = len(train_data)
            #     performCV = True
            # else:
            #     # No need to do CV; simply add the one example to one of the folds randomly
            #     performCV = False
            #

            # Do the 5 fold cross validation
            # Divide the new training data into five folds -- need to save this information somewhere? Or maybe just use a seed for random number generator to repeat this step, whenever necessary in future

            skf = KFold(n_splits=5, random_state=7654, shuffle=True)

            fold_number = 0
            # Get optimal model parameter value from the model name
            optimal_model_parameter = selectedModelName.split('Alpha=')[1]
            optimal_model_parameter = optimal_model_parameter.split('_')[0]
            optimal_model_parameter = float(optimal_model_parameter)
            for train_indices, test_indices in skf.split(all_train_tfidf):

                # Note: It is okay to use tf-idf transformed data for doing CV, because tf-idf is unsupervised
                X_train, X_test = all_train_tfidf[train_indices], all_train_tfidf[test_indices]
                y_train, y_test = np.array(all_train_target)[train_indices], np.array(all_train_target)[test_indices]

                model_fold = None
                if 'Multinomial_Naive_Bayes' in selectedModelName and 'One_vs_Rest' not in selectedModelName:
                    # Get optimal alpha for the model
                    # The best model is already computed and best parameter is already determined
                    mnb_alpha = optimal_model_parameter

                    model_fold = MultinomialNB(alpha=mnb_alpha).partial_fit(X_train.todense(), y_train,
                                                                            classes=np.unique(y_train))

                elif 'Logistic_Regression' in selectedModelName and 'One_vs_Rest' not in selectedModelName:
                    # Get optimal alpha for the model
                    lrl2_alpha = optimal_model_parameter

                    random_state = np.random.RandomState(seed=87654)
                    # output of the model is dependent on the interaction between alpha and the number of epochs (n_iter)
                    # When alpha is very small, n_iter must be large to compensate for the slower learning rate
                    # When using SGD, the partial_fit method has to be applied on different batches of the training data, and we need to epoch multiple times

                    model_fold = SGDClassifier(loss='log', penalty='l2', alpha=lrl2_alpha, class_weight=None,
                                               random_state=random_state)

                    train_tfidf_dense = X_train.toarray()

                    def batches(l, n):
                        for i in np.arange(0, len(l), n):
                            yield l[i:i + n]

                    n_iter = 25
                    np.random.seed(5647)
                    shuffledRange = np.arange(len(train_tfidf_dense))
                    for n in np.arange(n_iter):
                        np.random.shuffle(shuffledRange)
                        shuffled_train_tfidf = [train_tfidf_dense[i] for i in shuffledRange]
                        shuffled_train_target = [y_train[i] for i in shuffledRange]

                        # Training the model in 10 batches
                        for batch in batches(np.arange(len(shuffled_train_target)), 5):
                            model_fold.partial_fit(shuffled_train_tfidf[batch[0]:batch[-1] + 1],
                                                   shuffled_train_target[batch[0]:batch[-1] + 1],
                                                   classes=np.unique(y_train))


                elif 'Support_Vector_Machine' in selectedModelName and 'One_vs_Rest' not in selectedModelName:
                    # Get optimal alpha for the model, performance of 5-fold CV, and standard deviation of performance
                    svm_alpha = optimal_model_parameter

                    random_state = np.random.RandomState(seed=87654)

                    # When using SGD, the partial_fit method has to be applied on different batches of the training data, and we need to epoch multiple times
                    model_fold = SGDClassifier(loss='hinge', penalty='l2', alpha=svm_alpha, class_weight=None,
                                               random_state=random_state)

                    train_tfidf_dense = X_train.toarray()

                    def batches(l, n):
                        for i in np.arange(0, len(l), n):
                            yield l[i:i + n]

                    n_iter = 25
                    np.random.seed(5647)
                    shuffledRange = np.arange(len(train_tfidf_dense))
                    for n in np.arange(n_iter):
                        np.random.shuffle(shuffledRange)
                        shuffled_train_tfidf = [train_tfidf_dense[i] for i in shuffledRange]
                        shuffled_train_target = [y_train[i] for i in shuffledRange]

                        # Training the model in 10 batches
                        for batch in batches(np.arange(len(shuffled_train_target)), 5):
                            model_fold.partial_fit(shuffled_train_tfidf[batch[0]:batch[-1] + 1],
                                                   shuffled_train_target[batch[0]:batch[-1] + 1],
                                                   classes=np.unique(y_train))

                if 'One_vs_Rest_Multionomial_Naive_Bayes' in selectedModelName:
                    # Get optimal alpha for the model
                    mnb_alpha = optimal_model_parameter

                    model_fold = OneVsRestClassifier(MultinomialNB(alpha=mnb_alpha)).partial_fit(X_train.todense(),
                                                                                                 y_train,
                                                                                                 classes=np.unique(
                                                                                                     y_train))


                elif 'One_vs_Rest_Logistic_Regression' in selectedModelName:
                    # Get optimal alpha for the model
                    lrl2_alpha = optimal_model_parameter

                    random_state = np.random.RandomState(seed=87654)

                    # When using SGD, the partial_fit method has to be applied on different batches of the training data, and we need to epoch multiple times
                    model_fold = OneVsRestClassifier(
                        SGDClassifier(loss='log', penalty='l2', alpha=lrl2_alpha, class_weight=None,
                                      random_state=random_state))

                    train_tfidf_dense = X_train.toarray()

                    def batches(l, n):
                        for i in np.arange(0, len(l), n):
                            yield l[i:i + n]

                    n_iter = 25
                    np.random.seed(5647)
                    shuffledRange = np.arange(len(train_tfidf_dense))
                    for n in np.arange(n_iter):
                        np.random.shuffle(shuffledRange)
                        shuffled_train_tfidf = [train_tfidf_dense[i] for i in shuffledRange]
                        shuffled_train_target = [y_train[i] for i in shuffledRange]

                        # Training the model in 10 batches
                        for batch in batches(np.arange(len(shuffled_train_target)), 5):
                            model_fold.partial_fit(shuffled_train_tfidf[batch[0]:batch[-1] + 1],
                                                   shuffled_train_target[batch[0]:batch[-1] + 1],
                                                   classes=np.unique(y_train))

                elif 'One_vs_Rest_Support_Vector_Machine' in selectedModelName:
                    # Get optimal alpha for the model, performance of 5-fold CV, and standard deviation of performance
                    svm_alpha = optimal_model_parameter

                    random_state = np.random.RandomState(seed=87654)

                    # When using SGD, the partial_fit method has to be applied on different batches of the training data, and we need to epoch multiple times
                    model_fold = OneVsRestClassifier(
                        SGDClassifier(loss='hinge', penalty='l2', alpha=svm_alpha, class_weight=None,
                                      random_state=random_state))

                    train_tfidf_dense = X_train.toarray()

                    def batches(l, n):
                        for i in np.arange(0, len(l), n):
                            yield l[i:i + n]

                    n_iter = 25
                    np.random.seed(5647)
                    shuffledRange = np.arange(len(train_tfidf_dense))
                    for n in np.arange(n_iter):
                        np.random.shuffle(shuffledRange)
                        shuffled_train_tfidf = [train_tfidf_dense[i] for i in shuffledRange]
                        shuffled_train_target = [y_train[i] for i in shuffledRange]

                        # Training the model in 10 batches
                        for batch in batches(np.arange(len(shuffled_train_target)), 5):
                            model_fold.partial_fit(shuffled_train_tfidf[batch[0]:batch[-1] + 1],
                                                   shuffled_train_target[batch[0]:batch[-1] + 1],
                                                   classes=np.unique(y_train))

                # evaluate the model on the test data and record performances
                (accu, auc, micro_precision, macro_precision, micro_recall, macro_recall, micro_f1, macro_f1,
                 pred_y) = evaluate_model_MS(model_fold, X_train, y_train, list(set(y_train)))

                performances_all_measures_one_fold = []
                performances_all_measures_one_fold.append(accu)
                performances_all_measures_one_fold.append(auc)
                performances_all_measures_one_fold.append(macro_precision)
                performances_all_measures_one_fold.append(macro_recall)
                performances_all_measures_one_fold.append(macro_f1)
                performances_all_measures_all_folds.append(performances_all_measures_one_fold)

                fold_number += 1

                progressbar_value += 20
                es_conn.update(index=index_name_gloabals, id=id,
                               body={"doc": {"progressbar_value": progressbar_value}})

            all_measures_performances = np.average(performances_all_measures_all_folds, axis=0)
            standard_deviations_all_measures_all_folds = np.std(performances_all_measures_all_folds, axis=0)

            progress_text = progress_text + "\nThe five-fold cross validation performances have been successfully updated."
            es_conn.update(index=index_name_gloabals, id=id,
                           body={"doc": {"progress_text": progress_text}})

            # Manually set to 160
            progressbar_value = 160
            es_conn.update(index=index_name_gloabals, id=id,
                           body={"doc": {"progressbar_value": progressbar_value}})
            #
            # else:
            #     # Not perform CV, and add one training example to any one fold randomly
            #     random_fold_number = random.randint(1, 4)
            #
            #     fold_number = 0
            #
            #     performances_all_measures_all_folds = []
            #     standard_deviations_all_measures_all_folds = []
            #
            #     for fold_number in np.arange(0, 5):
            #         if fold_number == random_fold_number:
            #
            #             X_train = train_tfidf[0]
            #             y_train = np.array(train_target)[0]
            #
            #             # Need to convert the data to tfidf
            #             old_X_train = tfidf_vect.transform(fiveFoldTrainingDatasetTfidf[random_fold_number])
            #             old_y_train = fiveFoldTrainingLabels[random_fold_number]
            #             old_X_test = tfidf_vect.transform(fiveFoldTestingDatasetTfidf[random_fold_number])
            #             old_y_test = fiveFoldTestingLabels[random_fold_number]
            #
            #             # Update each of the models in the five folds one by one in this for loop, and evaluate its performance on the test data
            #             model_fold = fiveFoldModels[random_fold_number]
            #
            #             if 'Support_Vector_Machine' in selectedModelName or 'Logistic_Regression' in selectedModelName:
            #
            #                 def batches(l, n):
            #                     for i in np.arange(0, len(l), n):
            #                         yield l[i:i + n]
            #
            #                 train_tfidf_dense = np.concatenate((old_X_train.toarray(), X_train.toarray()), axis=0)
            #                 train_target_concatenated = np.concatenate((old_y_train, y_train), axis=0)
            #
            #                 n_iter = 25
            #                 np.random.seed(5647)
            #                 shuffledRange = np.arange(len(train_tfidf_dense))
            #                 for n in np.arange(n_iter):
            #                     np.random.shuffle(shuffledRange)
            #                     shuffled_train_tfidf = [train_tfidf_dense[i] for i in shuffledRange]
            #                     shuffled_train_target = [train_target_concatenated[i] for i in shuffledRange]
            #
            #                     # Training the model in 10 batches
            #                     for batch in batches(np.arange(len(shuffled_train_target)), 5):
            #                         model_fold.partial_fit(shuffled_train_tfidf[batch[0]:batch[-1] + 1],
            #                                                shuffled_train_target[batch[0]:batch[-1] + 1],
            #                                                classes=np.unique(train_target_concatenated))
            #
            #             else:
            #                 model_fold.partial_fit(np.concatenate((old_X_train.todense(), X_train.todense()), axis=0),
            #                                        np.concatenate((old_y_train, y_train), axis=0))
            #
            #             # evaluate the model on the test data and record performances
            #             (accu, auc, micro_precision, macro_precision, micro_recall, macro_recall, micro_f1, macro_f1,
            #              pred_y) = evaluate_model_MS(model_fold, old_X_test.todense(), old_y_test,
            #                                          list(set(np.concatenate((old_y_train, y_train), axis=0))))
            #
            #         else:
            #             # simply evauate train and test without adding any more examples to train or test parts
            #             model_fold = fiveFoldModels[random_fold_number]
            #
            #             (accu, auc, micro_precision, macro_precision, micro_recall, macro_recall, micro_f1, macro_f1,
            #              pred_y) = evaluate_model_MS(model_fold, old_X_test.todense(), old_y_test,
            #                                          list(set(np.concatenate((old_y_train, y_train), axis=0))))
            #
            #         performances_all_measures_one_fold = []
            #         performances_all_measures_one_fold.append(accu)
            #         performances_all_measures_one_fold.append(auc)
            #         performances_all_measures_one_fold.append(macro_precision)
            #         performances_all_measures_one_fold.append(macro_recall)
            #         performances_all_measures_one_fold.append(macro_f1)
            #         performances_all_measures_all_folds.append(performances_all_measures_one_fold)
            #
            #         progressbar_value += 20
            #         es_conn.update(index=index_name_gloabals, id=id,
            #                        body={"doc": {"progressbar_value": progressbar_value}})

            all_measures_performances = np.average(performances_all_measures_all_folds, axis=0)
            all_measures_standardDeviations = np.std(performances_all_measures_all_folds, axis=0)

            performances_dumps = pickle.dumps(all_measures_performances)
            s3.put_object(Body=performances_dumps, Bucket=AWS_STORAGE_BUCKET_NAME,
                          Key=key_value + 'performances.pkl')

            performancesStdev_dumps = pickle.dumps(all_measures_standardDeviations)
            s3.put_object(Body=performancesStdev_dumps, Bucket=AWS_STORAGE_BUCKET_NAME,
                          Key=key_value + 'performancesStdev.pkl')

            progress_text = progress_text + "Model has been incrementally updated with additional data and is saved sucessfully...\n"

            progressbar_value += 40

            programRunEndTime = datetime.now()

            timeDifference = relativedelta(programRunEndTime, programRunStartTime)

            programRunStartTimeLabel = "Program run took %d days %d hours %d minutes %d seconds." % (
                timeDifference.days, timeDifference.hours, timeDifference.minutes, timeDifference.seconds)
            progressbarlabel_text = programRunStartTimeLabel
            es_conn.update(index=index_name_gloabals, id=id,
                           body={"doc": {"progressbar_value": progressbar_value, "progress_text": progress_text,
                                         "progressbarlabel_text": progressbarlabel_text}})
            modelSavingTimeLabel = "Incrementally training the model took %d days %d hours %d minutes %d seconds." % (
                timeDifference.days, timeDifference.hours, timeDifference.minutes, timeDifference.seconds)

            append_text_to_history_file += '5-fold Cross Validation Performance after model update: ' + '\n'
            perfMeasuresStr = ['Accuracy:', 'AUC:', 'Precision:', 'Recall:', 'F1:']
            for i in range(len(all_measures_performances)):
                stringToWrite = '{:<10s}{:>10.2f}{:>4s}{:>10.2f}{:>1s}'.format(perfMeasuresStr[i],
                                                                               all_measures_performances[i] * 100.0,
                                                                               '% +/- ',
                                                                               all_measures_standardDeviations[
                                                                                   i] * 100.0, '%')
                append_text_to_history_file += stringToWrite + '\n'
            historyFile = ''
            historyFile.write(modelSavingTimeLabel + '.' + '\n' + '*' * 95 + '\n')

            append_text_to_history_file += '5-fold Cross Validation Performance after model update: ' + '\n'

            try:
                response2 = s3.get_object(Bucket=AWS_STORAGE_BUCKET_NAME,
                                          Key=key_value + 'history.txt')
                history_file_old_text = response2['Body'].read().decode('utf-8')
                append_text_to_history_file = history_file_old_text + append_text_to_history_file

                s3.put_object(Body=append_text_to_history_file, Bucket=AWS_STORAGE_BUCKET_NAME,
                              Key=key_value + 'history.txt')

            except botocore.exceptions.ClientError as e:
                if e.response['Error']['Code'] == "404":
                    # There is no existing history file, so create a new history file and write the history into that file in S3.
                    s3.put_object(Body=append_text_to_history_file, Bucket=AWS_STORAGE_BUCKET_NAME,
                                  Key=key_value + 'history.txt')
                elif e.response['Error']['Code'] == "NoSuchKey":
                    pass
                else:
                    pass

            progressbar_value += 40
            final_progress_value = 200
            es_conn.update(index=index_name_gloabals, id=id,
                           body={"doc": {"progressbar_value": progressbar_value,
                                         "final_progress_value": final_progress_value}})
            return JsonResponse({'finalResponse':
                                     "Incremental learning finished successfully. The model has been incrementally updated with the new data."})
    except Exception as e:
        return HttpResponse(e)


def preprocess_collection_incremental_learning(username, file_open, stopterms, existing_application_numbers, printInfo,
                                               progress_text):
    # de-duplication
    try:
        master = []  # list for used application numbers

        if len(existing_application_numbers) > 0:
            master = existing_application_numbers

        repeat = []  # list for duplicate application numbers
        file_temp = []  # updated collection file
        counter = 0
        num_docs = len(file_open)
        for index, doc in enumerate(file_open, start=1):
            try:
                apn = doc.split("\t")
                apn = apn[4].upper()
                if apn not in master:
                    file_temp.append(doc)
                    master.append(apn)
                    counter = counter + 1
                elif apn in master:
                    repeat.append(apn)
            except Exception as e:
                final_progress_value = 200
                progress_text = progress_text + "*" * 50 + "\n" + "ERROR: The document number %d in the file could not be processed" % index + "\n" + "-" * 50
                errorString = "The document number %d in the file could not be processed" % index
                index_name_gloabals = 'apolloglobals'
                query = {"query": {"bool": {"must": {"match": {"username.keyword": username}}}}}
                res = es_conn.search(index=index_name_gloabals, body=query)
                id = res['hits']['hits'][0]['_id']
                es_conn.update(index=index_name_gloabals, id=id,
                               body={"doc": {"final_progress_value": final_progress_value,
                                             "errorString": errorString, "progress_text": progress_text}})
                continue

        # step through collection docs (ie lines)
        file_proc = []
        file_stem = []
        file_stop_words_removed = []

        design_count = 0  # counter for design cases
        utility_count = 0  # counter for utility cases

        for file_index, file_line in enumerate(file_temp, start=1):
            file_line = file_line.split("\t")  # split by tab
            # take correct col number docs only
            try:
                no = str(file_index)
                file_t = file_line[1]  # title
                file_a = file_line[2]  # abstract
                file_c = file_line[3]  # claims
                apn = file_line[4].lower()
                apd = file_line[5]
                asgn = file_line[6].lower()
                if len(file_line) > 7:
                    upc = file_line[7].lower()
                if len(file_line) > 8:
                    label = file_line[8].lower()  # solve the issue if label has tab
            except Exception as e:
                # apollo4.globals.final_progress_value = 200
                # apollo4.globals.errorString = "The document number %d in the file could not be processed4430" % file_index
                progress_text = progress_text + "*" * 50 + "\n" + "ERROR: The document number %d in the file could not be processed" % file_index + "\n" + "-" * 50
                pass

            if apn.startswith("us2"):
                # filter out design cases
                progress_text = progress_text + "*" * 50 + "\n" + "Design patent found! App_No: %r\tUPC: %r" % (
                    apn, upc) + '\n' + "-" * 50
                design_count = design_count + 1

            elif apn.startswith("us"):
                # filter out non-apn lines (ie not patent data)
                utility_count = utility_count + 1

                # stop and stem title, abstract, claim
                file_t_stem = stop_and_stem(file_t, stopterms)
                file_a_stem = stop_and_stem(file_a, stopterms)
                file_c_stem = stop_and_stem(file_c, stopterms)

                # remove stopwords from the title, abstract, claim
                file_t_stop = remove_stopwords(file_t, stopterms)
                file_a_stop = remove_stopwords(file_a, stopterms)
                file_c_stop = remove_stopwords(file_c, stopterms)

                # Output the orginal clean version of utility patent
                file_new_line = '\t'.join(file_line)
                file_proc.append(file_new_line)

                # Output the preprocessed version of utility patent
                if len(file_line) > 7:
                    proc_doc = [no, file_t_stem, file_a_stem, file_c_stem,
                                apd, apn, asgn, upc]
                    proc_doc_stop = [no, file_t_stop, file_a_stop, file_c_stop,
                                     apd, apn, asgn, upc]

                else:
                    proc_doc = [no, file_t_stem, file_a_stem, file_c_stem,
                                apd, apn, asgn]
                    proc_doc_stop = [no, file_t_stop, file_a_stop, file_c_stop,
                                     apd, apn, asgn]

                if len(file_line) > 8:  # solve the issue if label has tab
                    proc_doc.append(label)
                    proc_doc_stop.append(label)

                proc_doc = '\t'.join(proc_doc)
                proc_doc_stop = '\t'.join(proc_doc_stop)

                file_stem.append(proc_doc)
                file_stop_words_removed.append(proc_doc_stop)

        # if printInfo:
        #     progress_text = progress_text + "stopwords removed, terms stemmed, documents de-duplicated, design removed\n" + \
        #                                     "%d unique documents out of %d total" % (counter, num_docs) + '\n' + \
        #                                     "%d design documents out of %d total" % (design_count, num_docs) + '\n' + \
        #                                     "%d utility documents out of %d total" % (utility_count, num_docs) + '\n'

        output = (file_proc, file_stem, file_stop_words_removed)
        return output
    except Exception as e:
        return HttpResponse(
            "Error running the program." + str(e))


def dedup_collection_journal_incremental_learning(file_open, existing_title_and_abstracts, uid, abstract_id):
    new_file_list = []
    new_item_list = []
    try:
        if len(existing_title_and_abstracts) > 0:
            new_item_list = existing_title_and_abstracts

        for doc in file_open:
            item = '\t'.join(doc.split('\t')[uid: abstract_id + 1])
            if item not in new_item_list:
                new_item_list.append(item)
                new_file_list.append(doc)
        return new_file_list
    except Exception as e:
        return HttpResponse(
            "Error running the program." + str(e))


def evaluate_model_MS(model, X_test, y_test, y_label_set):
    # initialize all peformance measures to -1 (i.e., not defined)
    accu = -1
    auc = -1
    micro_precision = -1
    macro_precision = -1
    micro_recall = -1
    macro_recall = -1
    micro_f1 = -1
    macro_f1 = -1
    try:
        # Note: The y_test needs to be a binary array of [n_samples, n_classes] where each value indicates presence/absence of class label in the respective column
        labelBinarizer = preprocessing.LabelBinarizer()
        labelBinarizer.fit(y_label_set)
        binarizedLabels = labelBinarizer.transform(y_test)

        if isinstance(model, MultinomialNB) or isinstance(model, KNeighborsClassifier) or isinstance(model,
                                                                                                     DeepLearningModel):
            # KNN Classifier's predict_proba function does not work with sparse matrices
            if isinstance(model, KNeighborsClassifier):
                y_probas = model.predict_proba(X_test)
            else:
                y_probas = model.predict_proba(X_test)
            try:
                auc = metrics.roc_auc_score(binarizedLabels, y_probas)
            except:
                # If AUC cannot be computed, set AUC value to -1, to represent not defined
                auc = -1
        elif isinstance(model, OneVsRestClassifier) and isinstance(model.estimator, MultinomialNB):
            y_probas = model.predict_proba(X_test)
            try:
                auc = metrics.roc_auc_score(binarizedLabels, y_probas)
            except:
                # If AUC cannot be computed, set AUC value to -1, to represent not defined
                auc = -1
        else:
            # For LR and SVM:
            y_decision = model.decision_function(X_test)
            try:
                auc = metrics.roc_auc_score(binarizedLabels, y_decision)
            except:
                # If AUC cannot be computed, set AUC value to -1, to represent not defined
                auc = -1

        pred_y = model.predict(X_test)
        accu = metrics.accuracy_score(y_test, pred_y)
        micro_precision = metrics.precision_score(y_test, pred_y, average='micro')
        macro_precision = metrics.precision_score(y_test, pred_y, average='macro')
        micro_recall = metrics.recall_score(y_test, pred_y, average='micro')
        macro_recall = metrics.recall_score(y_test, pred_y, average='macro')
        micro_f1 = metrics.f1_score(y_test, pred_y, average='micro')
        macro_f1 = metrics.f1_score(y_test, pred_y, average='macro')

        return (accu, auc, micro_precision, macro_precision, micro_recall, macro_recall, micro_f1, macro_f1, pred_y)
    except Exception as e:
        return HttpResponse(
            "Error running the program." + str(e))


@csrf_exempt
def run_IL_trainFromScratchFromGUI(request):
    # ...
    try:
        # Assuming that the client-side has already selected the options before running the program.
        # Assuming that the request from the client side will have all the fields necessary for running the program.
        if request.method == "POST":
            gui_parameters_data = request.body.decode('utf-8');
            username = request.user.username;
            gui_parameters = json.loads(gui_parameters_data);
            selectedProjectName = gui_parameters['saveProjectName']
            selectedModelName = gui_parameters['trainedModelName']
            programRunStartTime = datetime.now()
            modelValidation = retrainModelFromScratch(None, None, selectedProjectName, selectedModelName, None,
                                                      programRunStartTime, True, username)
            final_progress_value = 200
            index_name_gloabals = 'apolloglobals'
            query = {"query": {"bool": {"must": {"match": {"username.keyword": username}}}}}
            res = es_conn.search(index=index_name_gloabals, body=query)
            id = res['hits']['hits'][0]['_id']
            es_conn.update(index=index_name_gloabals, id=id,
                           body={"doc": {"final_progress_value": final_progress_value}})
            if modelValidation == True:
                return JsonResponse({'finalResponse': 'incremental learning sucessfully executed'})
            else:
                errorString = modelValidation.content
                es_conn.update(index=index_name_gloabals, id=id,
                               body={"doc": {"errorString": errorString}})
                return JsonResponse({'finalResponse': 'Error'})
    except Exception as e:
        final_progress_value = 200
        errorString = e
        es_conn.update(index=index_name_gloabals, id=id,
                       body={"doc": {"errorString": errorString, "final_progress_value": final_progress_value}})
        return HttpResponse(e)


def retrainModelFromScratch(train_data, train_target, selectedProjectName, selectedModelName, targetPerformanceMeasure,
                            programRunStartTime, runFromGUI, username):
    try:
        index_name_gloabals = 'apolloglobals'
        query = {"query": {"bool": {"must": {"match": {"username.keyword": username}}}}}
        res = es_conn.search(index=index_name_gloabals, body=query)
        id = res['hits']['hits'][0]['_id']

        trainingDataType = res['hits']['hits'][0]['_source']['trainingDataType']
        trainingFileName = res['hits']['hits'][0]['_source']['trainingFileName']
        progress_text = res['hits']['hits'][0]['_source']['progress_text']

        if runFromGUI:
            train_data_IL = res['hits']['hits'][0]['_source']['train_data']
            train_target_IL = res['hits']['hits'][0]['_source']['train_target']
            progressbar_value_IL = res['hits']['hits'][0]['_source']['progress_value']
            targetPerformanceMeasure_IL = res['hits']['hits'][0]['_source']['targetPerformanceMeasure_IL']

            train_data = train_data_IL
            train_target = train_target_IL
            progressbar_value = progressbar_value_IL
            targetPerformanceMeasure = targetPerformanceMeasure_IL

        # This function is only applicable to SUPERVISED learning mode, which may contain new classes. This does not apply to UNSUPERVISED learning case.
        # Retrain the selected model by combining the old patent data + old journal data + new data
        # output_folder = os.path.dirname(selectedTrainingFilePath)
        progressbar_value = 0
        # progress_text = ''
        progressbar_maximum = 80 + progressbar_value
        progress_text = progress_text + "Training the model from scratch using the additional data..."

        es_conn.update(index=index_name_gloabals, id=id,
                       body={"doc": {"progressbar_maximum": progressbar_maximum, "progress_text": progress_text}})

        # If old training data exists, load the data and preprocess the data
        # Once loaded, append the data to train_data and train_target to train the model from scratch

        s3 = boto3.client('s3')
        key_value = 'classificationprojects/'
        key_value += selectedProjectName + '/supervised/'
        key_value += selectedModelName + '/'

        try:
            response = s3.get_object(Bucket=AWS_STORAGE_BUCKET_NAME,
                                     Key=key_value + 'training_data_patents.txt')
            oldPatentsFile = response['Body'].read()

            file_sample_open = oldPatentsFile.decode('utf-8')
            file_sample_open = file_sample_open.split('\n')  # split by new line
            file_sample_open = list(filter(None, file_sample_open))  # delete empty lines
            # Now, the first line is header, so remove the first line
            file_sample_open = file_sample_open[1:]
            existing_application_numbers = [doc.split('\t')[4].upper() for doc in file_sample_open]

            progress_text = progress_text + "\nFound existing " + str(
                len(file_sample_open)) + " patents! \nPreprocessing documents...\n"

            # Set value of progressbar to 5 once the training dataset is loaded
            progressbar_value += 5
            es_conn.update(index=index_name_gloabals, id=id,
                           body={"doc": {"progressbar_value": progressbar_value, "progress_text": progress_text}})
            # Build the stop words
            stops = stopwords

            aux_stops = './static/AuxStops.txt'

            aux_stops = open(aux_stops, 'r').read()
            aux_stops = re.sub("[^a-zA-Z ]", "   ", aux_stops)  # remove non-alphanumeric
            aux_stops = " ".join(aux_stops.split())  # split by any whitespace and rejoin w/ space
            aux_stops = aux_stops.split(' ')
            aux_stops = list(filter(None, aux_stops))

            # append auxiliary stops
            stops = stops + aux_stops

            # append user-provided stop words
            try:
                response = s3.get_object(Bucket=AWS_STORAGE_BUCKET_NAME,
                                         Key=key_value + 'user_defined_stopwords.pkl')
                user_defined_stopwords = pickle.loads(response['Body'].read())
                # user_defined_stopwords = pickle.load(user_defined_stopwords_file)
            except botocore.exceptions.ClientError as e:
                if e.response['Error']['Code'] == "404":
                    # There is no existing patent (or journal) training data, so do nothing
                    user_defined_stopwords = []
                elif e.response['Error']['Code'] == "NoSuchKey":
                    pass
                else:
                    pass

            if len(user_defined_stopwords) > 0:
                stops = stops + user_defined_stopwords
            else:
                stops = stops

            # Bulid stopterm dictionary
            stopterms = {}
            for stop in stops:
                if stop in stopterms:
                    stopterms[stop] += 1
                else:
                    stopterms[stop] = 1

            # Preprocess the sample file
            (file_sample_proc, file_sample_stem, temp) = preprocess_collection(file_sample_open, stopterms, True,
                                                                               progress_text)
            file_sample = list(filter(None, file_sample_stem))

            title_samples = [doc.split('\t')[1].strip('\r').strip('\n') for doc in file_sample]
            abstract_samples = [doc.split('\t')[2].strip('\r').strip('\n') for doc in file_sample]
            claim_samples = [doc.split('\t')[3].strip('\r').strip('\n') for doc in file_sample]
            label_samples = [doc.split('\t')[8].lower().strip('\r').strip('\n') for doc in file_sample]
            labels = sorted(list(set(label_samples)))
            train_data_exisiting_patents = [' '.join(doc) for doc in
                                            zip(title_samples, abstract_samples, claim_samples)]
            train_target_existing_patents = label_samples

            # Append the existing data to new data
            train_data = train_data + train_data_exisiting_patents
            train_target = train_target + train_target_existing_patents

        except botocore.exceptions.ClientError as e:
            if e.response['Error']['Code'] == "404":
                # There is no existing patent training data, so no need to deduplicate based on previous data.
                existing_application_numbers = []
            elif e.response['Error']['Code'] == "NoSuchKey":
                existing_application_numbers = []
                pass
            else:
                existing_application_numbers = []
                pass

            # End patent training data

        try:
            response = s3.get_object(Bucket=AWS_STORAGE_BUCKET_NAME,
                                     Key=key_value + 'training_data_journals.txt')
            oldJournalsFile = response['Body'].read()

            file_sample_open = oldJournalsFile.decode('utf-8')
            file_sample_open = file_sample_open.split('\n')
            file_sample_open = list(filter(None, file_sample_open))  # delete empty lines

            # Now, the first line is header, so remove the first line
            file_sample_open = file_sample_open[1:]
            progress_text = progress_text + "\nFound " + str(
                len(file_sample_open)) + " journals! \nPreprocessing documents...\n"

            # Set value of progressbar to 5 once the training dataset is loaded
            progressbar_value += 5
            es_conn.update(index=index_name_gloabals, id=id,
                           body={"doc": {"progressbar_value": progressbar_value, "progress_text": progress_text}})
            # Remove the duplicated documents based on "title"
            file_sample_open = dedup_collection_journal(file_sample_open, 1, 2)

            # Preprocessing for scoupus data
            file_sample_open = preprocess_collection_journal(file_sample_open)

            # Take the stopwords from the GUI and add them to the stopwords list
            # append user-provided stop words
            try:
                response = s3.get_object(Bucket=AWS_STORAGE_BUCKET_NAME,
                                         Key=key_value + 'user_defined_stopwords.pkl')
                user_defined_stopwords_file = response['Body'].read()
                user_defined_stopwords = pickle.loads(user_defined_stopwords_file)
            except botocore.exceptions.ClientError as e:
                if e.response['Error']['Code'] == "404":
                    # There is no existing patent (or journal) training data, so do nothing
                    user_defined_stopwords = []
                elif e.response['Error']['Code'] == "NoSuchKey":
                    pass
                else:
                    pass

            if isinstance(user_defined_stopwords, str):
                if user_defined_stopwords.strip('\r').strip('\n') == '':
                    user_defined_stopwords = []

            file_sample_data = [' '.join([stop_and_stem_journal(doc.split('\t')[1], user_defined_stopwords)
                                             , stop_and_stem_journal(doc.split('\t')[2], user_defined_stopwords)
                                          ]) for doc in file_sample_open]

            # Training Phase
            label_samples = [doc.split('\t')[-1].lower().strip('\r').strip('\n') for doc in file_sample_open]
            labels = sorted(list(set(label_samples)))

            train_data_existing_journals = file_sample_data
            train_target_existing_journals = label_samples

            train_data = train_data + train_data_existing_journals
            train_target = train_target + train_target_existing_journals

        except botocore.exceptions.ClientError as e:
            if e.response['Error']['Code'] == "404":
                # There is no existing journal training data, so no need to deduplicate based on previous data.
                pass
            elif e.response['Error']['Code'] == "NoSuchKey":
                pass
            else:
                pass

            # end journal training data preprocessing

        # Display information once all data is pre-processed
        progress_text = progress_text + "Finished preprocessing the existing data" + "\nThe training data contains a total of " + str(
            len(train_data)) + " unique documents."

        # Set value of progressbar to 10 once the training dataset is preprocessed
        progressbar_value += 5
        es_conn.update(index=index_name_gloabals, id=id,
                       body={"doc": {"progressbar_value": progressbar_value, "progress_text": progress_text}})
        numInstancesTrainingData = len(train_data)
        trainingDataStats = []

        leastDocumentsForAClass = 5  # initialize to 5
        for label in set(train_target):
            distribution = str(np.round(train_target.count(label) * 100.0 / len(train_target) * 1.0, 2)) + '%'
            trainingDataStats.append([label, train_target.count(label), distribution])
            if train_target.count(label) < leastDocumentsForAClass:
                leastDocumentsForAClass = train_target.count(label)

        # Make sure that there are at least 5 documents for each class: this is required to perform 5-fold cross validation
        if leastDocumentsForAClass < 5:
            progress_text = progress_text + "*" * 50 + "\nThe program requires at least 5 training examples for each class. Please provide at least 5 training examples for each class and re-run the program." + "*" * 50
            progressbar_value = 0
            # Return this message and exit. No need to run anything further in this function.
            final_progress_value = 200
            errorString = "The program requires at least 5 training examples for each class. Please provide at least 5 training examples for each class and re-run the program."
            es_conn.update(index=index_name_gloabals, id=id,
                           body={"doc": {"progressbar_value": progressbar_value, "progress_text": progress_text,
                                         "final_progress_value": final_progress_value,
                                         "errorString": errorString}})
            return JsonResponse({'finalResponse': 'Error'})

        # Else, just continue with incremental learning
        progress_text = progress_text + "\nStarting model training..."

        cv = CountVectorizer()
        tfidf = TfidfTransformer()

        # Changed the n-grams to (1,5) in the line below, and max_df from 0.5 to 0.8, based on side-experiments
        tfidf_vect = TfidfVectorizer(analyzer='word',
                                     ngram_range=(1, 5),
                                     min_df=2,
                                     max_df=0.8,
                                     max_features=200000,
                                     stop_words='english',
                                     use_idf=True)

        # tf-idf with params
        train_tfidf = tfidf_vect.fit_transform(train_data)

        # Set value of progressbar to 15 once the training dataset is vectorized
        progress_text = progress_text + "\nOptimizing model parameters..."

        progressbar_value += 5
        es_conn.update(index=index_name_gloabals, id=id,
                       body={"doc": {"progressbar_value": progressbar_value, "progress_text": progress_text}})

        tfidfVectorizer = tfidf_vect

        # Model and model parameters
        model = None
        svm_alpha = 1.0  # default value
        mnb_alpha = 0.001  # default value
        lrl2_alpha = 1.0  # default value
        svm_kernel = 'linear'  # default value
        class_weight = None  # default value

        automatic_mode = False
        optimal_model_parameter = -1

        str_model_name = 'None'
        str_parameter_name = 'None'

        if 'One_vs_Rest_Multinomial_Naive_Bayes' in selectedModelName:
            str_model_name = 'ovrmnb'

        elif 'One_vs_Rest_Logistic_Regression' in selectedModelName:
            str_model_name = 'ovrlrl2'

        elif 'One_vs_Rest_Support_Vector_Machine' in selectedModelName:
            str_model_name = 'ovrsvm'

        elif 'Multinomial_Naive_Bayes' in selectedModelName:
            str_model_name = 'mnb'

        elif 'Logistic_Regression' in selectedModelName:
            str_model_name = 'lrl2'

        elif 'Support_Vector_Machine' in selectedModelName:
            str_model_name = 'svm'

        elif 'Deep_Learning_BERT' in selectedModelName:
            str_model_name = 'bert'

        elif 'Deep_Learning_RoBERTa' in selectedModelName:
            str_model_name = 'roberta'

        elif 'Deep_Learning_XLNet' in selectedModelName:
            str_model_name = 'xlnet'

        str_parameter_name = 'Alpha = '

        if str_model_name == 'mnb':
            # Get optimal alpha for the model
            mnb_alpha = -1
            if automatic_mode == False:
                mnb_alpha, max_performance_measure, standard_deviation, all_measures_performances, all_measures_standardDeviations = getOptimalParameterForMNB_alpha(
                    train_tfidf.todense(), train_data, train_target, targetPerformanceMeasure)
                trainingDataPerformances = all_measures_performances
                trainingDataPerformancesStandardDeviation = all_measures_standardDeviations
                fiveFoldTrainingDatasetTfidf = None
                fiveFoldTestingDatasetTfidf = None
                fiveFoldTrainingLabels = None
                fiveFoldTestingLabels = None
                fiveFoldModels = None

            else:
                # The best model is already computed and best parameter is already determined
                mnb_alpha = optimal_model_parameter

            model = MultinomialNB(alpha=mnb_alpha).partial_fit(train_tfidf.todense(), train_target,
                                                               classes=np.unique(train_target))
            trainedModel = model
            trainedModelName = 'Multinomial_Naive_Bayes_Alpha=' + str(mnb_alpha)

        elif str_model_name == 'lrl2':
            # Get optimal alpha for the model
            lrl2_alpha = -1
            if automatic_mode == False:
                lrl2_alpha, max_performance_measure, standard_deviation, all_measures_performances, all_measures_standardDeviations = getOptimalParameterForLR_alpha(
                    train_tfidf.todense(), train_data, train_target, targetPerformanceMeasure)
                trainingDataPerformances = all_measures_performances
                trainingDataPerformancesStandardDeviation = all_measures_standardDeviations
                fiveFoldTrainingDatasetTfidf = None
                fiveFoldTestingDatasetTfidf = None
                fiveFoldTrainingLabels = None
                fiveFoldTestingLabels = None
                fiveFoldModels = None

            else:
                lrl2_alpha = optimal_model_parameter

            random_state = np.random.RandomState(seed=87654)
            # output of the model is dependent on the interaction between alpha and the number of epochs (n_iter)
            # When alpha is very small, n_iter must be large to compensate for the slower learning rate
            # When using SGD, the partial_fit method has to be applied on different batches of the training data, and we need to epoch multiple times

            model = SGDClassifier(loss='log', penalty='l2', alpha=lrl2_alpha, class_weight=class_weight,
                                  random_state=random_state)

            train_tfidf_dense = train_tfidf.toarray()

            def batches(l, n):
                for i in np.arange(0, len(l), n):
                    yield l[i:i + n]

            n_iter = 25
            np.random.seed(5647)
            shuffledRange = np.arange(len(train_data))
            for n in np.arange(n_iter):
                np.random.shuffle(shuffledRange)
                shuffled_train_tfidf = [train_tfidf_dense[i] for i in shuffledRange]
                shuffled_train_target = [train_target[i] for i in shuffledRange]

                # Training the model in 10 batches
                for batch in batches(np.arange(len(shuffled_train_target)), 5):
                    model.partial_fit(shuffled_train_tfidf[batch[0]:batch[-1] + 1],
                                      shuffled_train_target[batch[0]:batch[-1] + 1], classes=np.unique(train_target))

            trainedModel = model
            trainedModelName = 'Logistic_Regression_Alpha=' + str(lrl2_alpha)

        elif str_model_name == 'svm':
            # Get optimal alpha for the model, performance of 5-fold CV, and standard deviation of performance
            if automatic_mode == False:
                svm_alpha, max_performance_measure, standard_deviation, all_measures_performances, all_measures_standardDeviations = getOptimalParameterForSVM_alpha(
                    train_tfidf.toarray(), train_data, train_target, targetPerformanceMeasure)
                trainingDataPerformances = all_measures_performances
                trainingDataPerformancesStandardDeviation = all_measures_standardDeviations
                fiveFoldTrainingDatasetTfidf = None
                fiveFoldTestingDatasetTfidf = None
                fiveFoldTrainingLabels = None
                fiveFoldTestingLabels = None
                fiveFoldModels = None

            else:
                svm_alpha = optimal_model_parameter

            random_state = np.random.RandomState(seed=87654)

            # When using SGD, the partial_fit method has to be applied on different batches of the training data, and we need to epoch multiple times
            model = SGDClassifier(loss='hinge', penalty='l2', alpha=svm_alpha, class_weight=class_weight,
                                  random_state=random_state)

            train_tfidf_dense = train_tfidf.toarray()

            def batches(l, n):
                for i in np.arange(0, len(l), n):
                    yield l[i:i + n]

            n_iter = 25
            np.random.seed(5647)
            shuffledRange = np.arange(len(train_data))
            for n in np.arange(n_iter):
                np.random.shuffle(shuffledRange)
                shuffled_train_tfidf = [train_tfidf_dense[i] for i in shuffledRange]
                shuffled_train_target = [train_target[i] for i in shuffledRange]

                # Training the model in 10 batches
                for batch in batches(np.arange(len(shuffled_train_target)), 5):
                    model.partial_fit(shuffled_train_tfidf[batch[0]:batch[-1] + 1],
                                      shuffled_train_target[batch[0]:batch[-1] + 1], classes=np.unique(train_target))

            trainedModel = model
            trainedModelName = 'Support_Vector_Machine_Alpha=' + str(svm_alpha)

            # Need to update the clibration model that will be used to output probabilities in the excel sheet
            model_isotonic_calibration = None
            model_sigmoid_calibration = None

            # calibrate probabilities that will be used by the excel sheet
            if len(train_target) > 500:
                model_isotonic_calibration = CalibratedClassifierCV(model, cv="prefit", method='isotonic')
                model_isotonic_calibration.fit(train_tfidf.todense(), train_target)

            else:
                model_sigmoid_calibration = CalibratedClassifierCV(model, cv="prefit", method='sigmoid')
                model_sigmoid_calibration.fit(train_tfidf.todense(), train_target)

        if str_model_name == 'ovrmnb':
            # Get optimal alpha for the model
            mnb_alpha = -1
            if automatic_mode == False:
                mnb_alpha, max_performance_measure, standard_deviation, all_measures_performances, all_measures_standardDeviations = getOptimalParameterForOVRMNB_alpha(
                    train_tfidf.todense(), train_data, train_target, targetPerformanceMeasure)
                trainingDataPerformances = all_measures_performances
                trainingDataPerformancesStandardDeviation = all_measures_standardDeviations
                fiveFoldTrainingDatasetTfidf = None
                fiveFoldTestingDatasetTfidf = None
                fiveFoldTrainingLabels = None
                fiveFoldTestingLabels = None
                fiveFoldModels = None

            else:
                # The best model is already computed and best parameter is already determined
                mnb_alpha = optimal_model_parameter

            model = OneVsRestClassifier(MultinomialNB(alpha=mnb_alpha)).partial_fit(train_tfidf.todense(), train_target,
                                                                                    classes=np.unique(train_target))
            trainedModel = model
            trainedModelName = 'One_vs_Rest_Multinomial_Naive_Bayes_Alpha=' + str(mnb_alpha)

        elif str_model_name == 'ovrlrl2':
            # Get optimal alpha for the model
            lrl2_alpha = -1
            if automatic_mode == False:
                lrl2_alpha, max_performance_measure, standard_deviation, all_measures_performances, all_measures_standardDeviations = getOptimalParameterForOVRLR_alpha(
                    train_tfidf.todense(), train_data, train_target, targetPerformanceMeasure)
                trainingDataPerformances = all_measures_performances
                fiveFoldTrainingDatasetTfidf = None
                fiveFoldTestingDatasetTfidf = None
                fiveFoldTrainingLabels = None
                fiveFoldTestingLabels = None
                fiveFoldModels = None

            else:
                lrl2_alpha = optimal_model_parameter

            random_state = np.random.RandomState(seed=87654)

            # When using SGD, the partial_fit method has to be applied on different batches of the training data, and we need to epoch multiple times
            model = OneVsRestClassifier(
                SGDClassifier(loss='log', penalty='l2', alpha=lrl2_alpha, class_weight=class_weight,
                              random_state=random_state))

            train_tfidf_dense = train_tfidf.toarray()

            def batches(l, n):
                for i in np.arange(0, len(l), n):
                    yield l[i:i + n]

            n_iter = 25
            np.random.seed(5647)
            shuffledRange = np.arange(len(train_data))
            for n in np.arange(n_iter):
                np.random.shuffle(shuffledRange)
                shuffled_train_tfidf = [train_tfidf_dense[i] for i in shuffledRange]
                shuffled_train_target = [train_target[i] for i in shuffledRange]

                # Training the model in 10 batches
                for batch in batches(np.arange(len(shuffled_train_target)), 5):
                    model.partial_fit(shuffled_train_tfidf[batch[0]:batch[-1] + 1],
                                      shuffled_train_target[batch[0]:batch[-1] + 1], classes=np.unique(train_target))

            trainedModel = model
            trainedModelName = 'One_vs_Rest_Logistic_Regression_Alpha=' + str(lrl2_alpha)

        elif str_model_name == 'ovrsvm':
            # Get optimal alpha for the model, performance of 5-fold CV, and standard deviation of performance

            if automatic_mode == False:
                svm_alpha, max_performance_measure, standard_deviation, all_measures_performances, all_measures_standardDeviations = getOptimalParameterForOVRSVM_alpha(
                    train_tfidf.todense(), train_data, train_target, targetPerformanceMeasure)
                trainingDataPerformances = all_measures_performances
                trainingDataPerformancesStandardDeviation = all_measures_standardDeviations
                fiveFoldTrainingDatasetTfidf = None
                fiveFoldTestingDatasetTfidf = None
                fiveFoldTrainingLabels = None
                fiveFoldTestingLabels = None
                fiveFoldModels = None

            else:
                svm_alpha = optimal_model_parameter

            random_state = np.random.RandomState(seed=87654)

            # When using SGD, the partial_fit method has to be applied on different batches of the training data, and we need to epoch multiple times
            model = OneVsRestClassifier(
                SGDClassifier(loss='hinge', penalty='l2', alpha=svm_alpha, class_weight=class_weight,
                              random_state=random_state))

            train_tfidf_dense = train_tfidf.toarray()

            def batches(l, n):
                for i in np.arange(0, len(l), n):
                    yield l[i:i + n]

            n_iter = 25
            np.random.seed(5647)
            shuffledRange = np.arange(len(train_data))
            for n in np.arange(n_iter):
                np.random.shuffle(shuffledRange)
                shuffled_train_tfidf = [train_tfidf_dense[i] for i in shuffledRange]
                shuffled_train_target = [train_target[i] for i in shuffledRange]

                # Training the model in 10 batches
                for batch in batches(np.arange(len(shuffled_train_target)), 5):
                    model.partial_fit(shuffled_train_tfidf[batch[0]:batch[-1] + 1],
                                      shuffled_train_target[batch[0]:batch[-1] + 1], classes=np.unique(train_target))

            # Need to update the clibration model that will be used to output probabilities in the excel sheet
            model_isotonic_calibration = None
            model_sigmoid_calibration = None

            # calibrate probabilities that will be used by the excel sheet
            if len(train_target) > 500:
                model_isotonic_calibration = CalibratedClassifierCV(model, cv="prefit", method='isotonic')
                model_isotonic_calibration.fit(train_tfidf.todense(), train_target)

            else:
                model_sigmoid_calibration = CalibratedClassifierCV(model, cv="prefit", method='sigmoid')
                model_sigmoid_calibration.fit(train_tfidf.todense(), train_target)

            trainedModel = model
            trainedModelName = 'One_vs_Rest_Support_Vector_Machine_Alpha=' + str(svm_alpha)

        elif str_model_name == 'bert' or str_model_name == 'roberta' or str_model_name == 'xlnet':
            # Get optimal alpha for the model, performance of 5-fold CV, and standard deviation of performance
            if str_model_name == 'bert':
                dl_model_type = 'bert'
                dl_model_name = 'bert-base-cased'
            elif str_model_name == 'roberta':
                dl_model_type = 'roberta'
                dl_model_name = 'roberta-base'
            elif str_model_name == 'xlnet':
                dl_model_type = 'xlnet'
                dl_model_name = 'xlnet-base-cased'

            if automatic_mode == False:
                torch.cuda.empty_cache()
                # Need to get userName from the GUI to use here
                DEEP_LEARNING_OUTPUT_DIR = './DeepLearningOutputs/' + username + '/'
                optimal_parameters, max_performance_measure, standard_deviation, all_measures_performances, all_measures_standardDeviations = getOptimalParametersForDeepLearning(
                    dl_model_type, dl_model_name, train_data, targetPerformanceMeasure, DEEP_LEARNING_OUTPUT_DIR)
                trainingDataPerformances = all_measures_performances
                trainingDataPerformancesStandardDeviation = all_measures_standardDeviations
                fiveFoldTrainingDatasetTfidf = None
                fiveFoldTestingDatasetTfidf = None
                fiveFoldTrainingLabels = None
                fiveFoldTestingLabels = None
                fiveFoldModels = None
            else:
                optimal_parameters = optimal_model_parameter

            random_state = np.random.RandomState(seed=87654)
            batchSize = optimal_parameters[0]
            maxSequenceLength = optimal_parameters[1]

            total_labels = len(Counter(train_target))

            # Set the output directory where temporary model results will be stored
            torch.cuda.empty_cache()
            # Need to get userName from the GUI to use here
            DEEP_LEARNING_OUTPUT_DIR = './DeepLearningOutputs/' + username + '/'
            trainedModel = DeepLearningModel(dl_model_type, dl_model_name, batchSize, maxSequenceLength, num_epochs=30,
                                             random_state=4987, output_dir=DEEP_LEARNING_OUTPUT_DIR)

            # Train the model
            trainedModel.fit(train_data)

            if str_model_name == 'bert':
                trainedModelName = 'Deep Learning_BERT_BatchSize=' + str(
                    optimal_parameters[0]) + '_MaxSequenceLength=' + str(optimal_parameters[1])
            elif str_model_name == 'roberta':
                trainedModelName = 'Deep Learning_RoBERTa_BatchSize=' + str(
                    optimal_parameters[0]) + '_MaxSequenceLength=' + str(optimal_parameters[1])
            elif str_model_name == 'xlnet':
                trainedModelName = 'Deep Learning_XLNet_BatchSize=' + str(
                    optimal_parameters[0]) + '_MaxSequenceLength=' + str(optimal_parameters[1])

        progressbar_value += 10

        historyFilename = 'history.txt'

        # Need to change this to elastic search Praveen remember
        trainingDataStatistics = trainingDataStats

        trainingDataStats_dumps = json.dumps(trainingDataStats)
        es_conn.update(index=index_name_gloabals, id=id,
                       body={"doc": {"progressbar_value": progressbar_value,
                                     "trainingDataStats_dumps": trainingDataStats_dumps}})
        INDEX_NAME = 'savemodelsupervised'
        query = {"query": {"bool": {"must": [{"match": {"saveProjectName.keyword": selectedProjectName}},
                                             {"match": {"model_data.trainedModelName.keyword": selectedModelName}}]}}}

        res = es_conn.search(index=INDEX_NAME, body=query)
        saveid = res['hits']['hits'][0]['_id']
        es_conn.update(index=INDEX_NAME, id=saveid,
                       body={"doc": {"model_data": {"trainingDataTables": trainingDataStats_dumps}}})

        progressbar_value += 1

        tfidfVectorizer_dumps = pickle.dumps(tfidfVectorizer)
        s3.put_object(Body=tfidfVectorizer_dumps, Bucket=AWS_STORAGE_BUCKET_NAME,
                      Key=key_value + 'tfidf_vect.pkl')
        progressbar_value += 1

        # Save the model, model's CV performance, and CV performance STDEV
        model_dumps = pickle.dumps(trainedModel)
        s3.put_object(Body=model_dumps, Bucket=AWS_STORAGE_BUCKET_NAME,
                      Key=key_value + 'trainedModel.pkl')
        progressbar_value += 1

        performances_dumps = list(trainingDataPerformances)
        es_conn.update(index=INDEX_NAME, id=saveid,
                       body={"doc": {"model_data": {"trainingDataPerformances": performances_dumps}}})
        progressbar_value += 1

        performancesStdev_dumps = list(trainingDataPerformancesStandardDeviation)
        # s3.put_object(Body=performancesStdev_dumps, Bucket=AWS_STORAGE_BUCKET_NAME,
        #               Key=key_value + 'performancesStdev.pkl')
        es_conn.update(index=INDEX_NAME, id=saveid,
                       body={"doc": {
                           "model_data": {"trainingDataPerformancesStandardDeviation": performancesStdev_dumps}}})
        progressbar_value += 1

        # fiveFoldTrainingDatasetTfidf_dumps = pickle.dumps(fiveFoldTrainingDatasetTfidf)
        # s3.put_object(Body=fiveFoldTrainingDatasetTfidf_dumps, Bucket=AWS_STORAGE_BUCKET_NAME,
        #               Key=key_value + 'fiveFoldTrainingDatasetTfidf.pkl')
        # progressbar_value += 1
        #
        # fiveFoldTestingDatasetTfidf_dumps = pickle.dumps(fiveFoldTestingDatasetTfidf)
        # s3.put_object(Body=fiveFoldTestingDatasetTfidf_dumps, Bucket=AWS_STORAGE_BUCKET_NAME,
        #               Key=key_value + 'fiveFoldTestingDatasetTfidf.pkl')
        # progressbar_value += 1
        #
        # fiveFoldTrainingLabels_dumps = pickle.dumps(fiveFoldTrainingLabels)
        # s3.put_object(Body=fiveFoldTrainingLabels_dumps, Bucket=AWS_STORAGE_BUCKET_NAME,
        #               Key=key_value + 'fiveFoldTrainingLabels.pkl')
        # progressbar_value += 1
        #
        # fiveFoldTestingLabels_dumps = pickle.dumps(fiveFoldTestingLabels)
        # s3.put_object(Body=fiveFoldTestingLabels_dumps, Bucket=AWS_STORAGE_BUCKET_NAME,
        #               Key=key_value + 'fiveFoldTestingLabels.pkl')
        # progressbar_value += 1
        #
        # fiveFoldModels_dumps = pickle.dumps(fiveFoldModels)
        # s3.put_object(Body=fiveFoldModels_dumps, Bucket=AWS_STORAGE_BUCKET_NAME,
        #               Key=key_value + 'fiveFoldModels.pkl')
        # progressbar_value += 1
        es_conn.update(index=index_name_gloabals, id=id,
                       body={"doc": {"progressbar_value": progressbar_value}})
        # Save the isotonic or sigmoid calibration model for SVM classifier
        if str_model_name == 'svm' or model == 'ovrsvm':
            if model_isotonic_calibration != None:
                # Save the isotonic calibration model
                model_isotonic_calibration_dumps = pickle.dumps(model_isotonic_calibration)
                s3.put_object(Body=model_isotonic_calibration_dumps, Bucket=AWS_STORAGE_BUCKET_NAME,
                              Key=key_value + 'model_isotonic_calibration.pkl')

                # Remove the sigmoid calibration model, because this model was trained on less than 500 training data examples

                try:
                    response2 = s3.get_object(Bucket=AWS_STORAGE_BUCKET_NAME,
                                              Key=key_value + 'model_sigmoid_calibration.pkl')

                    response2 = s3.delete_object(Bucket=AWS_STORAGE_BUCKET_NAME,
                                                 Key=key_value + 'model_sigmoid_calibration.pkl')

                except botocore.exceptions.ClientError as e:
                    if e.response['Error']['Code'] == "404":
                        pass
                    elif e.response['Error']['Code'] == "NoSuchKey":
                        pass
                    else:
                        pass

            if model_sigmoid_calibration != None:
                # Save the sigmoid calibration model
                model_sigmoid_calibration_dumps = pickle.dumps(model_sigmoid_calibration)
                s3.put_object(Body=model_sigmoid_calibration_dumps, Bucket=AWS_STORAGE_BUCKET_NAME,
                              Key=key_value + 'model_sigmoid_calibration.pkl')

        # pathToSaveTheProjectTrainingData = selectedModelName + 'trainingData/'

        # Save the training and testing data sets for future use
        # copyfile(selectedTrainingFilePath, pathToSaveTheProjectTrainingData + os.path.basename(selectedTrainingFilePath)) # Need to save in s3

        trainingDataPatents = 'training_data_patents.txt' + '_' + str(datetime.now())
        trainingDataJournals = 'training_data_journals.txt' + '_' + str(datetime.now())
        # Append the patent data to patent training file, and journal data to journal training file
        if trainingDataType == 'Patent':
            s3.put_object(Body=training_data, Bucket=AWS_STORAGE_BUCKET_NAME,
                          Key=key_value + trainingDataPatents)
        elif trainingDataType == 'Journal':
            s3.put_object(Body=training_data, Bucket=AWS_STORAGE_BUCKET_NAME,
                          Key=key_value + trainingDataJournals)

        progressbar_value += 1

        append_text_to_history_file = ""

        append_text_to_history_file += '-' * 100 + '\n' + 'username: ' + username + '\n'

        # append_text_to_history_file += "Program run started at " + apollo4.globals.programRunStartTime.strftime(
        #     "%I:%M%p on %B %d, %Y") + "\n"
        # historyFile = open('history.txt', 'a')
        append_text_to_history_file += '-' * 100 + '\n'

        if trainingDataType == 'Patent':
            append_text_to_history_file += 'Supervised learning model ' + trainedModelName + ' was trained on the PATENT training data file: ' + trainingFileName + '.' + '\n'
        elif trainingDataType == 'Journal':
            append_text_to_history_file += 'Supervised learning model ' + trainedModelName + ' was trained on the JOURNAL training data file: ' + trainingFileName + '.' + '\n'

        # write the number of instances and classes for tracking purposes
        numInstancesInTrainingData = 0
        stringToDisplayTrainingDataStats = '{:<40s}{:>20s}{:>20s}'.format('Class', '# Examples', 'Class %') + '\n'

        for entry in trainingDataStats:
            stringToDisplayTrainingDataStats += '{:<40s}{:>20s}{:>20s}'.format(str(entry[0]), str(entry[1]),
                                                                               str(entry[2])) + '\n'
            numInstancesInTrainingData += int(entry[1])

        append_text_to_history_file += 'Total number of documents in the training data: ' + str(
            numInstancesInTrainingData) + '\n'
        append_text_to_history_file += 'Total number of classes in the training data: ' + str(
            len(trainingDataStats)) + '\n'
        append_text_to_history_file += 'The model parameters were optimized for \'' + targetPerformanceMeasure + '\'.' + '\n'
        append_text_to_history_file += '5-fold Cross Validation Performance: ' + '\n'
        append_text_to_history_file += '5-fold Cross Validation Performance: ' + '\n'

        perfMeasuresStr = ['Accuracy:', 'AUC:', 'Precision:', 'Recall:', 'F1:']
        for i in range(len(trainingDataPerformances)):
            stringToWrite = '{:<10s}{:>10.2f}{:>4s}{:>10.2f}{:>1s}'.format(perfMeasuresStr[i],
                                                                           trainingDataPerformances[i] * 100.0,
                                                                           '% +/- ',
                                                                           trainingDataPerformancesStandardDeviation[
                                                                               i] * 100.0, '%')
            append_text_to_history_file += stringToWrite + '\n'

        # append_text_to_history_file += progressbarlabel_text + '.' + '\n'

        try:
            response = s3.get_object(Bucket=AWS_STORAGE_BUCKET_NAME,
                                     Key=key_value + 'history.txt')
            history_file_old_text = response['Body'].read().decode('utf-8')
            append_text_to_history_file = history_file_old_text + append_text_to_history_file

            s3.put_object(Body=append_text_to_history_file, Bucket=AWS_STORAGE_BUCKET_NAME,
                          Key=key_value + 'history.txt')

        except botocore.exceptions.ClientError as e:
            if e.response['Error']['Code'] == "404":
                # There is no existing history file, so create a new history file and write the history into that file in S3.
                s3.put_object(Body=append_text_to_history_file, Bucket=AWS_STORAGE_BUCKET_NAME,
                              Key=key_value + 'history.txt')
            else:
                s3.put_object(Body=append_text_to_history_file, Bucket=AWS_STORAGE_BUCKET_NAME,
                              Key=key_value + 'history.txt')

        progressbar_value += 1

        programRunEndTime = datetime.now()

        timeDifference = relativedelta(programRunEndTime, programRunStartTime)

        programRunStartTimeLabel = "Program run took %d days %d hours %d minutes %d seconds." % (
            timeDifference.days, timeDifference.hours, timeDifference.minutes, timeDifference.seconds)
        progressbarlabel_text = programRunStartTimeLabel

        progress_text = progress_text + "\nIncremental learning finished successfully."

        progressbar_value += 5
        final_progress_value = 200
        es_conn.update(index=index_name_gloabals, id=id,
                       body={"doc": {"progressbar_value": progressbar_value, "progress_text": progress_text,
                                     "progressbarlabel_text": progressbarlabel_text,
                                     "final_progress_value": final_progress_value}})

        return True
    except Exception as e:
        return HttpResponse(
            "Error running the program." + str(e))


@csrf_exempt
def incrementalUnsupervisedLearning(request):
    global response

    try:
        # Assuming that the client-side has already selected the options before running the program.
        # Assuming that the request from the client side will have all the fields necessary for running the program.
        if request.method == 'GET':
            return response
        elif request.method == "POST":
            gui_parameters = request.POST.getlist('inputData')[0]
            gui_parameters = json.loads(gui_parameters);
            finalTrainingData = request.FILES.getlist('trainFile')
            # training_data = request.FILES.getlist('file').read().decode("ISO-8859-1")
            training_data = Process_All_Files(finalTrainingData)
            trainingFileName = gui_parameters['training_file_name']
            trainingDataType = gui_parameters['training_data_type']
            selectedModelName = gui_parameters['trainedModelName']
            selectedProjectName = gui_parameters['saveProjectName']
            username = request.user.username;
            current_tab = 6

            progressbar_maximum = 200
            progressbar_value = 0
            progress_text = ''

            progress_text = progress_text + "Starting incremental learning process...\n"

            # Set the text in progressbarlabel
            programRunStartTime = datetime.now()
            programRunStartTimeLabel = 'Progress: Program run started at ' + programRunStartTime.strftime(
                "%I:%M%p on %B %d, %Y") + ' (UTC time). '

            progressbarlabel_text = programRunStartTimeLabel

            progress_text = progress_text + '-' * 75 + '\n' + "Program run started at " + programRunStartTime.strftime(
                "%I:%M%p on %B %d, %Y") + " (UTC time).\n" + \
                            '-' * 75 + '\n' + "Starting incremental learning process..."

            # The path to the model depends on the index of selected project and index of selected model
            numberOfClusters = int(re.search('NumClusters=(.*)_TopWords', selectedModelName).group(1))
            number_of_top_words = int(selectedModelName.split("TopWords=")[1].replace('/', ''))
            progressbar_maximum = 100 + 10 * numberOfClusters
            progressbar_value = 10
            index_name_gloabals = 'apolloglobalsunsupervised'
            query = {"query": {"bool": {"must": {"match": {"username.keyword": username}}}}}
            res = es_conn.search(index=index_name_gloabals, body=query)
            id = res['hits']['hits'][0]['_id']
            es_conn.update(index=index_name_gloabals, id=id,
                           body={"doc": {"progressbar_maximum": progressbar_maximum, "current_tab": current_tab,
                                         "progressbar_value": progressbar_value,
                                         "testingFileName": trainingFileName, "trainedModelName": selectedModelName,
                                         "testingDataType": trainingDataType, "progress_text": progress_text
                               , "progressbarlabel_text": progressbarlabel_text, "numberOfClusters": numberOfClusters,
                                         "number_of_top_words": number_of_top_words}})
            historyFilename = 'history.txt'
            s3 = boto3.client('s3')
            key_value = 'classificationprojects/'
            key_value += selectedProjectName + '/unsupervised/'
            key_value += selectedModelName + '/'

            # The code for loading and pre-processing the data is different for patent and journal data
            if trainingDataType == 'Patent':
                file_sample_open = training_data
                file_sample_open = file_sample_open.split('\n')  # split by new line
                file_sample_open = list(filter(None, file_sample_open))  # delete empty lines

                # Now, the first line is header, so remove the first line
                file_sample_open = file_sample_open[1:]
                progress_text = progress_text + "\nFound " + str(
                    len(file_sample_open)) + " documents! \nPreprocessing documents...\n"

                # Set value of progressbar to 10 once the training dataset is loaded
                progressbar_value += 10
                es_conn.update(index=index_name_gloabals, id=id,
                               body={"doc": {"progressbar_value": progressbar_value, "progress_text": progress_text}})

                # Build the stop words
                stops = stopwords

                aux_stops = './static/AuxStops.txt'
                aux_stops = open(aux_stops, 'r').read()
                aux_stops = re.sub("[^a-zA-Z ]", "   ", aux_stops)  # remove non-alphanumeric
                aux_stops = " ".join(aux_stops.split())  # split by any whitespace and rejoin w/ space
                aux_stops = aux_stops.split(' ')
                aux_stops = list(filter(None, aux_stops))

                # append auxiliary stops
                stops = stops + aux_stops

                user_defined_stopwords = []
                stops = stops + user_defined_stopwords

                # Bulid stopterm dictionary
                stopterms = {}
                for stop in stops:
                    if stop in stopterms:
                        stopterms[stop] += 1
                    else:
                        stopterms[stop] = 1

                # Preprocess the sample file
                # Also, we need to check the original data file for any duplicates in the new data

                # Load the latest training data for this model and use the application number column to deduplicate the documents from previous patents data

                try:
                    response = s3.get_object(Bucket=AWS_STORAGE_BUCKET_NAME,
                                             Key=key_value + 'training_data_patents.txt')
                    oldPatentsFile = response['Body'].read()
                    # file_old_training_data_open = codecs.open(oldPatentsFile).read()
                    file_old_training_data_open = oldPatentsFile.decode('utf-8')
                    file_old_training_data_open = file_old_training_data_open.split('\n')  # split by new line
                    file_old_training_data_open = list(filter(None, file_old_training_data_open))  # delete empty lines
                    # Now, the first line is header, so remove the first line
                    file_old_training_data_open = file_old_training_data_open[1:]
                    existing_application_numbers = [doc.split('\t')[4].upper() for doc in file_old_training_data_open]

                except botocore.exceptions.ClientError as e:
                    if e.response['Error']['Code'] == "404":
                        # There is no existing patent training data, so no need to deduplicate based on previous data.
                        existing_application_numbers = []
                    elif e.response['Error']['Code'] == "NoSuchKey":
                        existing_application_numbers = []
                        pass
                    else:
                        existing_application_numbers = []
                        pass

                progressbar_value += 10
                es_conn.update(index=index_name_gloabals, id=id,
                               body={"doc": {"progressbar_value": progressbar_value}})

                (file_test_proc, file_sample_stem,
                 unlabeled_data_stop_words_removed) = preprocess_collection_incremental_learning(username,
                                                                                                 file_sample_open,
                                                                                                 stopterms,
                                                                                                 existing_application_numbers,
                                                                                                 True, progress_text)
                file_sample = list(filter(None, file_sample_stem))

                progressbar_value += 20
                es_conn.update(index=index_name_gloabals, id=id,
                               body={"doc": {"progressbar_value": progressbar_value}})

                if len(file_sample) < 1:
                    progressbar_value = 100 + numberOfClusters
                    # Simply display the error message and exit the function.
                    final_progress_value = 200
                    errorString = 'The additional data file does not contain any new Patents for training the unsupervised learning model. \nCannot perform incremental learning in this case.5436'
                    es_conn.update(index=index_name_gloabals, id=id,
                                   body={"doc": {"progressbar_value": progressbar_value,
                                                 "final_progress_value": final_progress_value,
                                                 "errorString": errorString}})
                    return JsonResponse({'finalResponse': 'Error'})
                title_samples = [doc.split('\t')[1] for doc in file_sample]
                abstract_samples = [doc.split('\t')[2] for doc in file_sample]
                claim_samples = [doc.split('\t')[3] for doc in file_sample]

                unlabeled_data = [' '.join(doc) for doc in zip(title_samples, abstract_samples, claim_samples)]

                # End patent training data

            elif trainingDataType == 'Journal':
                file_sample_open = training_data
                file_sample_open = file_sample_open.split('\n')  # split by new line
                file_sample_open = list(filter(None, file_sample_open))  # delete empty lines

                # Now, the first line is header, so remove the first line
                file_sample_open = file_sample_open[1:]

                progress_text = progress_text + "\nFound " + str(
                    len(file_sample_open)) + " documents in the additional data file! \nPreprocessing documents...\n"

                # Set value of progressbar to 10 once the training dataset is loaded
                progressbar_value += 10
                es_conn.update(index=index_name_gloabals, id=id,
                               body={"doc": {"progressbar_value": progressbar_value,
                                             "progress_text": progress_text}})
                # Load the latest training data for this model and use the title and abstract columns to deduplicate the documents from previous journals data

                try:
                    response = s3.get_object(Bucket=AWS_STORAGE_BUCKET_NAME,
                                             Key=key_value + 'training_data_journals.txt')
                    oldJournalsFile = response['Body'].read()

                    # The training data already exists, and we need to append the new data to this data as well.
                    file_old_training_data_open = oldJournalsFile.decode('utf-8')
                    file_old_training_data_open = file_old_training_data_open.split('\n')  # split by new line
                    file_old_training_data_open = list(filter(None, file_old_training_data_open))  # delete empty lines

                    # Now, the first line is header, so remove the first line
                    file_old_training_data_open = file_old_training_data_open[1:]
                    existing_title_and_abstracts = ['\t'.join([doc.split('\t')[1], doc.split('\t')[2]]) for doc in
                                                    file_old_training_data_open]


                except botocore.exceptions.ClientError as e:
                    if e.response['Error']['Code'] == "404":
                        # There is no existing journal training data, so no need to deduplicate based on previous data.
                        existing_title_and_abstracts = []
                    elif e.response['Error']['Code'] == "NoSuchKey":
                        existing_title_and_abstracts = []
                        pass
                    else:
                        existing_title_and_abstracts = []
                        pass

                progressbar_value += 10
                es_conn.update(index=index_name_gloabals, id=id,
                               body={"doc": {"progressbar_value": progressbar_value}})
                # Remove the duplicated documents based on "title"
                file_data_open = dedup_collection_journal_incremental_learning(file_sample_open,
                                                                               existing_title_and_abstracts,
                                                                               1, 2)

                progressbar_value += 10
                es_conn.update(index=index_name_gloabals, id=id,
                               body={"doc": {"progressbar_value": progressbar_value}})

                if len(file_data_open) < 1:
                    progressbar_value = 100 + numberOfClusters
                    # Simply display the error message and exit the function.
                    final_progress_value = 200
                    errorString = 'The additional data file does not contain any new journals for training the unsupervised learning model. \nCannot perform incremental learning in this case.5503'
                    es_conn.update(index=index_name_gloabals, id=id,
                                   body={"doc": {"progressbar_value": progressbar_value,
                                                 "final_progress_value": final_progress_value,
                                                 "errorString": errorString}})
                    return JsonResponse({'finalResponse': 'Error'})
                # Preprocessing for scoupus data

                file_test_proc = preprocess_collection_journal(file_data_open)

                progressbar_value += 10
                es_conn.update(index=index_name_gloabals, id=id,
                               body={"doc": {"progressbar_value": progressbar_value}})

                user_defined_stopwords = []

                unlabeled_data = []
                unlabeled_data_stop_words_removed = []

                for doc in file_data_open:
                    stop_and_stem_document_title, stop_document_title = stop_and_stem_journal_2(doc.split('\t')[1],
                                                                                                user_defined_stopwords)
                    stop_and_stem_document_abstract, stop_document_abstract = stop_and_stem_journal_2(
                        doc.split('\t')[2],
                        user_defined_stopwords)
                    unlabeled_data.append(' '.join([stop_and_stem_document_title, stop_and_stem_document_abstract]))
                    unlabeled_data_stop_words_removed.append(' '.join([stop_document_title, stop_document_abstract]))

            progressbar_value += 10

            ## Set directory for the output folder
            # output_folder = os.path.dirname(selectedTrainingFilePath)

            progress_text = progress_text + "Starting unsupervised learning process..." + "\nFound " + str(
                len(unlabeled_data)) + " new documents in the additional data file!\n"
            es_conn.update(index=index_name_gloabals, id=id,
                           body={"doc": {"progressbar_value": progressbar_value, "progress_text": progress_text}})

            # The code for patent and journal testing data is different because it required different preprocessing

            user_defined_stopwords = []

            # Load the tfidf vectorizer that was previously saved
            response = s3.get_object(Bucket=AWS_STORAGE_BUCKET_NAME,
                                     Key=key_value + 'tfidf_vect.pkl')
            tfidf_vect_file = response['Body'].read()

            tfidf_vect = pickle.loads(tfidf_vect_file)
            unlabled_data_tfidf = tfidf_vect.transform(unlabeled_data)

            # Model and model parameters
            model = None
            number_of_clusters = numberOfClusters

            progress_text = progress_text + "\nPerforming clustering and topic extraction on the data..."
            es_conn.update(index=index_name_gloabals, id=id,
                           body={"doc": {"progress_text": progress_text}})

            output_filename = re.sub('.txt', '_Results_Topic_Modeling.txt', trainingFileName)

            # 1. Load the model

            response = s3.get_object(Bucket=AWS_STORAGE_BUCKET_NAME,
                                     Key=key_value + 'trainedModel.pkl')
            model_file = response['Body'].read()

            model = pickle.loads(model_file)

            model.partial_fit(unlabled_data_tfidf)
            modelDumps = pickle.dumps(model)

            # simply copy the file to the trainingData folder with the timestamp information
            # training_data = pickle.dumps(training_data)
            # s3.put_object(Body=training_data, Bucket=AWS_STORAGE_BUCKET_NAME,
            #                           Key=key_value + '/trainingData/' + os.path.basename(
            #                               trainingFileName) + '_' + datetime.now().strftime('%Y-%m-%d %H-%M-%S') +
            #                               os.path.splitext(trainingFileName)[1])

            # Append the patent data to patent training file, and journal data to journal training file
            if trainingDataType == 'Patent':
                s3.put_object(Body=training_data, Bucket=AWS_STORAGE_BUCKET_NAME,
                              Key=key_value + 'training_data_patents.txt')

            elif trainingDataType == 'Journal':
                s3.put_object(Body=training_data, Bucket=AWS_STORAGE_BUCKET_NAME,
                              Key=key_value + 'training_data_journals.txt')

            # Write the newly updated model back to S3; assuming that S3 overwrites the old file with the new one, and no settings have been changed for S3.
            s3.put_object(Body=modelDumps, Bucket=AWS_STORAGE_BUCKET_NAME,
                          Key=key_value + 'trainedModel.pkl')

            progressbar_value += 20
            es_conn.update(index=index_name_gloabals, id=id,
                           body={"doc": {"progressbar_value": progressbar_value}})

            clusters = model.labels_.tolist()

            test_patent_others = {'content': unlabeled_data, 'file_others': file_test_proc,
                                  'content_stop_words_removed': unlabeled_data_stop_words_removed, 'cluster': clusters}

            # 1. Load the previous data frame to run topic modeling again on each set of document clusters
            response = s3.get_object(Bucket=AWS_STORAGE_BUCKET_NAME,
                                     Key=key_value + 'topicModelingFrame.pkl')

            old_frame_file = response['Body'].read()
            oldFrame = pickle.loads(old_frame_file)

            frame = pd.DataFrame(test_patent_others, index=[clusters],
                                 columns=['content', 'file_others', 'content_stop_words_removed', 'cluster'])
            frame = frame.append(oldFrame)

            clusterTopicsAndCounts = []

            clustering_successful = False
            # output lda other topics
            fout_others = ''
            for no in range(numberOfClusters):
                try:
                    # sometimes, there is no document in the group, so handle that case with try and except
                    patent_group = frame.groupby(frame['cluster']).get_group(no)
                except:
                    # continue, because there is no document in this cluster. Move on to the topic modeling for next cluster
                    continue

                patent_tac = patent_group.ix[:, 0].tolist()
                patent_org = patent_group.ix[:, 1].tolist()

                lda_tf_vect = TfidfVectorizer(max_df=0.8, min_df=1,
                                              max_features=200000,
                                              ngram_range=(1, 5),
                                              use_idf=True,
                                              stop_words='english')
                tf = None
                try:
                    tf = lda_tf_vect.fit_transform(patent_tac)

                except Exception as e:
                    lda_tf_vect = TfidfVectorizer(max_df=1.0, min_df=1,
                                                  max_features=200000,
                                                  ngram_range=(1, 5),
                                                  use_idf=True,
                                                  stop_words='english')
                    tf = lda_tf_vect.fit_transform(patent_tac)

                # LDA Model
                lda = LatentDirichletAllocation(n_components=1, max_iter=20,
                                                learning_method='online',
                                                learning_offset=50,
                                                random_state=0).fit(tf)

                lda_feature_names = lda_tf_vect.get_feature_names()

                lda_topics = get_topic_list(lda, lda_feature_names, number_of_top_words)

                clusterTopicsAndCounts.append([len(patent_tac), lda_topics[0]])

                doc_topic = lda.transform(tf)
                doc_topic_index = doc_topic.argmax(axis=1)

                for doc, doc_topic_i in zip(patent_org, doc_topic_index):
                    fout_others += '\t'.join(
                        [doc.strip('\r').strip('\n'), lda_topics[doc_topic_i].strip('\r').strip('\n')]) + '\n'

                clustering_successful = True

                progressbar_value += 10
                es_conn.update(index=index_name_gloabals, id=id,
                               body={"doc": {"progressbar_value": progressbar_value}})

            if clustering_successful == True:
                progress_text = progress_text + '\nTopic extraction and clustering completed.'
                es_conn.update(index=index_name_gloabals, id=id,
                               body={"doc": {"progress_text": progress_text}})

                # Load the topic modeling results in the treeview
                # Compute the columns: (i) # instances, (ii) Topics extracted
                clusterTopicsAndCounts = clusterTopicsAndCounts
                dumps_clusterTopicsAndCounts = json.dumps(clusterTopicsAndCounts)

            progress_text = progress_text + "\nPlease download the " + output_filename + " file and check all the results in the file."
            es_conn.update(index=index_name_gloabals, id=id,
                           body={"doc": {"progress_text": progress_text}})
            # historyFile = open(historyFilename, append_write)
            append_text_to_history_file = "\n"
            append_text_to_history_file += '-' * 100 + '\n'
            append_text_to_history_file += 'username: ' + username
            append_text_to_history_file += "Program run started at " + programRunStartTime.strftime(
                "%I:%M%p on %B %d, %Y") + " (UTC time).\n" + '-' * 100 + '\n'

            ## Update the history based on whether the updated model is supervised or unsupervised model.

            if trainingDataType == 'Patent':
                append_text_to_history_file += 'Unsupervised learning model ' + selectedModelName + ' was incrementally updated using the PATENT training data file: ' + trainingFileName + '.' + '\n'
            elif trainingDataType == 'Journal':
                append_text_to_history_file += 'Unsupervised learning model ' + selectedModelName + ' was incrementally updated using the JOURNAL training data file: ' + trainingFileName + '.' + '\n'

            INDEX_NAME = 'savemodelunsupervised'
            query = {"query": {"bool": {"must": [{"match": {"saveProjectName.keyword": selectedProjectName}},
                                                 {"match": {
                                                     "model_data.trainedModelName.keyword": selectedModelName}}]}}}

            res = es_conn.search(index=INDEX_NAME, body=query)
            saveid = res['hits']['hits'][0]['_id']
            es_conn.update(index=INDEX_NAME, id=saveid,
                           body={"doc": {"model_data": {"clusterTopicsAndCounts": dumps_clusterTopicsAndCounts}}})

            programRunEndTime = datetime.now()

            timeDifference = relativedelta(programRunEndTime, programRunStartTime)

            programRunStartTimeLabel = "Program run took %d days %d hours %d minutes %d seconds." % (
                timeDifference.days, timeDifference.hours, timeDifference.minutes, timeDifference.seconds)
            progressbarlabel_text = programRunStartTimeLabel

            modelSavingTimeLabel = "Incrementally updating the model took %d days %d hours %d minutes %d seconds." % (
                timeDifference.days, timeDifference.hours, timeDifference.minutes, timeDifference.seconds)

            es_conn.update(index=index_name_gloabals, id=id,
                           body={"doc": {"progressbarlabel_text": progressbarlabel_text}})
            append_text_to_history_file += modelSavingTimeLabel + '.' + '\n' + '*' * 95 + '\n'

            # Need to add history file in all tabs and uncomment below code
            try:
                response = s3.get_object(Bucket=AWS_STORAGE_BUCKET_NAME,
                                         Key=key_value + 'history.txt')
                history_file_old_text = response['Body'].read().decode('utf-8')
                append_text_to_history_file = history_file_old_text + append_text_to_history_file

                s3.put_object(Body=append_text_to_history_file, Bucket=AWS_STORAGE_BUCKET_NAME,
                              Key=key_value + 'history.txt')

            except botocore.exceptions.ClientError as e:
                if e.response['Error']['Code'] == "404":
                    # There is no existing history file, so create a new history file and write the history into that file in S3.
                    s3.put_object(Body=append_text_to_history_file, Bucket=AWS_STORAGE_BUCKET_NAME,
                                  Key=key_value + 'history.txt')
                elif e.response['Error']['Code'] == "NoSuchKey":
                    s3.put_object(Body=append_text_to_history_file, Bucket=AWS_STORAGE_BUCKET_NAME,
                                  Key=key_value + 'history.txt')
                    pass
                else:
                    s3.put_object(Body=append_text_to_history_file, Bucket=AWS_STORAGE_BUCKET_NAME,
                                  Key=key_value + 'history.txt')
                    pass

            progressbar_value += 20
            final_progress_value = 200
            es_conn.update(index=index_name_gloabals, id=id,
                           body={"doc": {"progressbar_value": progressbar_value,
                                         "final_progress_value": final_progress_value}})

            response = HttpResponse(content=fout_others,
                                    content_type='text/plain')
            response['Content-Disposition'] = 'attachment; filename=' + output_filename
        return HttpResponse('success');
    except Exception as e:
        final_progress_value = 200
        errorString = errorString
        es_conn.update(index=index_name_gloabals, id=id,
                       body={"doc": {"errorString": errorString,
                                     "final_progress_value": final_progress_value}})
        return JsonResponse({'finalResponse': 'Error'})


@csrf_exempt
def getUserName(request):
    try:
        if request.method == "POST":
            userName = request.user.username;
        return HttpResponse(userName)
    except Exception as e:
        return HttpResponse(
            "Error running the program." + str(e))


@csrf_exempt
def saveAndContinue(request):
    HOST_URLS = ["https://search-dataannotation-unp43zc5slowdwllbqukty73wy.us-east-1.es.amazonaws.com"]

    #  In production or deploying to aws uncomment this line and comment local server settings.
    es_conn = Elasticsearch(HOST_URLS, timeout=30)
    AWS_STORAGE_BUCKET_NAME = 'dataannotate'
    s3 = boto3.client('s3')

    try:
        if request.method == "POST":
            userName = request.user.username;
            save_doc_Name = request.body.decode('utf-8');
            saveandcontinue = json.loads(save_doc_Name);
            categoryName = saveandcontinue['categoryName']
            timeTaken = saveandcontinue['timeTaken']
            keywords = saveandcontinue['keywords']
            index = saveandcontinue['index']
            headers = saveandcontinue['headers']

            key_value = 'catogorizedfiles/'
            key_value += userName + '/'

            index_name = 'annotatefile'
            query = {"query": {"bool": {"must": {"match": {"username.keyword": userName}}}}}

            if es_conn.indices.exists(index_name):
                res = es_conn.search(index=index_name, body=query)
            else:
                save_response = es_conn.indices.create(index=index_name, ignore=400)
                res = es_conn.search(index=index_name, body=query)

            fileName = res['hits']['hits'][0]['_source']['filename']
            training_data = res['hits']['hits'][0]['_source']['training_data']

            file_sample_open = training_data[1:]
            file_sample = list(filter(None, file_sample_open))
            identificationNumber = [doc.split('\t')[1].strip('\r').strip('\n') for doc in file_sample]
            identificationNumber = identificationNumber[0]
            file_sample_current_document = file_sample[0]
            task = [doc.split('\t')[0].strip('\r').strip('\n') for doc in file_sample]
            task = task[0]

            saveAndContinueData = {
                'identificationNumber': identificationNumber,
                'username': userName,
                'fileName': fileName,
                'categoryName': categoryName,
                'timespent': str(timeTaken) + ' sec',
                'keywords': keywords,
                'index': index,
                'headers': headers,
                'task': task,
            }
            INDEX_NAME = 'dataannotateresult'
            # save_response = es_conn.indices.create(index=INDEX_NAME, ignore=400)
            if es_conn.indices.exists(INDEX_NAME):
                savingResponse = es_conn.create(index=INDEX_NAME, doc_type=TYPE_NAME_USER, body=saveAndContinueData,
                                                id=uuid.uuid4())
            else:
                save_response = es_conn.indices.create(index=INDEX_NAME, ignore=400)
                savingResponse = es_conn.create(index=INDEX_NAME, doc_type=TYPE_NAME_USER, body=saveAndContinueData,
                                                id=uuid.uuid4())

            if categoryName != 'skipButton':
                labeleddocument = file_sample_current_document.strip('\r').strip('\n') + '\t' + categoryName + '\n'
                try:
                    response2 = s3.get_object(Bucket=AWS_STORAGE_BUCKET_NAME,
                                              Key=key_value + fileName)
                    oldlabeleddocument = response2['Body'].read().decode('utf-8')
                    labeleddocument = oldlabeleddocument + labeleddocument

                    s3.put_object(Body=labeleddocument, Bucket=AWS_STORAGE_BUCKET_NAME,
                                  Key=key_value + fileName)

                except botocore.exceptions.ClientError as e:
                    if e.response['Error']['Code'] == "404":
                        # There is no existing history file, so create a new history file and write the history into that file in S3.
                        header = training_data[0].strip('\r').strip('\n') + '\t' + 'Category'
                        labeleddocument = header + '\n' + labeleddocument
                        s3.put_object(Body=labeleddocument, Bucket=AWS_STORAGE_BUCKET_NAME,
                                      Key=key_value + fileName)
                    elif e.response['Error']['Code'] == "NoSuchKey":
                        header = training_data[0].strip('\r').strip('\n') + '\t' + 'Category'
                        labeleddocument = header + '\n' + labeleddocument
                        s3.put_object(Body=labeleddocument, Bucket=AWS_STORAGE_BUCKET_NAME,
                                      Key=key_value + fileName)
                        pass
                    else:
                        pass
        return HttpResponse('success')
    except Exception as e:
        return HttpResponse(
            "Error running the program.")


@csrf_exempt
def customFileUpload(request):
    # global training_data
    index_name = 'annotatefile'
    HOST_URLS = ["https://search-dataannotation-unp43zc5slowdwllbqukty73wy.us-east-1.es.amazonaws.com"]

    #  In production or deploying to aws uncomment this line and comment local server settings.
    es_conn = Elasticsearch(HOST_URLS, timeout=30)
    AWS_STORAGE_BUCKET_NAME = 'dataannotate'
    try:
        if request.method == 'POST':
            shuffledArray = request.POST.getlist('shuffledArray')[0].split('#@@#')

            userName = request.user.username;
            # training_data = open(request.FILES.get('file').temporary_file_path(), 'r').read()
            # es_conn.indices.create(index=index_name)
            # form = FileUploadForm(request.POST, request.FILES)
            training_data = request.FILES.get('file').read().decode("ISO-8859-1")
            fileName = request.FILES['file'].name
            fileName = str(
                fileName.split('.')[0] + '_' + datetime.now().strftime("%Y%m%d-%H%M%S") + '.' + fileName.split('.')[1])

            # res = es_conn.update(index=index_name,body={"doc": {"match_all": {}}})
            headerTemplate = ['meta data', 'identification number', 'task', 'title', 'abstract', 'claims',
                              'application number', 'application date', 'assignee', 'current assignee', 'upc',
                              'appln. no.', 'appl. date', 'authors', 'affiliation', 'year', 'category']

            file_sample_open = training_data
            file_sample_open = file_sample_open.split('\n')  # split by new line
            file_sample_open = list(filter(None, file_sample_open))  # delete empty lines
            file_sample_headers = file_sample_open[0].strip('\r').split('\t')
            header_lower = [x.lower() for x in file_sample_headers]

            if set(header_lower).issubset(set(headerTemplate)):
                task_file_open = shuffledArray
                # task_file_open = task_file_open.split('\n')  # split by new line
                # task_file_open = list(filter(None, task_file_open))
                task_file_open = task_file_open[1:]
                task = [doc.split('\t')[0].strip('\r').strip('\n') for doc in task_file_open]
                task = task[0]
                # Now, the first line is header, so remove the first line
                query = {"query": {"bool": {"must": {"match": {"username.keyword": userName}}}}}
                # save_response = es_conn.indices.create(index=index_name, ignore=400)
                if es_conn.indices.exists(index_name):
                    es_conn.delete_by_query(index=index_name, body=query)
                else:
                    save_response = es_conn.indices.create(index=index_name, ignore=400)
                    es_conn.delete_by_query(index=index_name, body=query)

                if len(shuffledArray) > 0:
                    datafile = {
                        'username': userName,
                        'training_data': shuffledArray,
                        'filename': fileName,
                        'task': task
                    }

                    es_conn.create(index=index_name, doc_type='_doc', body=datafile, id=uuid.uuid4())

                    key_value = 'inputfiles/'
                    key_value += userName + '/'

                    s3 = boto3.client('s3')
                    file_sample_open = json.dumps(file_sample_open, ensure_ascii=False)
                    s3.put_object(Body=file_sample_open, Bucket=AWS_STORAGE_BUCKET_NAME,
                                  Key=key_value + fileName)
                dataResponse = 'sucess'
            else:
                dataResponse = 'errorfile'

            return HttpResponse(dataResponse)
    except Exception as e:
        return HttpResponse(
            "Error running the program.")


@csrf_exempt
def removeData(request):
    HOST_URLS = ["https://search-dataannotation-unp43zc5slowdwllbqukty73wy.us-east-1.es.amazonaws.com"]

    #  In production or deploying to aws uncomment this line and comment local server settings.
    es_conn = Elasticsearch(HOST_URLS, timeout=30)
    try:
        if request.method == "POST":
            userName = request.user.username;
            index_name = 'annotatefile'
            query = {"query": {"bool": {"must": {"match": {"username.keyword": userName}}}}}
            if es_conn.indices.exists(index_name):
                res = es_conn.search(index=index_name, body=query)
            else:
                save_response = es_conn.indices.create(index=index_name, ignore=400)
                res = es_conn.search(index=index_name, body=query)
            id = res['hits']['hits'][0]['_id']
            training_data = res['hits']['hits'][0]['_source']['training_data']

            training_data.pop(1)
            if len(training_data) == 1:
                es_conn.delete_by_query(index=index_name, body=query)
            # save_response = es_conn.indices.create(index=INDEX_NAME, ignore=400)
            else:
                es_conn.update(index=index_name, id=id,
                               body={"doc": {"training_data": training_data}})
        return HttpResponse('success')
    except Exception as e:
        return HttpResponse(
            "Error running the program.")


@csrf_exempt
def resetDocument(request):
    # global training_data
    index_name = 'annotatefile'
    HOST_URLS = ["https://search-dataannotation-unp43zc5slowdwllbqukty73wy.us-east-1.es.amazonaws.com"]

    #  In production or deploying to aws uncomment this line and comment local server settings.
    es_conn = Elasticsearch(HOST_URLS, timeout=30)
    try:
        if request.method == 'POST':
            userName = request.user.username;

            query = {"query": {"bool": {"must": {"match": {"username.keyword": userName}}}}}

            if es_conn.indices.exists(index_name):
                es_conn.delete_by_query(index=index_name, body=query)
            else:
                save_response = es_conn.indices.create(index=index_name, ignore=400)
                es_conn.delete_by_query(index=index_name, body=query)

            return HttpResponse('sucess')
    except Exception as e:
        return HttpResponse(
            "Error running the program.")


@csrf_exempt
def dataWhenReload(request):
    HOST_URLS = ["https://search-dataannotation-unp43zc5slowdwllbqukty73wy.us-east-1.es.amazonaws.com"]

    #  In production or deploying to aws uncomment this line and comment local server settings.
    es_conn = Elasticsearch(HOST_URLS, timeout=30)
    try:
        if request.method == "POST":
            userName = request.user.username;
            index_name = 'annotatefile'
            query = {"query": {"bool": {"must": {"match": {"username.keyword": userName}}}}}
            if es_conn.indices.exists(index_name):
                res = es_conn.search(index=index_name, body=query)
            else:
                save_response = es_conn.indices.create(index=index_name, ignore=400)
                res = es_conn.search(index=index_name, body=query)
            training_data = res['hits']['hits'][0]['_source']['training_data']

        return JsonResponse({"training_data": training_data})
    except Exception as e:
        return HttpResponse(
            "Error running the program.")


@csrf_exempt
def categoryName(request):
    HOST_URLS = ["https://search-dataannotation-unp43zc5slowdwllbqukty73wy.us-east-1.es.amazonaws.com"]

    #  In production or deploying to aws uncomment this line and comment local server settings.
    es_conn = Elasticsearch(HOST_URLS, timeout=30)
    try:
        if request.method == "POST":
            userName = request.user.username;
            index_name = 'annotatefile'
            query = {"_source": ["task"], "query": {"bool": {"must": {"match": {"username.keyword": userName}}}}}
            if es_conn.indices.exists(index_name):
                res = es_conn.search(index=index_name, body=query)
            else:
                save_response = es_conn.indices.create(index=index_name, ignore=400)
                res = es_conn.search(index=index_name, body=query)

            task = res['hits']['hits'][0]['_source']['task']

            index_name = 'dataannotateresult'
            query = {"query": {"bool": {"must": {"match": {"task": task}}}}, "size": 10000}
            if es_conn.indices.exists(index_name):
                res = es_conn.search(index=index_name, body=query)
            else:
                save_response = es_conn.indices.create(index=index_name, ignore=400)
                res = es_conn.search(index=index_name, body=query)

        return JsonResponse(res)
    except Exception as e:
        return HttpResponse(
            "Error running the program.")


@csrf_exempt
def downloads3file(request):
    #  In production or deploying to aws uncomment this line and comment local server settings.
    HOST_URLS = ["https://admin:Samsung1!@search-apolloannotate-llycnkz2qhhenqpe4gj4tl2thy.us-east-1.es.amazonaws.com"]
    es_conn = Elasticsearch(HOST_URLS, timeout=30)
    AWS_STORAGE_BUCKET_NAME = 'dataannotate'
    s3 = boto3.client('s3')
    response1 = ''

    try:
        if request.method == "GET":
            userName = request.user.username;
            index_name = 'annotatefile'
            query = {"_source": ["filename"], "query": {"bool": {"must": {"match": {"username.keyword": userName}}}}}
            res = es_conn.search(index=index_name, body=query)
            filename = res['hits']['hits'][0]['_source']['filename']
            key_value = 'catogorizedfiles/'
            key_value += userName + '/'
            response = s3.get_object(Bucket=AWS_STORAGE_BUCKET_NAME,
                                     Key=key_value + filename)
            filecontent = response['Body'].read().decode('utf-8')

            response1 = HttpResponse(filecontent, content_type='text/plain')
            response1['Content-Disposition'] = 'attachment; filename=' + filename
            # workbook.save(response1);
        return response1
    except Exception as e:
        return HttpResponse(
            "Error running the program.")